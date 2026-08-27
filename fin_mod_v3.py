import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import io
import subprocess
import sys

# Автоматическая установка openpyxl в случае отсутствия (полезно для Streamlit Cloud при нестандартном имени файла зависимостей)
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Установка конфигурации страницы
st.set_page_config(
    page_title="CENTURY 21 Financial Model MVP v2",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Пользовательский CSS-стилинг в корпоративных цветах Century 21 (Золотой #C2A24D, Черный #111111, Серый)
st.markdown("""
<style>
    .reportview-container {
        background-color: #111111;
        color: #ffffff;
    }
    .stMetric {
        background-color: #1a1a1a;
        padding: 15px;
        border-radius: 8px;
        border-left: 5px solid #C2A24D;
        box-shadow: 2px 2px 5px rgba(0,0,0,0.3);
    }
    .stMetric label {
        color: #cccccc !important;
        font-weight: bold;
    }
    .stMetric div[data-testid="stMetricValue"] {
        color: #C2A24D !important;
        font-size: 24px;
    }
    .main-title {
        color: #C2A24D;
        font-size: 32px;
        font-weight: bold;
        text-align: center;
        margin-bottom: 5px;
        text-transform: uppercase;
        letter-spacing: 2px;
    }
    .sub-title {
        color: #ffffff;
        font-size: 16px;
        text-align: center;
        margin-bottom: 30px;
        font-style: italic;
    }
    div.stButton > button:first-child {
        background-color: #C2A24D;
        color: #111111;
        font-weight: bold;
        border: none;
        border-radius: 5px;
        transition: all 0.3s ease;
        width: 100%;
        padding: 10px;
    }
    div.stButton > button:first-child:hover {
        background-color: #ffffff;
        color: #111111;
        box-shadow: 0 4px 15px rgba(194, 162, 77, 0.4);
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 10px;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #1a1a1a;
        border: 1px solid #333333;
        border-radius: 5px 5px 0px 0px;
        padding: 10px 20px;
        color: #ffffff;
    }
    .stTabs [aria-selected="true"] {
        background-color: #C2A24D !important;
        color: #111111 !important;
        font-weight: bold;
    }
</style>
""", unsafe_style_with_html=True)

# ----------------------------------------------------
# ОПРЕДЕЛЕНИЕ БАЗОВЫХ ДАННЫХ (На основе источников C21)
# ----------------------------------------------------

# Список месяцев
months = [f"Месяц {i}" for i in range(1, 25)]

# Сценарий "Гипотеза 1" - Базовые показатели
hyp1_baseline = {
    'leads': [60, 120, 180, 240, 300, 360, 420, 480, 540, 600, 660, 720,
              840, 900, 960, 1020, 1080, 1140, 1200, 1260, 1320, 1380, 1440, 1500],
    'deals_rent': [3, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1, 1,
                   3, 3, 4, 4, 4, 4, 5, 5, 6, 6, 6, 6],
    'deals_suburban': [0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1, 1,
                       2, 2, 2, 2, 3, 3, 4, 4, 5, 5, 5, 5],
    'revenue': [0, 0, 0, 1520000, 957400, 1096100, 1811500, 2619650, 2322329, 2559567, 3376806, 3614044,
                3974182, 4247600, 4551568, 5070055, 5874023, 6357991, 6941959, 7245928, 8049896, 8313864, 8837832, 8881800],
    'agent_comm': [0, 0, 0, 612100, 398312, 460152, 892040, 1274170, 1109098, 1215943, 1628083, 1734144,
                   2053406, 2216022, 2372104, 2625039, 3002856, 3369888, 3678751, 3834367, 4214813, 4349624, 4587930, 4640042],
    'salaries': [214935, 214935, 214935, 214935, 214935, 144935, 144935, 144935, 144935, 144935, 144935, 144935,
                 957423, 957814, 960140, 1118661, 1120987, 1453313, 1598638, 1600964, 1753290, 1755616, 1757942, 1760267],
    'salary_taxes': [44935]*12 + [137423, 137814, 140140, 148661, 150987, 153313, 198638, 200964, 203290, 205616, 207942, 210267],
    'opex': [168800, 186300, 204900, 223500, 359600, 395700, 431800, 467900, 504000, 540100, 581200, 727300,
             460550, 463210, 621050, 978890, 641730, 650070, 663410, 671250, 684590, 1181930, 699770, 707610],
    'hq_royalty_fix': [0, 0, 0, 45675, 45675, 45675, 45675, 78750, 78750, 78750, 78750, 78750,
                       111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825],
    'hq_royalty_deal': [0, 0, 0, 8400, 6752, 5927, 9303, 12679, 11855, 13130, 16506, 17782,
                        18863, 20267, 21670, 24477, 27980, 29383, 32887, 34290, 37793, 39197, 40600, 42003],
    'hq_crm_and_others': [19400, 5500, 5500, 18500, 18500, 18500, 18500, 18500, 45740, 50309, 54878, 59447,
                          40068, 33446, 33635, 33824, 34013, 100202, 34391, 34580, 34769, 34958, 101147, 35336],
    'capex': [0]*24
}

# Сценарий "Оптимальный" - Базовые показатели
optimal_baseline = {
    'leads': [440, 880, 1320, 1760, 2200, 2640, 3080, 3520, 3960, 4400, 4840, 5280,
              6160, 6600, 7040, 7480, 7920, 8360, 8800, 9240, 9680, 10120, 10560, 11000],
    'deals_rent': [3, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1,
                   3, 3, 4, 4, 4, 4, 5, 5, 6, 6, 6, 6],
    'deals_suburban': [0, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1,
                       2, 2, 2, 2, 3, 3, 4, 4, 5, 5, 5, 5],
    'revenue': [0, 0, 0, 800000, 1520000, 2623800, 2717000, 3714700, 4985694, 4985609, 5565525, 6725440,
                7814356, 8482712, 9167967, 10435378, 11620634, 12485890, 13451145, 14136401, 15321656, 15966912, 16872168, 17297423],
    'agent_comm': [0, 0, 0, 338500, 612100, 1053409, 1314265, 1784222, 2332633, 2318694, 2629481, 3171647,
                   4103166, 4502709, 4855674, 5449734, 6035280, 6593359, 7120893, 7473269, 8061122, 8387655, 8822681, 9066321],
    'salaries': [214935, 214935, 214935, 214935, 214935, 144935, 144935, 144935, 144935, 144935, 144935, 144935,
                 963068, 964023, 966913, 1126564, 1129454, 1462344, 1608235, 1611125, 1764015, 1766905, 1769796, 1772686],
    'salary_taxes': [44935]*12 + [143068, 144023, 146913, 156564, 159454, 162344, 208235, 211125, 214015, 216905, 219796, 222686],
    'opex': [150000, 286300, 321300, 357400, 393500, 464600, 535700, 606800, 677900, 749000, 820100, 896200,
             569750, 601530, 769250, 1136970, 809690, 827910, 851130, 868850, 892070, 1399290, 927010, 944730],
    'hq_royalty_fix': [0, 0, 0, 45675, 45675, 45675, 45675, 78750, 78750, 78750, 78750, 78750,
                       111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825],
    'hq_royalty_deal': [0, 0, 0, 4200, 8400, 13556, 14574, 19793, 25011, 26030, 29148, 34367,
                        39134, 42564, 45994, 52855, 58385, 61816, 67346, 70776, 76307, 79737, 83167, 86598],
    'hq_crm_and_others': [19400, 5500, 5500, 18500, 18500, 18500, 18500, 18500, 85147, 96247, 107347, 118470,
                          43344, 37268, 37730, 38192, 38654, 105116, 39578, 40040, 40502, 40964, 107426, 41888],
    'capex': [0]*24
}

# ----------------------------------------------------
# ШАПКА ПРИЛОЖЕНИЯ
# ----------------------------------------------------

st.markdown("<div class='main-title'>CENTURY 21 Financial Application (v2)</div>", unsafe_style_with_html=True)
st.markdown("<div class='sub-title'>Интерактивная финансовая модель и MVP-калькулятор для клиентов</div>", unsafe_style_with_html=True)

# ----------------------------------------------------
# БОКОВАЯ ПАНЕЛЬ С ДРАЙВЕРАМИ И ПАРАМЕТРАМИ
# ----------------------------------------------------

st.sidebar.markdown("<h2 style='color:#C2A24D; text-align:center;'>Панель управления</h2>", unsafe_style_with_html=True)

# 1. Сценарий
scenario = st.sidebar.selectbox(
    "📊 Выберите базовый сценарий проекта:",
    options=["Гипотеза 1 (Региональный рост)", "Бюджет Оптимальный (Интенсивный старт)"]
)

# Загрузка базовых данных под сценарий
if "Гипотеза 1" in scenario:
    base = hyp1_baseline.copy()
    default_conversion_deal = 15.0  # % встреч в договоры
    scenario_name = "Гипотеза 1"
    capex_m0 = 594500  # CAPEX на старте
    start_opex = 50000
else:
    base = optimal_baseline.copy()
    default_conversion_deal = 5.0   # % встреч в договоры
    scenario_name = "Оптимальный"
    capex_m0 = 1929500 # CAPEX на старте
    start_opex = 150000

# 2. Драйверы воронки продаж
st.sidebar.markdown("### 📈 Воронка продаж")
scaling_leads = st.sidebar.slider("Количество лидов (звонков), % к базе", min_value=50, max_value=200, value=100, step=10, help="Масштабирование воронки")
conversion_meeting = st.sidebar.slider("Конверсия лид -> встреча, %", min_value=2.0, max_value=20.0, value=10.0, step=0.5)
conversion_deal = st.sidebar.slider("Конверсия встреча -> договор, %", min_value=1.0, max_value=30.0, value=default_conversion_deal, step=0.5)

# 3. Средние чеки
st.sidebar.markdown("### 💰 Средние чеки сделок (₽)")
comm_secondary = st.sidebar.number_input("Вторичный рынок (базовая: 360к)", min_value=100000, max_value=1000000, value=360000, step=10000)
comm_primary = st.sidebar.number_input("Первичный рынок (базовая: 440к)", min_value=100000, max_value=1000000, value=440000, step=10000)
comm_suburban = st.sidebar.number_input("Загородная (базовая: 500к)", min_value=100000, max_value=1000000, value=500000, step=10000)

# 4. ФОТ и комиссия агентов
st.sidebar.markdown("### 👥 Персонал и Выплаты")
agent_commission_pct = st.sidebar.slider("Средний % выплат агентам", min_value=25, max_value=60, value=38, step=1)
backoffice_salary_mult = st.sidebar.slider("Оклады бэк-офиса, % к базе", min_value=50, max_value=150, value=100, step=5)

# 5. OPEX и Роялти
st.sidebar.markdown("### 🏢 Аренда и Франшиза")
office_rent_custom = st.sidebar.number_input("Аренда офиса (в месяц), ₽", min_value=30000, max_value=500000, value=150000, step=10000)
royalty_deal_custom = st.sidebar.number_input("Роялти за каждую сделку, ₽", min_value=1000, max_value=5000, value=2100, step=100)
tax_rate = st.sidebar.slider("Ставка налога УСН, %", min_value=3, max_value=15, value=7, step=1)

# ----------------------------------------------------
# ВЫЧИСЛИТЕЛЬНОЕ ЯДРО МОДЕЛИ
# ----------------------------------------------------

adj_leads = []
adj_meetings = []
adj_contracts = []
adj_prepayments = []
adj_deals_secondary = []
adj_deals_primary = []
adj_deals_rent = []
adj_deals_suburban = []

adj_revenue_secondary = []
adj_revenue_primary = []
adj_revenue_rent = []
adj_revenue_suburban = []
adj_revenue_total = []

adj_taxes_income = []
adj_agent_payouts = []
adj_backoffice_salaries = []
adj_backoffice_taxes = []
adj_opex_total = []
adj_hq_royalty_fix_total = []
adj_hq_royalty_deal_total = []
adj_hq_crm_total = []

adj_expenses_total = []
adj_capex_total = [0] * 24

# CAPEX Расписание в зависимости от выбранного сценария
if scenario_name == "Гипотеза 1":
    adj_capex_total[14] = 1335000  # 15-й Месяц (индекс 14)
    adj_capex_total[19] = 790000   # 20-й Месяц (индекс 19)
else:
    adj_capex_total[14] = 1335000
    adj_capex_total[19] = 790000

leads_mult = scaling_leads / 100.0
conv_meet_rate = conversion_meeting / 100.0
conv_deal_rate = conversion_deal / 100.0

for i in range(24):
    # Лиды и Воронка продаж
    L = int(base['leads'][i] * leads_mult)
    adj_leads.append(L)
    
    meetings = int(L * conv_meet_rate)
    adj_meetings.append(meetings)
    
    contracts = int(meetings * conv_deal_rate)
    adj_contracts.append(contracts)
    
    prepayments = contracts * 0.75
    adj_prepayments.append(prepayments)
    
    # Сделки
    deals_sec = prepayments * 0.90
    adj_deals_secondary.append(deals_sec)
    
    deals_prim = prepayments * 0.10
    adj_deals_primary.append(deals_prim)
    
    deals_r = base['deals_rent'][i]
    adj_deals_rent.append(deals_r)
    
    deals_sub = base['deals_suburban'][i]
    adj_deals_suburban.append(deals_sub)
    
    # Расчет доходов (Комиссия ВКД)
    rev_sec = deals_sec * comm_secondary
    adj_revenue_secondary.append(rev_sec)
    
    rev_prim = deals_prim * comm_primary
    adj_revenue_primary.append(rev_prim)
    
    rev_rent = deals_r * 80000  # Комиссия за аренду фиксирована на 80к
    adj_revenue_rent.append(rev_rent)
    
    rev_sub = deals_sub * comm_suburban
    adj_revenue_suburban.append(rev_sub)
    
    # Прочие доходы пропорционально лидам
    rev_others = (base['revenue'][i] - (base['deals_rent'][i]*80000 + base['deals_suburban'][i]*500000)) * (L / base['leads'][i] if base['leads'][i] > 0 else 0)
    if rev_others < 0:
        rev_others = 0
        
    total_r = rev_sec + rev_prim + rev_rent + rev_sub + rev_others
    adj_revenue_total.append(total_r)
    
    # Налоги
    tax = total_r * (tax_rate / 100.0)
    adj_taxes_income.append(tax)
    
    # Выплаты агентам
    agent_p = total_r * (agent_commission_pct / 100.0)
    adj_agent_payouts.append(agent_p)
    
    # Оклады бэк-офиса и налоги на ФОТ
    sal = base['salaries'][i] * (backoffice_salary_mult / 100.0)
    adj_backoffice_salaries.append(sal)
    sal_tax = sal * 0.209  # соотношение налогов на ФОТ к окладам
    adj_backoffice_taxes.append(sal_tax)
    
    # OPEX
    op = base['opex'][i]
    if scenario_name == "Гипотеза 1":
        if i >= 4:
            op = op - 150000 + office_rent_custom
        else:
            op = op - 50000 + (office_rent_custom / 3)
    else:
        op = op - 150000 + office_rent_custom
    adj_opex_total.append(op)
    
    # Роялти FIX
    adj_hq_royalty_fix_total.append(base['hq_royalty_fix'][i])
    
    # Роялти за сделку
    total_deals_count = deals_sec + deals_prim + deals_r + deals_sub
    adj_hq_royalty_deal_total.append(total_deals_count * royalty_deal_custom)
    
    # CRM и сервисы ЦО
    adj_hq_crm_total.append(base['hq_crm_and_others'][i])
    
    # Итого расходы
    tot_exp = agent_p + sal + sal_tax + op + base['hq_royalty_fix'][i] + (total_deals_count * royalty_deal_custom) + base['hq_crm_and_others'][i] + tax
    adj_expenses_total.append(tot_exp)
    
    # Чистая прибыль месяца
    net_p = total_r - tot_exp - adj_capex_total[i]
    adj_net_profit.append(net_p)

# Расчет накопленного денежного потока
cum_flow = []
running_sum = -capex_m0 - start_opex  # Месяц 0: CAPEX + стартовый OPEX
cum_flow.append(running_sum)

for i in range(24):
    running_sum += adj_net_profit[i]
    cum_flow.append(running_sum)

# ----------------------------------------------------
# ИНДИКАТОРЫ И KPI
# ----------------------------------------------------
total_revenue_2years = sum(adj_revenue_total)
total_profit_2years = sum(adj_net_profit) - capex_m0 - start_opex
max_investment = abs(min(cum_flow))

# Месяц окупаемости (первый месяц, где cum_flow >= 0)
payback_month = "Н/Д"
for idx, val in enumerate(cum_flow):
    if val >= 0:
        payback_month = f"{idx} мес."
        break

# Точка безубыточности (первый месяц с положительной чистой прибылью)
breakeven_month = "Н/Д"
for idx, val in enumerate(adj_net_profit):
    if val > 0:
        breakeven_month = f"{idx + 1} мес."
        break

# Отображение ключевых карточек KPI
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("💼 Общая выручка (2 года)", f"{total_revenue_2years:,.0f} ₽".replace(",", " "))
with col2:
    st.metric("📈 Итоговая чистая прибыль", f"{total_profit_2years:,.0f} ₽".replace(",", " "))
with col3:
    st.metric("📉 Пик инвестиций (CAPEX+OPEX)", f"{max_investment:,.0f} ₽".replace(",", " "))
with col4:
    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("🎯 Без-ость", breakeven_month)
    with col_b:
        st.metric("⏳ Окупаемость", payback_month)

# ----------------------------------------------------
# ВКЛАДКИ ДЛЯ КЛИЕНТА (ГРАФИКИ, ТАБЛИЦА, МЕТОДОЛОГИЯ)
# ----------------------------------------------------
tab1, tab2, tab3 = st.tabs(["📊 Интерактивные Графики", "📋 Детальный P&L (Помесячно)", "📝 Описание Модели"])

with tab1:
    st.markdown("### Визуализация финансового развития")
    
    # 1. График накопленного денежного потока
    fig_cum = go.Figure()
    fig_cum.add_trace(go.Scatter(
        x=list(range(25)),
        y=cum_flow,
        mode='lines+markers',
        name='Накопленный кэш-флоу',
        line=dict(color='#C2A24D', width=3),
        marker=dict(size=6, color='#ffffff', line=dict(color='#C2A24D', width=1))
    ))
    fig_cum.add_shape(
        type="line", x0=0, y0=0, x1=24, y1=0,
        line=dict(color="#ffffff", width=1, dash="dash")
    )
    fig_cum.update_layout(
        title="Кривая накопленного кэш-флоу (Выход из инвестиционной фазы)",
        xaxis_title="Месяцы проекта",
        yaxis_title="Баланс кэша, ₽",
        template="plotly_dark",
        plot_bgcolor="#1a1a1a",
        paper_bgcolor="#111111",
        font=dict(color="#ffffff")
    )
    st.plotly_chart(fig_cum, use_container_width=True)
    
    # 2. Столбчатый график Доходы vs Расходы
    fig_rev_exp = go.Figure()
    fig_rev_exp.add_trace(go.Bar(
        x=months,
        y=adj_revenue_total,
        name='Доходы (Комиссия ВКД)',
        marker_color='#C2A24D'
    ))
    fig_rev_exp.add_trace(go.Bar(
        x=months,
        y=adj_expenses_total,
        name='Расходы',
        marker_color='#555555'
    ))
    fig_rev_exp.update_layout(
        title="Ежемесячные доходы и расходы",
        xaxis_title="Месяцы",
        yaxis_title="Сумма, ₽",
        barmode='group',
        template="plotly_dark",
        plot_bgcolor="#1a1a1a",
        paper_bgcolor="#111111",
        font=dict(color="#ffffff")
    )
    st.plotly_chart(fig_rev_exp, use_container_width=True)

    # 3. Круговая диаграмма структуры расходов
    avg_expenses = {
        'Выплаты агентам': sum(adj_agent_payouts),
        'Оклады бэк-офиса': sum(adj_backoffice_salaries) + sum(adj_backoffice_taxes),
        'Аренда и OPEX': sum(adj_opex_total),
        'Роялти и CRM ЦО': sum(adj_hq_royalty_fix_total) + sum(adj_hq_royalty_deal_total) + sum(adj_hq_crm_total),
        'Налоги': sum(adj_taxes_income)
    }
    fig_pie = px.pie(
        names=list(avg_expenses.keys()),
        values=list(avg_expenses.values()),
        color_discrete_sequence=['#C2A24D', '#E2C26D', '#555555', '#333333', '#888888'],
        title="Структура совокупных расходов за 2 года"
    )
    fig_pie.update_layout(
        template="plotly_dark",
        paper_bgcolor="#111111",
        font=dict(color="#ffffff")
    )
    st.plotly_chart(fig_pie, use_container_width=True)

with tab2:
    st.markdown("### Детальная таблица доходов и расходов (P&L)")
    
    p_and_l_data = {
        'Метрика': [
            'Лиды / Звонки', 'Встречи', 'Договоры', 'Задатки', 
            'Сделки Вторичка', 'Сделки Первичка', 'Сделки Аренда', 'Сделки Загородная',
            'ВЫРУЧКА (Комиссия ВКД)', 'Налог УСН', 
            'Расходы: Выплаты агентам', 'Расходы: Оклады бэк-офиса', 'Расходы: Налоги на ФОТ',
            'Расходы: Офис и OPEX', 'Расходы: Роялти FIX', 'Расходы: Роялти со сделки', 
            'Расходы: CRM и сервисы ЦО', 'Капитальные затраты (CAPEX)', 'ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ'
        ]
    }
    
    for idx, m in enumerate(months):
        p_and_l_data[m] = [
            adj_leads[idx],
            adj_meetings[idx],
            adj_contracts[idx],
            f"{adj_prepayments[idx]:.1f}",
            f"{adj_deals_secondary[idx]:.1f}",
            f"{adj_deals_primary[idx]:.1f}",
            adj_deals_rent[idx],
            adj_deals_suburban[idx],
            f"{adj_revenue_total[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_taxes_income[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_agent_payouts[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_backoffice_salaries[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_backoffice_taxes[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_opex_total[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_hq_royalty_fix_total[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_hq_royalty_deal_total[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_hq_crm_total[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_capex_total[idx]:,.0f} ₽".replace(",", " "),
            f"{adj_net_profit[idx]:,.0f} ₽".replace(",", " ")
        ]
        
    df_p_and_l = pd.DataFrame(p_and_l_data)
    st.dataframe(df_p_and_l, use_container_width=True, height=600)
    
    # ----------------------------------------------------
    # ГЕНЕРАЦИЯ ФОРМАТИРОВАННОГО ЭКСКЕЛЯ С openpyxl
    # ----------------------------------------------------
    
    def generate_excel():
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "C21 P&L Export"
        
        ws.views.sheetView[0].showGridLines = True
        
        # Корпоративные цвета и стили
        gold_fill = PatternFill(start_color="C2A24D", end_color="C2A24D", fill_type="solid")
        dark_fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
        light_gold_fill = PatternFill(start_color="F9F5EB", end_color="F9F5EB", fill_type="solid")
        
        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_header = Font(name="Calibri", size=11, bold=True, color="111111")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        
        thin_border_side = Side(border_style="thin", color="D3D3D3")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        double_bottom_border = Border(bottom=Side(border_style="double", color="111111"), top=thin_border_side)
        
        # Заголовок
        ws.merge_cells("A1:Y1")
        ws["A1"] = f"ФИНАНСОВАЯ МОДЕЛЬ CENTURY 21 - СЦЕНАРИЙ: {scenario_name.upper()}"
        ws["A1"].font = font_title
        ws["A1"].fill = dark_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # Заголовки столбцов
        headers = ["Метрика / Период"] + months
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = gold_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 28
        
        # Данные по строкам
        metrics_rows = [
            ('Количество лидов (звонков)', adj_leads, "0"),
            ('Встречи', adj_meetings, "0"),
            ('Договоры', adj_contracts, "0"),
            ('Задатки', adj_prepayments, "0.0"),
            ('Сделки: Вторичный рынок', adj_deals_secondary, "0.0"),
            ('Сделки: Первичный рынок', adj_deals_primary, "0.0"),
            ('Сделки: Аренда', adj_deals_rent, "0"),
            ('Сделки: Загородная недвижимость', adj_deals_suburban, "0"),
            ('ВЫРУЧКА ИТОГО (Комиссия ВКД), ₽', adj_revenue_total, "#,##0"),
            ('Налог с доходов УСН, ₽', adj_taxes_income, "#,##0"),
            ('Выплаты агентам, ₽', adj_agent_payouts, "#,##0"),
            ('Оклады бэк-офиса, ₽', adj_backoffice_salaries, "#,##0"),
            ('Налоги на ФОТ оклады, ₽', adj_backoffice_taxes, "#,##0"),
            ('Офисные и прочие OPEX, ₽', adj_opex_total, "#,##0"),
            ('Роялти FIX, ₽', adj_hq_royalty_fix_total, "#,##0"),
            ('Роялти за сделки в ЦО, ₽', adj_hq_royalty_deal_total, "#,##0"),
            ('CRM и прочие IT-сервисы ЦО, ₽', adj_hq_crm_total, "#,##0"),
            ('Капитальные расходы (CAPEX), ₽', adj_capex_total, "#,##0"),
            ('ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ, ₽', adj_net_profit, "#,##0")
        ]
        
        for r_idx, (m_name, data_list, num_format) in enumerate(metrics_rows, 4):
            ws.cell(row=r_idx, column=1, value=m_name).font = font_bold if "ИТОГО" in m_name or "ПРИБЫЛЬ" in m_name else font_regular
            ws.cell(row=r_idx, column=1).border = thin_border
            if "ИТОГО" in m_name or "ПРИБЫЛЬ" in m_name:
                ws.cell(row=r_idx, column=1).fill = light_gold_fill
                
            for c_idx, val in enumerate(data_list, 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.number_format = num_format
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right")
                
                if "ИТОГО" in m_name or "ПРИБЫЛЬ" in m_name:
                    cell.font = font_bold
                    cell.fill = light_gold_fill
                    if "ПРИБЫЛЬ" in m_name:
                        cell.border = double_bottom_border
                else:
                    cell.font = font_regular
                    
            ws.row_dimensions[r_idx].height = 20
            
        # Автоподгон ширины колонок
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        ws.column_dimensions["A"].width = 38
        
        wb.save(output)
        return output.getvalue()
    
    excel_data = generate_excel()
    st.download_button(
        label="📥 Скачать финансовую модель в Excel",
        data=excel_data,
        file_name=f"C21_Financial_Model_{scenario_name}_v2.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

with tab3:
    st.markdown("""
    ### 📖 Справочная методология моделирования
    
    Эта интерактивная модель разработана для симуляции финансового плана агентства недвижимости **Century 21**:
    
    1. **Базовый сценарий:**
       * **Гипотеза 1 (Региональный рост):** Постепенный выход на плановое количество лидов, умеренные стартовые инвестиции (CAPEX ~595k) и поэтапная закупка оборудования во 2-й год.
       * **Бюджет Оптимальный (Интенсивный старт):** Агрессивный старт с полной меблировкой и закупкой оборудования с 1-го месяца (CAPEX ~1.93 млн), максимальные рекламные вложения с первых месяцев для ускоренного роста.
       
    2. **Конверсия воронки:**
       * **Звонки -> Встречи (базово 10%):** Отражает эффективность работы администраторов и рекрутеров при первичном контакте.
       * **Встречи -> Договоры:** Отражает квалификацию агентов (базовая для Гипотезы 1 — 15%, для Оптимального — 5%).
       
    3. **Финансовые результаты:**
       * На основе настроенных средних чеков (Комиссии ВКД) и процента выплат агентам рассчитывается операционная рентабельность, позволяя увидеть срок возврата инвестиций (Payback) и точку безубыточности (Break-even).
    """)
