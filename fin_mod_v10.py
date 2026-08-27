import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import io

# Проверяем доступность openpyxl для экспорта в Excel без падения приложения
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    openpyxl_available = True
except ImportError:
    openpyxl_available = False

# Установка конфигурации страницы
st.set_page_config(
    page_title="CENTURY 21 Financial Model MVP v10",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Пользовательский CSS-стилинг в корпоративных светлых цветах Century 21 (Золотой #BEAF87, Черный #252526, Светлый фон #F9F9FB)
st.markdown("""
<style>
    .stApp {
        background-color: #F9F9FB;
        color: #252526;
        font-family: 'Segoe UI', Arial, sans-serif;
    }
    [data-testid="stSidebar"] {
        background-color: #F1F1F3 !important;
        border-right: 1px solid #E6E7E8;
    }
    .stMetric {
        background-color: #FFFFFF;
        padding: 20px;
        border-radius: 10px;
        border-left: 5px solid #BEAF87;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        margin-bottom: 10px;
    }
    .stMetric label {
        color: #777779 !important;
        font-weight: 600 !important;
        font-size: 14px !important;
    }
    .stMetric div[data-testid="stMetricValue"] {
        color: #A19276 !important;
        font-size: 26px !important;
        font-weight: 800 !important;
    }
    .main-title {
        color: #252526;
        font-size: 34px;
        font-weight: 900;
        text-align: center;
        margin-bottom: 5px;
        text-transform: uppercase;
        letter-spacing: 1.5px;
    }
    .brand-accent {
        color: #BEAF87;
    }
    .sub-title {
        color: #777779;
        font-size: 16px;
        text-align: center;
        margin-bottom: 30px;
        font-style: italic;
    }
    div.stButton > button:first-child {
        background-color: #BEAF87;
        color: #FFFFFF;
        font-weight: bold;
        border: none;
        border-radius: 6px;
        padding: 12px 24px;
        transition: all 0.3s ease;
        box-shadow: 0 2px 4px rgba(190, 175, 135, 0.2);
    }
    div.stButton > button:first-child:hover {
        background-color: #252526;
        color: #FFFFFF;
        box-shadow: 0 4px 12px rgba(37, 37, 38, 0.2);
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #FFFFFF;
        border: 1px solid #E6E7E8;
        border-radius: 6px 6px 0px 0px;
        padding: 12px 24px;
        color: #777779;
        font-weight: 600;
    }
    .stTabs [aria-selected="true"] {
        background-color: #BEAF87 !important;
        color: #FFFFFF !important;
        border-color: #BEAF87 !important;
    }
    .section-header {
        background-color: #F1F1F3;
        padding: 8px 12px;
        border-radius: 4px;
        font-weight: bold;
        margin-top: 15px;
        color: #252526;
    }
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# ОПРЕДЕЛЕНИЕ БАЗОВЫХ ДАННЫХ (На основе источников C21)
# ----------------------------------------------------

months = [f"Месяц {i}" for i in range(1, 25)]

# Сценарий "Гипотеза 1" - Базовые показатели
hyp1_baseline = {
    'leads': [60, 120, 180, 240, 300, 360, 420, 480, 540, 600, 660, 720,
              840, 900, 960, 1020, 1080, 1140, 1200, 1260, 1320, 1380, 1440, 1500],
    'meetings': [6, 12, 18, 24, 30, 36, 42, 48, 54, 60, 66, 72,
                 84, 90, 96, 102, 108, 114, 120, 126, 132, 138, 144, 150],
    'contracts': [0, 1, 2, 3, 4, 5, 5, 6, 7, 8, 9, 10,
                  11, 13, 14, 14, 15, 16, 17, 18, 19, 20, 21, 22],
    'prepayments': [0.0, 0.8, 1.5, 2.3, 3.0, 3.8, 3.8, 4.5, 5.3, 6.0, 6.8, 7.5,
                    8.3, 9.8, 10.5, 10.5, 11.3, 12.0, 12.8, 13.5, 14.3, 15.0, 15.8, 16.5],
    
    # Декомпозиция ДОХОДОВ
    'rev_sec': [0, 0, 0, 1080000, 437400, 656100, 874800, 1093500, 1312200, 1530900, 1749600, 1968300,
                2187000, 2405700, 2624400, 3061800, 3280500, 3499200, 3717900, 3936600, 4155300, 4374000, 4592700, 4811400],
    'rev_prim': [0, 0, 0, 440000, 440000, 440000, 880000, 880000, 880000, 880000, 880000, 880000,
                 267300, 294030, 320760, 374220, 400950, 427680, 454410, 481140, 507870, 534600, 561330, 588060],
    'rev_rent': [0, 0, 0, 0, 80000, 0, 0, 80000, 0, 0, 80000, 80000,
                 240000, 240000, 280000, 280000, 320000, 320000, 360000, 400000, 440000, 440000, 480000, 480000],
    'rev_sub': [0, 0, 0, 0, 0, 0, 0, 500000, 0, 0, 500000, 500000,
                1000000, 1000000, 1000000, 1000000, 1500000, 1500000, 2000000, 2000000, 2500000, 2500000, 2500000, 2500000],
    'rev_overseas': [0]*17 + [220000] + [0]*4 + [220000] + [0],
    'rev_other_p': [0]*12 + [75600]*12,
    'rev_mort': [0]*8 + [43376, 50605, 57834, 65063, 72293, 79522, 86751, 101210, 108439, 115668, 122897, 130127, 137356, 144585, 151814, 159044],
    'rev_ins': [0]*8 + [11154, 13013, 14872, 16731, 18590, 20448, 22307, 26025, 27884, 29743, 31602, 33461, 35320, 37179, 39038, 40897],
    'rev_legal': [0]*6 + [56700, 66150, 75600, 85050, 94500, 103950, 113400, 132300, 141750, 151200, 160650, 170100, 179550, 189000, 198450, 207900, 217350, 226800],
    'revenue': [0, 0, 0, 1520000, 957400, 1096100, 1811500, 2619650, 2322329, 2559567, 3376806, 3614044,
                3974182, 4247600, 4551568, 5070055, 5874023, 6357991, 6941959, 7245928, 8049896, 8313864, 8837832, 8881800],

    # Декомпозиция РАСХОДОВ - ФОТ
    'salaries': [214935, 214935, 214935, 214935, 214935, 144935, 144935, 144935, 144935, 144935, 144935, 144935,
                 957423, 957814, 960140, 1118661, 1120987, 1453313, 1598638, 1600964, 1753290, 1755616, 1757942, 1760267],
    'salary_taxes': [44935]*12 + [137423, 137814, 140140, 148661, 150987, 153313, 198638, 200964, 203290, 205616, 207942, 210267],
    'agent_comm': [0, 0, 0, 612100, 398312, 460152, 892040, 1274170, 1109098, 1215943, 1628083, 1734144,
                   2053406, 2216022, 2372104, 2625039, 3002856, 3369888, 3678751, 3834367, 4214813, 4349624, 4587930, 4640042],

    # Декомпозиция OPEX
    'opex': [168800, 186300, 204900, 223500, 359600, 395700, 431800, 467900, 504000, 540100, 581200, 727300,
             460550, 463210, 621050, 978890, 641730, 650070, 663410, 671250, 684590, 1181930, 699770, 707610],

    # Декомпозиция ВЫПЛАТ ЦО
    'hq_royalty_fix': [0, 0, 0, 45675, 45675, 45675, 45675, 78750, 78750, 78750, 78750, 78750,
                       111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825],
    'hq_royalty_deal': [0, 0, 0, 8400, 6752, 5927, 9303, 12679, 11855, 13130, 16506, 17782,
                        18863, 20267, 21670, 24477, 27980, 29383, 32887, 34290, 37793, 39197, 40600, 42003],
    'hq_crm_and_others': [19400, 5500, 5500, 18500, 18500, 18500, 18500, 18500, 45740, 50309, 54878, 59447,
                          40068, 33446, 33635, 33824, 34013, 100202, 34391, 34580, 34769, 34958, 101147, 35336],
    'capex': [0]*24
}

# Сценарий "Бюджет Оптимальный" - Базовые показатели
optimal_baseline = {
    'leads': [440, 880, 1320, 1760, 2200, 2640, 3080, 3520, 3960, 4400, 4840, 5280,
              6160, 6600, 7040, 7480, 7920, 8360, 8800, 9240, 9680, 10120, 10560, 11000],
    'meetings': [44, 88, 132, 176, 220, 264, 308, 352, 396, 440, 484, 528,
                 616, 660, 704, 748, 792, 836, 880, 924, 968, 1012, 1056, 1100],
    'contracts': [0, 2, 4, 7, 9, 11, 13, 15, 18, 20, 22, 24,
                  26, 31, 33, 35, 37, 40, 42, 44, 46, 48, 51, 53],
    'prepayments': [0.0, 1.5, 3.0, 5.3, 6.8, 8.3, 9.8, 11.3, 13.5, 15.0, 16.5, 18.0,
                    19.5, 23.3, 24.8, 26.3, 27.8, 30.0, 31.5, 33.0, 34.5, 36.0, 38.3, 39.8],
    
    # Декомпозиция ДОХОДОВ
    'rev_sec': [0, 0, 0, 360000, 1080000, 1603800, 2138400, 2673000, 3207600, 3742200, 4276800, 4811400,
                5346000, 5880600, 6415200, 7484400, 8019000, 8553600, 9088200, 9622800, 10157400, 10692000, 11226600, 11761200],
    'rev_prim': [0, 0, 0, 440000, 440000, 440000, 440000, 880000, 880000, 880000, 880000, 880000,
                 653400, 718740, 784080, 914760, 980100, 1045440, 1110780, 1176120, 1241460, 1306800, 1372140, 1437480],
    'rev_rent': [0, 0, 0, 0, 0, 80000, 0, 0, 80000, 0, 0, 80000,
                 240000, 240000, 280000, 280000, 320000, 320000, 360000, 400000, 440000, 440000, 480000, 480000],
    'rev_sub': [0, 0, 0, 0, 0, 500000, 0, 0, 500000, 0, 0, 500000,
                1000000, 1000000, 1000000, 1000000, 1500000, 1500000, 2000000, 2000000, 2500000, 2500000, 2500000, 2500000],
    'rev_overseas': [0]*17 + [220000] + [0]*4 + [220000] + [0],
    'rev_other_p': [0]*12 + [75600]*12,
    'rev_mort': [0]*8 + [106029, 123701, 141372, 159044, 176715, 194387, 212058, 247401, 265073, 282744, 300416, 318087, 335759, 353430, 371102, 388773],
    'rev_ins': [0]*8 + [27265, 31809, 36353, 40897, 45441, 49985, 54529, 63617, 68162, 72706, 77250, 81794, 86338, 90882, 95426, 99970],
    'rev_legal': [0]*6 + [138600, 161700, 184800, 207900, 231000, 254100, 277200, 323400, 346500, 369600, 392700, 415800, 438900, 462000, 485100, 508200, 531300, 554400],
    'revenue': [0, 0, 0, 800000, 1520000, 2623800, 2717000, 3714700, 4985694, 4985609, 5565525, 6725440,
                7814356, 8482712, 9167967, 10435378, 11620634, 12485890, 13451145, 14136401, 15321656, 15966912, 16872168, 17297423],

    # Декомпозиция РАСХОДОВ - ФОТ
    'salaries': [214935, 214935, 214935, 214935, 214935, 144935, 144935, 144935, 144935, 144935, 144935, 144935,
                 963068, 964023, 966913, 1126564, 1129454, 1462344, 1608235, 1611125, 1764015, 1766905, 1769796, 1772686],
    'salary_taxes': [44935]*12 + [143068, 144023, 146913, 156564, 159454, 162344, 208235, 211125, 214015, 216905, 219796, 222686],
    'agent_comm': [0, 0, 0, 338500, 612100, 1053409, 1314265, 1784222, 2332633, 2318694, 2629481, 3171647,
                   4103166, 4502709, 4855674, 5449734, 6035280, 6593359, 7120893, 7473269, 8061122, 8387655, 8822681, 9066321],

    # Декомпозиция OPEX
    'opex': [150000, 286300, 321300, 357400, 393500, 464600, 535700, 606800, 677900, 749000, 820100, 896200,
             569750, 601530, 769250, 1136970, 809690, 827910, 851130, 868850, 892070, 1399290, 927010, 944730],

    # Декомпозиция ВЫПЛАТ ЦО
    'hq_royalty_fix': [0, 0, 0, 45675, 45675, 45675, 45675, 78750, 78750, 78750, 78750, 78750,
                       111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825],
    'hq_royalty_deal': [0, 0, 0, 4200, 8400, 13556, 14574, 19793, 25011, 26030, 29148, 34367,
                        39134, 42564, 45994, 52855, 58385, 61816, 67346, 70776, 76307, 79737, 83167, 86598],
    'hq_crm_and_others': [19400, 5500, 5500, 68375, 72575, 77731, 78749, 117043, 188908, 201034, 215260, 231587,
                          194303, 191657, 195549, 202872, 208864, 278757, 218749, 222641, 228634, 232526, 302418, 240311],
    'capex': [0]*24
}

# Региональные параметры и пресеты
REGIONAL_PRESETS = {
    "Москва": {
        "pau_fee": 577500,
        "comm_sec": 360000,
        "comm_prim": 440000,
        "comm_sub": 500000,
        "comm_rent": 80000,
        "rent_base": 150000,
        "salary_mult": 1.0,
        "royalty_fix_factor": 1.0,
        "royalty_deal": 2100,
    },
    "Московская область (МО)": {
        "pau_fee": 500000,
        "comm_sec": 280000,
        "comm_prim": 330000,
        "comm_sub": 400000,
        "comm_rent": 60000,
        "rent_base": 100000,
        "salary_mult": 0.85,
        "royalty_fix_factor": 0.85,
        "royalty_deal": 2000,
    },
    "Крупные города (население > 800к)": {
        "pau_fee": 450000,
        "comm_sec": 220000,
        "comm_prim": 260000,
        "comm_sub": 320000,
        "comm_rent": 45000,
        "rent_base": 80000,
        "salary_mult": 0.75,
        "royalty_fix_factor": 0.75,
        "royalty_deal": 2000,
    },
    "Средние города (население 400к - 800к)": {
        "pau_fee": 350000,
        "comm_sec": 160000,
        "comm_prim": 200000,
        "comm_sub": 240000,
        "comm_rent": 35000,
        "rent_base": 55000,
        "salary_mult": 0.60,
        "royalty_fix_factor": 0.60,
        "royalty_deal": 1800,
    },
    "Малые города (население < 400к)": {
        "pau_fee": 270000,
        "comm_sec": 110000,
        "comm_prim": 140000,
        "comm_sub": 180000,
        "comm_rent": 25000,
        "rent_base": 38000,
        "salary_mult": 0.45,
        "royalty_fix_factor": 0.45,
        "royalty_deal": 1500,
    }
}

# ----------------------------------------------------
# ШАПКА ПРИЛОЖЕНИЯ
# ----------------------------------------------------
st.markdown("<div class='main-title'>💼 CENTURY 21 <span class='brand-accent'>Financial Application MVP</span></div>", unsafe_allow_html=True)
st.markdown("<div class='sub-title'>Интерактивная финансовая модель и декомпозиция доходов/расходов по регионам РФ</div>", unsafe_allow_html=True)

# ----------------------------------------------------
# БОКОВАЯ ПАНЕЛЬ С УПРАВЛЕНИЕМ & ТУЛТИПАМИ
# ----------------------------------------------------
st.sidebar.markdown("<h2 style='color:#252526; text-align:center; font-weight:800;'>Управление моделью</h2>", unsafe_allow_html=True)

# Регион
region_select = st.sidebar.selectbox(
    "🌍 Выберите регион / тип города:",
    options=list(REGIONAL_PRESETS.keys()),
    help="Каждый пресет автоматически настраивает средние чеки сделок, стоимость аренды офиса, оклады персонала и fixed-роялти франшизы под экономический уровень выбранной территории."
)
preset = REGIONAL_PRESETS[region_select]

# Сценарий
scenario = st.sidebar.selectbox(
    "📊 Базовый сценарий проекта:",
    options=["Гипотеза 1 (Региональный рост)", "Бюджет Оптимальный (Интенсивный старт)"],
    help="Гипотеза 1 – экономичный запуск с поэтапной закупкой оборудования и умеренной воронкой лидов. Бюджет Оптимальный – агрессивный запуск с полным оснащением офиса с первых месяцев."
)

st.sidebar.markdown("### 🎚️ Настройки калибровки")

# Паушальный
regional_pau_fee = st.sidebar.number_input(
    "Паушальный взнос (с НДС), ₽",
    min_value=150000,
    max_value=1500000,
    value=preset["pau_fee"],
    step=10000,
    help="Единоразовый лицензионный платеж за франшизу. Изменяется на основе пресета выбранного города."
)

# Воронка
st.sidebar.markdown("### 📈 Воронка продаж")
scaling_leads = st.sidebar.slider(
    "Количество лидов (звонков), % к базе",
    min_value=50,
    max_value=200,
    value=100,
    step=10,
    help="Регулирует входящий поток клиентов. Увеличение лидов прямо пропорционально масштабирует доходы от сделок по вторичному и первичному рынку."
)

conversion_meeting = st.sidebar.slider(
    "Конверсия лид -> встреча, %",
    min_value=2.0,
    max_value=20.0,
    value=10.0,
    step=0.5,
    help="Какая часть позвонивших клиентов доходит до личной встречи в офисе. Базовый корпоративный стандарт Century 21 составляет 10%."
)

default_conversion_deal = 15.0 if "Гипотеза 1" in scenario else 5.0
conversion_deal = st.sidebar.slider(
    "Конверсия встреча -> договор, %",
    min_value=1.0,
    max_value=30.0,
    value=default_conversion_deal,
    step=0.5,
    help="Эффективность работы агентов при переговорах. Показывает долю встреч, переходящих в подписанные эксклюзивные договоры."
)

# Средние чеки
st.sidebar.markdown("### 💰 Средние комиссии (чеки сделок)")
comm_secondary = st.sidebar.number_input(
    "Вторичный рынок, ₽",
    min_value=50000,
    max_value=1000000,
    value=preset["comm_sec"],
    step=10000,
    help="Средний комиссионный доход агентства с одной сделки купли-продажи на вторичном рынке недвижимости."
)
comm_primary = st.sidebar.number_input(
    "Первичный рынок, ₽",
    min_value=50000,
    max_value=1000000,
    value=preset["comm_prim"],
    step=10000,
    help="Средняя комиссия, получаемая от девелоперов за продажу новостройки."
)
comm_suburban = st.sidebar.number_input(
    "Загородная недвижимость, ₽",
    min_value=50000,
    max_value=1000000,
    value=preset["comm_sub"],
    step=10000,
    help="Средний комиссионный чек со сделок по загородным домам, коттеджам и участкам."
)
comm_rent_val = st.sidebar.number_input(
    "Аренда (жилая/коммерческая), ₽",
    min_value=10000,
    max_value=200000,
    value=preset["comm_rent"],
    step=5000,
    help="Комиссионный доход со сделок аренды жилой и коммерческой недвижимости."
)

# ФОТ
st.sidebar.markdown("### 👥 Персонал и Выплаты")
agent_commission_pct = st.sidebar.slider(
    "Средний % выплат агентам",
    min_value=25,
    max_value=60,
    value=38,
    step=1,
    help="Средний процент от комиссии по сделкам, выплачиваемый агентам. Отражает мотивационную сетку Century 21 (35%-50% в зависимости от категории D, C, B, A)."
)
backoffice_salary_mult = st.sidebar.slider(
    "Индекс окладов бэк-офиса, % к базе",
    min_value=50,
    max_value=150,
    value=100,
    step=5,
    help="Шкала фиксированных окладов бэк-офиса (РОП, HR/рекрутер, юрист, ипотечный брокер, листинг-менеджер, администратор)."
)

# OPEX и Роялти
st.sidebar.markdown("### 🏢 Аренда и Налоги")
office_rent_custom = st.sidebar.number_input(
    "Аренда офиса в месяц, ₽",
    min_value=10000,
    max_value=500000,
    value=preset["rent_base"],
    step=5000,
    help="Стоимость аренды офисного помещения. Применяются поправки на ранних этапах запуска офиса."
)
royalty_deal_custom = st.sidebar.number_input(
    "Роялти за каждую сделку, ₽",
    min_value=1000,
    max_value=5000,
    value=preset["royalty_deal"],
    step=100,
    help="Фиксированный платеж в пользу управляющей компании за факт проведения транзакции."
)
tax_rate = st.sidebar.slider(
    "Ставка налога УСН, %",
    min_value=3,
    max_value=15,
    value=7,
    step=1,
    help="Действующая налоговая ставка по УСН (доходы) с учетом региональных субсидий и льгот."
)

# ----------------------------------------------------
# ВЫЧИСЛИТЕЛЬНОЕ ЯДРО МОДЕЛИ (Полная декомпозиция)
# ----------------------------------------------------

# Загрузка базовых данных под сценарий
if "Гипотеза 1" in scenario:
    base = hyp1_baseline.copy()
    scenario_name = "Гипотеза 1"
    start_opex = 50000 # Офисный OPEX Месяца 0
else:
    base = optimal_baseline.copy()
    scenario_name = "Оптимальный"
    start_opex = 150000

# Региональный коэффициент ФОТ и Роялти
salary_mult = preset["salary_mult"]
royalty_fix_mult = preset["royalty_fix_factor"]

# Инициализация всех списков для декомпозиции P&L
adj_leads = []
adj_meetings = []
adj_contracts = []
adj_prepayments = []
adj_deals_secondary = []
adj_deals_primary = []
adj_deals_rent = []
adj_deals_suburban = []

# Доходы
rev_secondary_list = []
rev_primary_list = []
rev_rent_list = []
rev_suburban_list = []
rev_overseas_list = []
rev_other_p_list = []
rev_mortgage_list = []
rev_insurance_list = []
rev_legal_list = []
rev_total_list = []

# Расходы - ФОТ
payouts_agents_list = []
salaries_backoffice_list = []
taxes_payroll_list = []
total_payroll_list = []

# Расходы - OPEX
rent_list = []
internet_list = []
mobile_list = []
kanc_list = []
reklama_list = []
hh_list = []
buh_list = []
bank_list = []
cleaning_list = []
gsm_list = []
courier_list = []
events_list = []
total_opex_list = []

# Расходы - Роялти и ЦО
hq_royalty_fix_list = []
hq_royalty_deal_list = []
hq_crm_list = []
hq_kc_list = []
hq_mort_list = []
hq_overseas_list = []
hq_legal_list = []
hq_insurance_list = []
total_hq_payments_list = []

# Налоги
taxes_usn_list = []

# Капитальные затраты (CAPEX)
capex_franchise_list = [0]*24
capex_renovation_list = [0]*24
capex_equipment_list = [0]*24
total_capex_list = []

# Итоговая чистая прибыль
adj_net_profit = []

# Заполнение CAPEX календаря (Месяц 1 - 24)
if scenario_name == "Гипотеза 1":
    # Месяц 5 (индекс 4)
    capex_renovation_list[4] = 520000 # ремонт + брендирование
    capex_equipment_list[4] = 815000 # оборудование + мебель
    # Месяц 15 (индекс 14)
    capex_renovation_list[14] = 520000
    capex_equipment_list[14] = 815000
    # Месяц 20 (индекс 19)
    capex_equipment_list[19] = 790000
else:
    # Оптимальный
    capex_renovation_list[14] = 520000
    capex_equipment_list[14] = 815000
    capex_equipment_list[19] = 790000

leads_mult = scaling_leads / 100.0
conv_meet_rate = conversion_meeting / 10.0 # Относительно базовых 10%
conv_deal_rate = conversion_deal / default_conversion_deal # Относительно базовых 15% или 5%
overall_voeronka_factor = leads_mult * conv_meet_rate * conv_deal_rate

# Подготовка OPEX календаря (Гипотеза 1 или Оптимальный)
base_internet = [5000]*12 + [950]*12
if scenario_name == "Гипотеза 1":
    base_mobile = [7800]*2 + [8400] + [9000] + [9600] + [10200] + [10800] + [11400] + [12000] + [12600] + [13200] + [13800] + [18000] + [7560] + [8100] + [8640] + [9180] + [9720] + [10260] + [10800] + [11340] + [11880] + [12420] + [12960]
    base_rent_list = [50000]*4 + [150000]*8 + [150000]*2 + [300000]*10
else:
    base_mobile = [7800]*2 + [8400] + [9000] + [9600] + [10200] + [10800] + [11400] + [12000] + [12600] + [13200] + [13800] + [18000] + [18480] + [19800] + [21120] + [22440] + [23760] + [25080] + [26400] + [27720] + [29040] + [30360] + [31680]
    base_rent_list = [150000]*12 + [150000]*2 + [300000]*10

base_kanc = [6500]*2 + [7000] + [7500] + [8000] + [8500] + [9000] + [9500] + [10000] + [10500] + [11000] + [11500] + [15000] + [15500] + [16500] + [17500] + [18500] + [20000] + [21500] + [22500] + [24000] + [24500] + [25500] + [26500]
if scenario_name == "Гипотеза 1":
    base_reklama = [17500] + [35000] + [52500] + [70000] + [105000] + [140000] + [175000] + [210000] + [245000] + [280000] + [315000] + [350000] + [75600] + [88200] + [94500] + [100800] + [107100] + [113400] + [119700] + [126000] + [132300] + [138600] + [144900] + [151200]
else:
    base_reklama = [35000] + [70000] + [105000] + [140000] + [210000] + [280000] + [350000] + [420000] + [490000] + [560000] + [630000] + [700000] + [184800] + [215600] + [231000] + [246400] + [261800] + [277200] + [292600] + [308000] + [323400] + [338800] + [354200] + [369600]

base_hh = [45000]*12 + [90000]*12
base_buh = [20000]*12 + [40000]*12
base_bank = [2000]*12 + [6000]*12
base_cleaning = [15000]*12 + [45000]*12
base_gsm = [0]*10 + [5000]*2 + [10000]*4 + [15000]*2 + [20000]*2 + [25000]*4
base_courier = [0]*11 + [10000]*13

base_events = [0]*24
base_events[11] = 100000
base_events[15] = 350000
base_events[21] = 490000

for i in range(24):
    # 1. Лиды и воронка
    L = int(base['leads'][i] * leads_mult)
    adj_leads.append(L)
    adj_meetings.append(int(base['meetings'][i] * leads_mult * conv_meet_rate))
    adj_contracts.append(int(base['contracts'][i] * overall_voeronka_factor))
    adj_prepayments.append(base['prepayments'][i] * overall_voeronka_factor)
    
    # 2. ДОХОДЫ с учетом 1-3 месяцев строго равных нулю
    if i < 3: # 1, 2, 3 месяцы нет доходов
        rev_sec = 0.0
        rev_prim = 0.0
        rev_rent = 0.0
        rev_sub = 0.0
        rev_overseas = 0.0
        rev_other_p = 0.0
        rev_mort = 0.0
        rev_ins = 0.0
        rev_legal = 0.0
    else:
        rev_sec = base['rev_sec'][i] * overall_voeronka_factor * (comm_secondary / 360000.0)
        rev_prim = base['rev_prim'][i] * overall_voeronka_factor * (comm_primary / 440000.0)
        rev_rent = base['rev_rent'][i] * (comm_rent_val / 80000.0)
        rev_sub = base['rev_sub'][i] * (comm_suburban / 500000.0)
        rev_overseas = base['rev_overseas'][i]
        rev_other_p = base['rev_other_p'][i]
        rev_mort = base['rev_mort'][i] * overall_voeronka_factor * (comm_secondary / 360000.0)
        rev_ins = base['rev_ins'][i] * overall_voeronka_factor * (comm_secondary / 360000.0)
        rev_legal = base['rev_legal'][i] * overall_voeronka_factor * (comm_secondary / 360000.0)
        
    rev_secondary_list.append(rev_sec)
    rev_primary_list.append(rev_prim)
    rev_rent_list.append(rev_rent)
    rev_suburban_list.append(rev_sub)
    rev_overseas_list.append(rev_overseas)
    rev_other_p_list.append(rev_other_p)
    rev_mortgage_list.append(rev_mort)
    rev_insurance_list.append(rev_ins)
    rev_legal_list.append(rev_legal)
    
    total_revenue_month = rev_sec + rev_prim + rev_rent + rev_sub + rev_overseas + rev_other_p + rev_mort + rev_ins + rev_legal
    rev_total_list.append(total_revenue_month)
    
    # 3. РАСХОДЫ - ФОТ
    agent_payouts = base['agent_comm'][i] * (agent_commission_pct / 38.0) * overall_voeronka_factor
    payouts_agents_list.append(agent_payouts)
    
    salaries_back = base['salaries'][i] * (backoffice_salary_mult / 100.0) * salary_mult
    salaries_backoffice_list.append(salaries_back)
    
    taxes_payroll = base['salary_taxes'][i] * (backoffice_salary_mult / 100.0) * salary_mult
    taxes_payroll_list.append(taxes_payroll)
    
    total_payroll_list.append(agent_payouts + salaries_back + taxes_payroll)
    
    # 4. РАСХОДЫ - OPEX
    rent_list.append(office_rent_custom * (base_rent_list[i] / 150000.0))
    internet_list.append(base_internet[i])
    mobile_list.append(base_mobile[i])
    kanc_list.append(base_kanc[i])
    reklama_list.append(base_reklama[i] * leads_mult)
    hh_list.append(base_hh[i])
    buh_list.append(base_buh[i])
    bank_list.append(base_bank[i])
    cleaning_list.append(base_cleaning[i])
    gsm_list.append(base_gsm[i])
    courier_list.append(base_courier[i])
    events_list.append(base_events[i])
    
    total_opex = (office_rent_custom * (base_rent_list[i] / 150000.0) + base_internet[i] + base_mobile[i] + base_kanc[i] + 
                  base_reklama[i] * leads_mult + base_hh[i] + base_buh[i] + base_bank[i] + base_cleaning[i] + base_gsm[i] + 
                  base_courier[i] + base_events[i])
    total_opex_list.append(total_opex)
    
    # 5. РАСХОДЫ - ВЫПЛАТЫ ЦО (Франшиза)
    hq_fix = base['hq_royalty_fix'][i] * royalty_fix_mult
    hq_royalty_fix_list.append(hq_fix)
    
    hq_deal = base['hq_royalty_deal'][i] * (royalty_deal_custom / 2100.0) * overall_voeronka_factor
    hq_royalty_deal_list.append(hq_deal)
    
    # Разделение crm и других IT
    hq_crm = 5500 if i < 12 else 11000
    hq_crm_list.append(hq_crm)
    
    hq_kc_val = base['hq_crm_and_others'][i] - hq_crm
    # Очистка остатка
    if hq_kc_val < 0: hq_kc_val = 0
    hq_kc_list.append(hq_kc_val)
    
    total_hq_payments = hq_fix + hq_deal + hq_crm + hq_kc_val
    total_hq_payments_list.append(total_hq_payments)
    
    # 6. БИЗНЕС-НАЛОГИ
    tax_month = total_revenue_month * (tax_rate / 100.0)
    taxes_usn_list.append(tax_month)
    
    # 7. КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX) Месячный тотал
    tot_capex_month = capex_franchise_list[i] + capex_renovation_list[i] + capex_equipment_list[i]
    total_capex_list.append(tot_capex_month)
    
    # ИТОГО РАСХОДЫ Месяца (включая налоги)
    total_expenses_month = agent_payouts + salaries_back + taxes_payroll + total_opex + total_hq_payments + tax_month
    
    # Чистая прибыль за месяц
    net_profit_month = total_revenue_month - total_expenses_month - tot_capex_month
    adj_net_profit.append(net_profit_month)

# Расчет накопленного денежного потока начиная с Месяца 0 (CAPEX + OPEX старт)
capex_m0_franchise = regional_pau_fee + 17000 # паушальный + роспатент
capex_m0_renovation = 520000 if scenario_name == "Оптимальный" else 0
capex_m0_equipment = 815000 if scenario_name == "Оптимальный" else 0
total_capex_m0 = capex_m0_franchise + capex_m0_renovation + capex_m0_equipment

cum_flow = []
running_sum = -total_capex_m0 - start_opex
cum_flow.append(running_sum) # месяц 0

for i in range(24):
    running_sum += adj_net_profit[i]
    cum_flow.append(running_sum)

# ----------------------------------------------------
# KPI И ПОКАЗАТЕЛИ ЭФФЕКТИВНОСТИ
# ----------------------------------------------------
total_revenue_2years = sum(rev_total_list)
total_profit_2years = sum(adj_net_profit) - total_capex_m0 - start_opex
max_investment = abs(min(cum_flow))

# Месяц окупаемости (первый месяц, где cum_flow >= 0)
payback_month = "Н/Д"
for idx, val in enumerate(cum_flow):
    if val >= 0:
        payback_month = f"{idx} мес."
        break

# Точка безубыточности (первый месяц с положительной чистой прибылью)
breakeven_month = "Н/Д"
for idx, val in enumerate(adj_net_profit):
    if val > 0:
        breakeven_month = f"{idx + 1} мес."
        break

# Отображение KPI Карточек
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("💼 Общая выручка (2 года)", f"{total_revenue_2years:,.0f} ₽".replace(",", " "))
with col2:
    st.metric("📈 Итоговая чистая прибыль", f"{total_profit_2years:,.0f} ₽".replace(",", " "))
with col3:
    st.metric("📉 Пик инвестиций (CAPEX+OPEX)", f"{max_investment:,.0f} ₽".replace(",", " "))
with col4:
    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("🎯 Без-ость", breakeven_month)
    with col_b:
        st.metric("⏳ Окупаемость", payback_month)

# ----------------------------------------------------
# ТАБЫ: ГРАФИКИ, ДЕТАЛИЗАЦИЯ И СРАВНЕНИЕ
# ----------------------------------------------------
tab1, tab2, tab3, tab4 = st.tabs(["📊 Интерактивные Графики", "📋 Декомпозиция P&L (Все статьи)", "🌍 Сравнение Регионов", "📝 Описание Модели"])

with tab1:
    st.markdown("### Визуализация финансового развития")
    
    # 1. График накопленного денежного потока
    fig_cum = go.Figure()
    fig_cum.add_trace(go.Scatter(
        x=list(range(25)),
        y=cum_flow,
        mode='lines+markers',
        name='Накопленный кэш-флоу',
        line=dict(color='#BEAF87', width=3),
        marker=dict(size=6, color='#FFFFFF', line=dict(color='#BEAF87', width=1.5))
    ))
    fig_cum.add_shape(
        type="line", x0=0, y0=0, x1=24, y1=0,
        line=dict(color="#777779", width=1.5, dash="dash")
    )
    fig_cum.update_layout(
        title="Кривая накопленного кэш-флоу (J-Curve)",
        xaxis_title="Месяцы проекта (0 - старт)",
        yaxis_title="Баланс кэша, ₽",
        template="plotly_white",
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        font=dict(color="#252526")
    )
    st.plotly_chart(fig_cum, use_container_width=True)
    
    # 2. Столбчатый график Доходы vs Расходы
    fig_rev_exp = go.Figure()
    fig_rev_exp.add_trace(go.Bar(
        x=months,
        y=rev_total_list,
        name='Доходы (Комиссия ВКД)',
        marker_color='#BEAF87'
    ))
    
    expenses_no_capex = [rev_total_list[idx] - adj_net_profit[idx] - total_capex_list[idx] for idx in range(24)]
    fig_rev_exp.add_trace(go.Bar(
        x=months,
        y=expenses_no_capex,
        name='Операционные расходы + Налоги',
        marker_color='#252526'
    ))
    fig_rev_exp.update_layout(
        title="Ежемесячные доходы и операционные расходы",
        xaxis_title="Месяцы",
        yaxis_title="Сумма, ₽",
        barmode='group',
        template="plotly_white",
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        font=dict(color="#252526")
    )
    st.plotly_chart(fig_rev_exp, use_container_width=True)

    # 3. Круговая диаграмма структуры расходов
    avg_expenses = {
        'Выплаты агентам': sum(payouts_agents_list),
        'Оклады бэк-офиса': sum(salaries_backoffice_list),
        'Налоги на ФОТ оклады': sum(taxes_payroll_list),
        'Аренда офиса': sum(rent_list),
        'Реклама и маркетинг': sum(reklama_list),
        'HH.ru и подбор': sum(hh_list),
        'Роялти и ЦО': sum(total_hq_payments_list),
        'УСН налоги': sum(taxes_usn_list),
        'Прочие OPEX': sum(internet_list) + sum(mobile_list) + sum(kanc_list) + sum(buh_list) + sum(bank_list) + sum(cleaning_list) + sum(gsm_list) + sum(courier_list) + sum(events_list)
    }
    fig_pie = px.pie(
        names=list(avg_expenses.keys()),
        values=list(avg_expenses.values()),
        color_discrete_sequence=['#BEAF87', '#A19276', '#252526', '#777779', '#CCCCCC', '#E6E7E8', '#999999', '#BEAF87', '#EAEAEA'],
        title="Детализированная структура расходов за 2 года"
    )
    fig_pie.update_layout(
        template="plotly_white",
        paper_bgcolor="#FFFFFF",
        font=dict(color="#252526")
    )
    st.plotly_chart(fig_pie, use_container_width=True)

with tab2:
    st.markdown("### 📋 Детализированный отчет о прибылях и убытках (P&L)")
    
    # Построение иерархической таблицы данных
    p_and_l_decomposed = {
        'Категория / Статья': [
            '► ДОХОДЫ (ВКД)',
            '  ├─ Вторичный рынок',
            '  ├─ Первичный рынок',
            '  ├─ Аренда (жилая/коммерческая)',
            '  ├─ Загородная недвижимость',
            '  ├─ Зарубежная недвижимость',
            '  ├─ Сделки: прочее (МЛС, срочновыкуп)',
            '  ├─ Сервисы: ипотека',
            '  ├─ Сервисы: страхование',
            '  └─ Сервисы: юр. сопровождение',
            '★ ИТОГО ДОХОДЫ',
            
            '► РАСХОДЫ: ФОТ',
            '  ├─ Выплаты агентам (% комиссионных)',
            '  ├─ Оклады бэк-офиса (оклады)',
            '  └─ Налоги на ФОТ оклады',
            '★ ИТОГО ФОТ',
            
            '► РАСХОДЫ: OPEX',
            '  ├─ Аренда офиса',
            '  ├─ Интернет',
            '  ├─ Сотовая связь',
            '  ├─ Канцелярия',
            '  ├─ Реклама объектов',
            '  ├─ HeadHunter.ru',
            '  ├─ Бухгалтерия: аутсорс',
            '  ├─ Услуги банка',
            '  ├─ Уборка офиса',
            '  ├─ ГСМ',
            '  ├─ Доставка/курьер',
            '  └─ Корпоративные мероприятия',
            '★ ИТОГО OPEX',
            
            '► ВЫПЛАТЫ ЦО (ФРАНШИЗА)',
            '  ├─ Роялти FIX (вкл. НРФ)',
            '  ├─ Роялти со сделок',
            '  ├─ CRM-система',
            '  └─ Колл-центр и прочие сервисы',
            '★ ИТОГО ВЫПЛАТЫ ЦО',
            
            '★ НАЛОГ УСН',
            '★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ',
            
            '► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)',
            '  ├─ Франшиза (Паушальный взнос + Роспатент)',
            '  ├─ Ремонт и Брендирование офиса',
            '  └─ Мебель, компьютеры и оборудование',
            '★ ИТОГО CAPEX',
            
            '★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ',
            '★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК'
        ]
    }
    
    # Заполняем таблицу по месяцам
    for idx, m in enumerate(months):
        p_and_l_decomposed[m] = [
            "", # Заголовок доходов
            f"{rev_secondary_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_primary_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_rent_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_suburban_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_overseas_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_other_p_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_mortgage_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_insurance_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_legal_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_total_list[idx]:,.0f} ₽".replace(",", " "),
            
            "", # Заголовок ФОТ
            f"{payouts_agents_list[idx]:,.0f} ₽".replace(",", " "),
            f"{salaries_backoffice_list[idx]:,.0f} ₽".replace(",", " "),
            f"{taxes_payroll_list[idx]:,.0f} ₽".replace(",", " "),
            f"{total_payroll_list[idx]:,.0f} ₽".replace(",", " "),
            
            "", # Заголовок OPEX
            f"{rent_list[idx]:,.0f} ₽".replace(",", " "),
            f"{internet_list[idx]:,.0f} ₽".replace(",", " "),
            f"{mobile_list[idx]:,.0f} ₽".replace(",", " "),
            f"{kanc_list[idx]:,.0f} ₽".replace(",", " "),
            f"{reklama_list[idx]:,.0f} ₽".replace(",", " "),
            f"{hh_list[idx]:,.0f} ₽".replace(",", " "),
            f"{buh_list[idx]:,.0f} ₽".replace(",", " "),
            f"{bank_list[idx]:,.0f} ₽".replace(",", " "),
            f"{cleaning_list[idx]:,.0f} ₽".replace(",", " "),
            f"{gsm_list[idx]:,.0f} ₽".replace(",", " "),
            f"{courier_list[idx]:,.0f} ₽".replace(",", " "),
            f"{events_list[idx]:,.0f} ₽".replace(",", " "),
            f"{total_opex_list[idx]:,.0f} ₽".replace(",", " "),
            
            "", # Заголовок франшизы
            f"{hq_royalty_fix_list[idx]:,.0f} ₽".replace(",", " "),
            f"{hq_royalty_deal_list[idx]:,.0f} ₽".replace(",", " "),
            f"{hq_crm_list[idx]:,.0f} ₽".replace(",", " "),
            f"{hq_kc_list[idx]:,.0f} ₽".replace(",", " "),
            f"{total_hq_payments_list[idx]:,.0f} ₽".replace(",", " "),
            
            f"{taxes_usn_list[idx]:,.0f} ₽".replace(",", " "),
            f"{(total_payroll_list[idx] + total_opex_list[idx] + total_hq_payments_list[idx] + taxes_usn_list[idx]):,.0f} ₽".replace(",", " "),
            
            "", # Заголовок CAPEX
            f"{capex_franchise_list[idx]:,.0f} ₽".replace(",", " "),
            f"{capex_renovation_list[idx]:,.0f} ₽".replace(",", " "),
            f"{capex_equipment_list[idx]:,.0f} ₽".replace(",", " "),
            f"{total_capex_list[idx]:,.0f} ₽".replace(",", " "),
            
            f"{adj_net_profit[idx]:,.0f} ₽".replace(",", " "),
            f"{cum_flow[idx+1]:,.0f} ₽".replace(",", " ")
        ]
        
    df_p_and_l_decomposed = pd.DataFrame(p_and_l_decomposed)
    st.dataframe(df_p_and_l_decomposed, use_container_width=True, height=600)
    
    # ----------------------------------------------------
    # EXCEL EXPORT С ПОЛНОЙ ДЕКОМПОЗИЦИЕЙ
    # ----------------------------------------------------
    if openpyxl_available:
        def generate_excel():
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "C21 Decomposed P&L"
            
            ws.views.sheetView[0].showGridLines = True
            
            # Цветовые схемы Century 21
            gold_fill = PatternFill(start_color="BEAF87", end_color="BEAF87", fill_type="solid")
            dark_fill = PatternFill(start_color="252526", end_color="252526", fill_type="solid")
            light_gold_fill = PatternFill(start_color="F9F6EE", end_color="F9F6EE", fill_type="solid")
            category_fill = PatternFill(start_color="F1F1F3", end_color="F1F1F3", fill_type="solid")
            
            font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_bold = Font(name="Calibri", size=11, bold=True)
            font_regular = Font(name="Calibri", size=11)
            
            thin_border_side = Side(border_style="thin", color="D3D3D3")
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            double_bottom_border = Border(bottom=Side(border_style="double", color="111111"), top=thin_border_side)
            
            # Заголовок листа
            ws.merge_cells("A1:Y1")
            ws["A1"] = f"ДЕТАЛИЗИРОВАННАЯ ФИНАНСОВАЯ МОДЕЛЬ CENTURY 21 - РЕГИОН: {region_select.upper()} ({scenario_name.upper()})"
            ws["A1"].font = font_title
            ws["A1"].fill = dark_fill
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40
            
            # Столбцы
            headers = ["Показатель / Месяцы"] + months
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = gold_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[3].height = 28
            
            # Месячный экспорт строк
            metrics_rows = [
                ('► ДОХОДЫ (ВКД)', [""], ""),
                ('  ├─ Вторичный рынок', rev_secondary_list, "#,##0"),
                ('  ├─ Первичный рынок', rev_primary_list, "#,##0"),
                ('  ├─ Аренда (жилая/коммерческая)', rev_rent_list, "#,##0"),
                ('  ├─ Загородная недвижимость', rev_suburban_list, "#,##0"),
                ('  ├─ Зарубежная недвижимость', rev_overseas_list, "#,##0"),
                ('  ├─ Сделки: прочее (МЛС, срочновыкуп)', rev_other_p_list, "#,##0"),
                ('  ├─ Сервисы: ипотека', rev_mortgage_list, "#,##0"),
                ('  ├─ Сервисы: страхование', rev_insurance_list, "#,##0"),
                ('  └─ Сервисы: юр. сопровождение', rev_legal_list, "#,##0"),
                ('★ ИТОГО ДОХОДЫ', rev_total_list, "#,##0"),
                
                ('► РАСХОДЫ: ФОТ', [""], ""),
                ('  ├─ Выплаты агентам (% комиссионных)', payouts_agents_list, "#,##0"),
                ('  ├─ Оклады бэк-офиса (оклады)', salaries_backoffice_list, "#,##0"),
                ('  └─ Налоги на ФОТ оклады', taxes_payroll_list, "#,##0"),
                ('★ ИТОГО ФОТ', total_payroll_list, "#,##0"),
                
                ('► РАСХОДЫ: OPEX', [""], ""),
                ('  ├─ Аренда офиса', rent_list, "#,##0"),
                ('  ├─ Интернет', internet_list, "#,##0"),
                ('  ├─ Сотовая связь', mobile_list, "#,##0"),
                ('  ├─ Канцелярия', kanc_list, "#,##0"),
                ('  ├─ Реклама объектов', reklama_list, "#,##0"),
                ('  ├─ HeadHunter.ru', hh_list, "#,##0"),
                ('  ├─ Бухгалтерия: аутсорс', buh_list, "#,##0"),
                ('  ├─ Услуги банка', bank_list, "#,##0"),
                ('  ├─ Уборка офиса', cleaning_list, "#,##0"),
                ('  ├─ ГСМ', gsm_list, "#,##0"),
                ('  ├─ Доставка/курьер', courier_list, "#,##0"),
                ('  └─ OPEX: Корпоративы', events_list, "#,##0"),
                ('★ ИТОГО OPEX', total_opex_list, "#,##0"),
                
                ('► ВЫПЛАТЫ ЦО (ФРАНШИЗА)', [""], ""),
                ('  ├─ Роялти FIX (вкл. НРФ)', hq_royalty_fix_list, "#,##0"),
                ('  ├─ Роялти со сделок', hq_royalty_deal_list, "#,##0"),
                ('  ├─ CRM-система', hq_crm_list, "#,##0"),
                ('  └─ Колл-центр и прочие сервисы', hq_kc_list, "#,##0"),
                ('★ ИТОГО ВЫПЛАТЫ ЦО', total_hq_payments_list, "#,##0"),
                
                ('★ НАЛОГ УСН', taxes_usn_list, "#,##0"),
                ('★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ', [total_payroll_list[idx]+total_opex_list[idx]+total_hq_payments_list[idx]+taxes_usn_list[idx] for idx in range(24)], "#,##0"),
                
                ('► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)', [""], ""),
                ('  ├─ Франшиза (Паушальный взнос + Роспатент)', capex_franchise_list, "#,##0"),
                ('  ├─ Ремонт и Брендирование офиса', capex_renovation_list, "#,##0"),
                ('  └─ Мебель, компьютеры и оборудование', capex_equipment_list, "#,##0"),
                ('★ ИТОГО CAPEX', total_capex_list, "#,##0"),
                
                ('★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ', adj_net_profit, "#,##0"),
                ('★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК', cum_flow[1:], "#,##0")
            ]
            
            for r_idx, (m_name, data_list, num_format) in enumerate(metrics_rows, 4):
                is_header = "►" in m_name
                is_total = "★" in m_name
                
                cell_a = ws.cell(row=r_idx, column=1, value=m_name)
                cell_a.font = font_bold if (is_header or is_total) else font_regular
                cell_a.border = thin_border
                
                if is_header:
                    cell_a.fill = category_fill
                elif is_total:
                    cell_a.fill = light_gold_fill
                
                for c_idx in range(1, 25):
                    if is_header:
                        cell_val = ws.cell(row=r_idx, column=c_idx+1, value="")
                        cell_val.fill = category_fill
                        cell_val.border = thin_border
                        continue
                        
                    val = data_list[c_idx-1] if len(data_list) >= c_idx else 0
                    cell_val = ws.cell(row=r_idx, column=c_idx+1, value=val)
                    cell_val.number_format = num_format
                    cell_val.alignment = Alignment(horizontal="right")
                    cell_val.border = thin_border
                    
                    if is_total:
                        cell_val.font = font_bold
                        cell_val.fill = light_gold_fill
                        if "ЧИСТАЯ ПРИБЫЛЬ" in m_name:
                            cell_val.border = double_bottom_border
                    else:
                        cell_val.font = font_regular
                
                ws.row_dimensions[r_idx].height = 20
                
            ws.column_dimensions["A"].width = 45
            for col in range(2, 26):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 15
                
            wb.save(output)
            return output.getvalue()
            
        excel_data = generate_excel()
        st.download_button(
            label="📥 Скачать детализированный P&L в Excel",
            data=excel_data,
            file_name=f"C21_P_and_L_Decomposed_{region_select}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("⚠️ **Экспорт в Excel временно недоступен.**")

with tab3:
    st.markdown("### 🌍 Сравнение ключевых показателей регионов")
    st.write("Сравнение базовых финансово-экономических параметров по региональным сегментам Century 21:")
    
    comp_data = {
        'Региональный сегмент': list(REGIONAL_PRESETS.keys()),
        'Паушальный взнос (база)': [f"{v['pau_fee']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Вторичка': [f"{v['comm_sec']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Первичка': [f"{v['comm_prim']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Загородная': [f"{v['comm_sub']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Оклады бэк-офиса': [f"{v['salary_mult']*100:.0f}% к Москве" for v in REGIONAL_PRESETS.values()],
        'Аренда офиса (база)': [f"{v['rent_base']:,} ₽/мес.".replace(",", " ") for v in REGIONAL_PRESETS.values()],
    }
    df_comp = pd.DataFrame(comp_data)
    st.table(df_comp)

with tab4:
    st.markdown("""
    ### 📖 Справочник и методология декомпозиции
    
    Эта интерактивная модель разделяет финансовый план агентства недвижимости **Century 21** на 5 ключевых блоков:
    
    1. **Доходы (Комиссия ВКД):** 
       * Рассчитывается динамически на основе воронки лидов. 
       * Включает доп. сервисы — **комиссионные вознаграждения от банков по ипотеке, страховых компаний и юр. услуг**.
       
    2. **ФОТ (Зарплатный фонд):**
       * **Переменная часть (агенты):** Прямо привязана к объему сделок и настраивается слайдером (базово 38%).
       * **Постоянная часть (оклады):** Включает фиксированные оклады РОПа, HR, юристов, рекрутеров, скорректированные под выбранный регион.
       
    3. **OPEX (Операционные расходы):**
       * Включает аренду офиса, затраты на лидогенерацию (маркетинг), сервисы по набору персонала (HeadHunter), а также жизнеобеспечение (уборка, интернет, связь, канцелярия).
       
    4. **Выплаты ЦО (Франшиза):**
       * Фиксированные роялти (шкала по месяцам), переменные роялти с каждой сделки (настраиваются слайдером) и IT-сервисы ЦО.
       
    5. **CAPEX (Капитальные вложения):**
       * Регистрационные расходы, запуск офиса (ремонт, мебель, компьютеры) на старте и календарное расширение рабочих мест во второй год.
    """)
