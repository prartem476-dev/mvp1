import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import io

# Проверяем доступность openpyxl для экспорта в Excel без падения приложения
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    openpyxl_available = True
except ImportError:
    openpyxl_available = False

# Установка конфигурации страницы
st.set_page_config(
    page_title="CENTURY 21 Financial Model (Бюджет Оптимальный) v8",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ----------------------------------------------------
# ПОМОЩНИКИ ДЛЯ РУЧНЫХ КОРРЕКТИРОВОК P&L
# ----------------------------------------------------
if 'pl_overrides' not in st.session_state:
    st.session_state['pl_overrides'] = {}

def get_override(metric_name, col_name, default_val):
    key = (metric_name, col_name)
    if key in st.session_state['pl_overrides']:
        return st.session_state['pl_overrides'][key]
    return default_val

def parse_value(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(' ', '').replace('₽', '').replace(',', '.').strip()
    if s == "" or s == "-" or s == "0":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


# ----------------------------------------------------
# АВТОРИЗАЦИЯ ПОЛЬЗОВАТЕЛЯ
# ----------------------------------------------------
if 'authenticated' not in st.session_state:
    st.session_state['authenticated'] = False

if not st.session_state['authenticated']:
    # Beautiful custom styling for login
    st.markdown("""
    <style>
        /* Импорт шрифта */
        @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@300;400;600;700;800&display=swap');
        
        html, body, [data-testid="stAppViewContainer"] {
            font-family: 'Segoe UI', Arial, sans-serif;
            background-color: #FFFFFF;
            color: #252526;
        }
        
        .login-title {
            color: #252526;
            font-size: 26px;
            font-weight: 800;
            text-align: center;
            margin-top: 10px;
            margin-bottom: 5px;
            text-transform: uppercase;
            letter-spacing: 2px;
        }
        .brand-accent {
            color: #BEAF87; /* Relentless Gold */
        }
        .login-sub {
            color: #808285; /* Medium Grey */
            font-size: 14px;
            text-align: center;
            margin-bottom: 30px;
        }
        div.stButton > button {
            background-color: #BEAF87 !important;
            color: #FFFFFF !important;
            font-weight: bold !important;
            border: none !important;
            border-radius: 5px !important;
            padding: 12px 24px !important;
            transition: all 0.3s ease !important;
            width: 100% !important;
            box-shadow: 0 2px 5px rgba(190, 175, 135, 0.3) !important;
        }
        div.stButton > button:hover {
            background-color: #A19276 !important; /* Dark Gold */
            box-shadow: 0 4px 12px rgba(161, 146, 118, 0.5) !important;
            transform: translateY(-1px);
        }
        
        /* Контейнер для монограммы с соблюдением охранного поля 0.5X */
        .brand-monogram-container {
            display: block;
            margin: 35px auto; /* Охранное поле 0.5X для монограммы высотой 70px */
            text-align: center;
        }
    </style>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1.4, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        # Встраиваем монограмму C21 согласно Brandbook (в один цвет Relentless Gold, с охранным полем 0.5X)
        st.markdown("""
        <div class="brand-monogram-container">
            <svg viewBox="0 0 100 100" width="70" height="70" style="display: block; margin: 0 auto;">
                <path d="M 75,25 A 35,35 0 1,0 75,75" fill="none" stroke="#BEAF87" stroke-width="8" stroke-linecap="round" />
                <text x="48" y="59" font-family="'Segoe UI', Arial, sans-serif" font-weight="bold" font-size="26" fill="#BEAF87" text-anchor="middle">21</text>
            </svg>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div class='login-title'>CENTURY 21 <span class='brand-accent'>Financial MVP</span></div>", unsafe_allow_html=True)
        st.markdown("<div class='login-sub'>Интерактивная финансовая модель | Кейс: Бюджет Оптимальный</div>", unsafe_allow_html=True)
        
        with st.container(border=True):
            st.markdown("<h4 style='text-align: center; color: #252526; font-weight: 700; margin-bottom: 20px;'>Вход в личный кабинет</h4>", unsafe_allow_html=True)
            username = st.text_input("Логин", placeholder="Введите имя пользователя")
            password = st.text_input("Пароль", type="password", placeholder="Введите ваш пароль")
            
            st.markdown("<div style='margin-top: 20px;'>", unsafe_allow_html=True)
            if st.button("Войти в систему"):
                if username == "admin" and password == "c21admin":
                    st.session_state['authenticated'] = True
                    st.success("Успешная авторизация! Загрузка модели...")
                    st.rerun()
                else:
                    st.error("Неверный логин или пароль. Попробуйте еще раз.")
            st.markdown("</div>", unsafe_allow_html=True)
        st.stop()

# Пользовательский CSS-стилинг в корпоративных цветах Century 21 (Premium Light Theme)
st.markdown("""
<style>
    /* Импорт шрифта */
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@300;400;600;700;800&display=swap');
    
    html, body, [data-testid="stAppViewContainer"] {
        font-family: 'Segoe UI', Arial, sans-serif;
        background-color: #FFFFFF;
        color: #252526;
    }
    
    /* Стилизация боковой панели */
    [data-testid="stSidebar"] {
        background-color: #E6E7E8 !important;
        border-right: 1px solid #D3D3D3;
        padding-top: 20px;
    }
    
    [data-testid="stSidebar"] * {
        color: #252526 !important;
    }
    
    /* Красивые карточки KPI */
    div[data-testid="stMetric"] {
        background-color: #F8F9FA;
        border: 1px solid #E6E7E8;
        border-left: 5px solid #BEAF87;
        padding: 15px 20px;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    
    div[data-testid="stMetric"] label {
        color: #777779 !important;
        font-size: 13px !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        color: #252526 !important;
        font-size: 22px !important;
        font-weight: 700 !important;
        margin-top: 5px;
    }
    
    /* Кнопки */
    div.stButton > button {
        background-color: #BEAF87 !important;
        color: #FFFFFF !important;
        font-weight: bold !important;
        border: none !important;
        border-radius: 5px !important;
        padding: 12px 24px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 2px 5px rgba(190, 175, 135, 0.3) !important;
        width: 100% !important;
    }
    
    div.stButton > button:hover {
        background-color: #A19276 !important;
        color: #FFFFFF !important;
        box-shadow: 0 4px 12px rgba(161, 146, 118, 0.5) !important;
        transform: translateY(-1px);
    }
    
    /* Заголовки */
    .main-title {
        color: #252526;
        font-size: 30px;
        font-weight: 800;
        text-align: center;
        margin-top: 10px;
        margin-bottom: 5px;
        letter-spacing: 0.5px;
    }
    
    .brand-accent {
        color: #A19276;
    }
    
    .sub-title {
        color: #777779;
        font-size: 15px;
        text-align: center;
        margin-bottom: 30px;
        font-weight: 400;
    }
    
    /* Вкладки */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        border-bottom: 2px solid #E6E7E8;
    }
    
    .stTabs [data-baseweb="tab"] {
        background-color: #F8F9FA;
        border: 1px solid #E6E7E8;
        border-bottom: none;
        border-radius: 6px 6px 0px 0px;
        padding: 10px 24px;
        color: #777779;
        font-weight: 600;
        transition: all 0.2s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #FFFFFF !important;
        color: #A19276 !important;
        border-top: 3px solid #BEAF87 !important;
        border-left: 1px solid #BEAF87 !important;
        border-right: 1px solid #BEAF87 !important;
    }
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# ОПРЕДЕЛЕНИЕ ДАННЫХ ЛИСТА "БЮДЖЕТ ОПТИМАЛЬНЫЙ" (C21)
# ----------------------------------------------------

months = [f"Месяц {i}" for i in range(1, 25)]

optimal_baseline = {
    'leads': [440, 880, 1320, 1760, 2200, 2640, 3080, 3520, 3960, 4400, 4840, 5280,
              6160, 6600, 7040, 7480, 7920, 8360, 8800, 9240, 9680, 10120, 10560, 11000],
    'meetings': [44, 88, 132, 176, 220, 264, 308, 352, 396, 440, 484, 528,
                 616, 660, 704, 748, 792, 836, 880, 924, 968, 1012, 1056, 1100],
    'contracts': [0, 2, 4, 7, 9, 11, 13, 15, 18, 20, 22, 24,
                  26, 31, 33, 35, 37, 40, 42, 44, 46, 48, 51, 53],
    'prepayments': [0.0, 1.5, 3.0, 5.3, 6.8, 8.3, 9.8, 11.3, 13.5, 15.0, 16.5, 18.0,
                    19.5, 23.3, 24.8, 26.3, 27.8, 30.0, 31.5, 33.0, 34.5, 36.0, 38.3, 39.8],
    
    # Декомпозиция ДОХОДОВ (ВКД)
    'rev_sec': [0, 0, 0, 360000, 1080000, 1603800, 2138400, 2673000, 3207600, 3742200, 4276800, 4811400,
                5346000, 5880600, 6415200, 7484400, 8019000, 8553600, 9088200, 9622800, 10157400, 10692000, 11226600, 11761200],
    'rev_prim': [0, 0, 0, 440000, 440000, 440000, 440000, 880000, 880000, 880000, 880000, 880000,
                 653400, 718740, 784080, 914760, 980100, 1045440, 1110780, 1176120, 1241460, 1306800, 1372140, 1437480],
    'rev_rent': [0, 0, 0, 0, 0, 80000, 0, 0, 80000, 0, 0, 80000,
                 240000, 240000, 280000, 280000, 320000, 320000, 360000, 400000, 440000, 440000, 480000, 480000],
    'rev_sub': [0, 0, 0, 0, 0, 500000, 0, 0, 500000, 0, 0, 500000,
                1000000, 1000000, 1000000, 1000000, 1500000, 1500000, 2000000, 2000000, 2500000, 2500000, 2500000, 2500000],
    'rev_overseas': [0]*17 + [220000] + [0]*4 + [220000] + [0],
    'rev_other_p': [0]*12 + [75600]*12,
    'rev_mort': [0]*8 + [106029, 123701, 141372, 159044, 176715, 194387, 212058, 247401, 265073, 282744, 300416, 318087, 335759, 353430, 371102, 388773],
    'rev_ins': [0]*8 + [27265, 31809, 36353, 40897, 45441, 49985, 54529, 63617, 68162, 72706, 77250, 81794, 86338, 90882, 95426, 99970],
    'rev_legal': [0]*6 + [138600, 161700, 184800, 207900, 231000, 254100, 277200, 323400, 346500, 369600, 392700, 415800, 438900, 462000, 485100, 508200, 531300, 554400],
    'revenue': [0, 0, 0, 800000, 1520000, 2623800, 2717000, 3714700, 4985694, 4985609, 5565525, 6725440,
                7814356, 8482712, 9167967, 10435378, 11620634, 12485890, 13451145, 14136401, 15321656, 15966912, 16872168, 17297423],

    # Декомпозиция РАСХОДОВ - ФОТ
    'salaries': [214935, 214935, 214935, 214935, 214935, 144935, 144935, 144935, 144935, 144935, 144935, 144935,
                 963068, 964023, 966913, 1126564, 1129454, 1462344, 1608235, 1611125, 1764015, 1766905, 1769796, 1772686],
    'salary_taxes': [44935]*12 + [143068, 144023, 146913, 156564, 159454, 162344, 208235, 211125, 214015, 216905, 219796, 222686],
    'agent_comm': [0, 0, 0, 338500, 612100, 1053409, 1314265, 1784222, 2332633, 2318694, 2629481, 3171647,
                   4103166, 4502709, 4855674, 5449734, 6035280, 6593359, 7120893, 7473269, 8061122, 8387655, 8822681, 9066321],

    # Декомпозиция OPEX
    'opex': [150000, 286300, 321300, 357400, 393500, 464600, 535700, 606800, 677900, 749000, 820100, 896200,
             569750, 601530, 769250, 1136970, 809690, 827910, 851130, 868850, 892070, 1399290, 927010, 944730],

    # Декомпозиция ВЫПЛАТ ЦО
    'hq_royalty_fix': [0, 0, 0, 45675, 45675, 45675, 45675, 78750, 78750, 78750, 78750, 78750,
                       111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825],
    'hq_royalty_deal': [0, 0, 0, 4200, 8400, 13556, 14574, 19793, 25011, 26030, 29148, 34367,
                        39134, 42564, 45994, 52855, 58385, 61816, 67346, 70776, 76307, 79737, 83167, 86598],
    'hq_crm_and_others': [19400, 5500, 5500, 68375, 72575, 77731, 78749, 117043, 188908, 201034, 215260, 231587,
                          194303, 191657, 195549, 202872, 208864, 278757, 218749, 222641, 228634, 232526, 302418, 240311],
    'capex': [0]*24
}

# Региональные параметры и пресеты
REGIONAL_PRESETS = {
    "Москва": {
        "pau_fee": 577500,
        "comm_sec": 360000,
        "comm_prim": 440000,
        "comm_sub": 500000,
        "comm_rent": 80000,
        "rent_base": 150000,
        "salary_mult": 1.0,
        "royalty_fix_factor": 1.0,
        "royalty_deal": 2100,
    },
    "Московская область (МО)": {
        "pau_fee": 500000,
        "comm_sec": 280000,
        "comm_prim": 330000,
        "comm_sub": 400000,
        "comm_rent": 60000,
        "rent_base": 100000,
        "salary_mult": 0.85,
        "royalty_fix_factor": 0.85,
        "royalty_deal": 2000,
    },
    "Крупные города (население > 800к)": {
        "pau_fee": 450000,
        "comm_sec": 220000,
        "comm_prim": 260000,
        "comm_sub": 320000,
        "comm_rent": 45000,
        "rent_base": 80000,
        "salary_mult": 0.75,
        "royalty_fix_factor": 0.75,
        "royalty_deal": 2000,
    },
    "Средние города (население 400к - 800к)": {
        "pau_fee": 350000,
        "comm_sec": 160000,
        "comm_prim": 200000,
        "comm_sub": 240000,
        "comm_rent": 35000,
        "rent_base": 55000,
        "salary_mult": 0.60,
        "royalty_fix_factor": 0.60,
        "royalty_deal": 1800,
    },
    "Малые города (население < 400к)": {
        "pau_fee": 270000,
        "comm_sec": 110000,
        "comm_prim": 140000,
        "comm_sub": 180000,
        "comm_rent": 25000,
        "rent_base": 38000,
        "salary_mult": 0.45,
        "royalty_fix_factor": 0.45,
        "royalty_deal": 1500,
    }
}

# ----------------------------------------------------
# ШАПКА ПРИЛОЖЕНИЯ
# ----------------------------------------------------
st.markdown("<div class='main-title'>💼 CENTURY 21 <span class='brand-accent'>Optimal Budget Case</span></div>", unsafe_allow_html=True)
st.markdown("<div class='sub-title'>Интерактивная финансовая модель и декомпозиция доходов/расходов на основе листа «Бюджет Оптимальный»</div>", unsafe_allow_html=True)

# ----------------------------------------------------
# БОКОВАЯ ПАНЕЛЬ С УПРАВЛЕНИЕМ & ТУЛТИПАМИ
# ----------------------------------------------------
st.sidebar.markdown("""
<div style='text-align: center; margin: 30px auto; display: block;'>
    <!-- Монограмма C21 в один цвет (Obsessed Grey) согласно Brandbook -->
    <svg viewBox="0 0 100 100" width="60" height="60" style="display: block; margin: 0 auto;">
        <path d="M 75,25 A 35,35 0 1,0 75,75" fill="none" stroke="#252526" stroke-width="8" stroke-linecap="round" />
        <text x="48" y="59" font-family="'Segoe UI', Arial, sans-serif" font-weight="bold" font-size="26" fill="#252526" text-anchor="middle">21</text>
    </svg>
</div>
""", unsafe_allow_html=True)

st.sidebar.markdown("<h2 style='color:#252526; text-align:center; font-weight:800; margin-top: 10px;'>Управление моделью</h2>", unsafe_allow_html=True)

if st.sidebar.button("🔓 Выйти из системы", use_container_width=True):
    st.session_state['authenticated'] = False
    st.rerun()

st.sidebar.markdown("---")

# Регион
region_select = st.sidebar.selectbox(
    "🌍 Выберите регион / тип города:",
    options=list(REGIONAL_PRESETS.keys()),
    help="Каждый пресет автоматически настраивает средние чеки сделок, стоимость аренды офиса, оклады бэк-офиса и fixed-роялти франшизы под экономический уровень выбранной территории."
)
preset = REGIONAL_PRESETS[region_select]

# Фиксированный сценарий "Бюджет Оптимальный"
st.sidebar.info("📌 **Выбран кейс: Бюджет Оптимальный (Интенсивный старт)**\n\nДанный кейс моделирует агрессивный запуск с первого месяца: в Месяце 0 производится закупка оборудования и ремонт офиса на сумму 1 335 000 ₽, а воронка масштабируется на базе 440 базовых лидов на старте с выходом на 11 000 лидов к 24 месяцу.")

# --- РЕГУЛИРОВКА ПОКАЗАТЕЛЕЙ ПО РАЗДЕЛАМ (Сворачиваемые экспандеры в Sidebar) ---

# 1. Воронка продаж
with st.sidebar.expander("📈 1. Воронка продаж", expanded=True):
    scaling_leads = st.slider(
        "Количество лидов (звонков), % к базе",
        min_value=50, max_value=200, value=100, step=10,
        help="Регулирует входящий поток клиентов. Масштабирует доходы от сделок по вторичному и первичному рынку."
    )
    conversion_meeting = st.slider(
        "Конверсия лид -> встреча, %",
        min_value=2.0, max_value=20.0, value=10.0, step=0.5,
        help="Какая часть позвонивших клиентов доходит до личной встречи в офисе. Базовый корпоративный стандарт Century 21 составляет 10%."
    )
    conversion_deal = st.slider(
        "Конверсия встреча -> договор, %",
        min_value=1.0, max_value=30.0, value=5.0, step=0.5,
        help="Эффективность работы агентов при переговорах. Показывает долю встреч, переходящих в подписанные эксклюзивные договоры. Базовый уровень для оптимального бюджета равен 5%."
    )

# 2. Средние чеки (Комиссии)
with st.sidebar.expander("💰 2. Средние комиссии", expanded=False):
    comm_secondary = st.number_input(
        "Вторичный рынок, ₽", min_value=50000, max_value=1000000, value=preset["comm_sec"], step=10000,
        help="Средний комиссионный доход агентства с одной сделки купли-продажи на вторичном рынке недвижимости."
    )
    comm_primary = st.number_input(
        "Первичный рынок, ₽", min_value=50000, max_value=1000000, value=preset["comm_prim"], step=10000,
        help="Средняя комиссия, получаемая от девелоперов за продажу новостройки."
    )
    comm_suburban = st.number_input(
        "Загородная недвижимость, ₽", min_value=50000, max_value=1000000, value=preset["comm_sub"], step=10000,
        help="Средний комиссионный чек со сделок по загородным домам, коттеджам и участкам."
    )
    comm_rent_val = st.number_input(
        "Аренда (жилая/коммерческая), ₽", min_value=10000, max_value=200000, value=preset["comm_rent"], step=5000,
        help="Комиссионный доход со сделок аренды жилой и коммерческой недвижимости."
    )

# 3. ФОТ и Выплаты
with st.sidebar.expander("👥 3. Персонал и Выплаты", expanded=False):
    agent_commission_pct = st.slider(
        "Средний % выплат агентам", min_value=25, max_value=60, value=38, step=1,
        help="Средний процент от комиссии по сделкам, выплачиваемый агентам. Отражает мотивационную сетку Century 21 (35%-50% в зависимости от категории)."
    )
    backoffice_salary_mult = st.slider(
        "Индекс окладов бэк-офиса, % к базе", min_value=50, max_value=150, value=100, step=5,
        help="Шкала фиксированных окладов бэк-офиса (РОП, HR/рекрутер, юрист, ипотечный брокер, листинг-менеджер, администратор)."
    )

# 4. OPEX (Аренда и операционные затраты)
with st.sidebar.expander("🏢 4. OPEX и Налоги", expanded=False):
    office_rent_custom = st.number_input(
        "Аренда офиса в месяц, ₽", min_value=10000, max_value=500000, value=preset["rent_base"], step=5000,
        help="Стоимость аренды офисного помещения. Применяются поправки на ранних этапах запуска офиса."
    )
    tax_rate = st.sidebar.slider(
        "Ставка налога УСН, %", min_value=3, max_value=15, value=7, step=1,
        help="Действующая налоговая ставка по УСН (доходы) с учетом региональных субсидий и льгот."
    )
    marketing_mult = st.slider(
        "Реклама и маркетинг, % к базе", min_value=50, max_value=200, value=100, step=10,
        help="Регулирует объем затрат на рекламу объектов и лидогенерацию."
    )

# 5. CAPEX (Капитальные затраты)
with st.sidebar.expander("🛠️ 5. Регулировка CAPEX", expanded=False):
    capex_renovation_val = st.number_input("Ремонт и брендирование (всего), ₽", min_value=100000, max_value=2000000, value=520000, step=50000, help="Затраты на ремонт офиса и рекламные вывески.")
    capex_equipment_val = st.number_input("Мебель и оборудование (всего), ₽", min_value=100000, max_value=2000000, value=815000, step=50000, help="Столы, стулья, компьютеры, принтеры и роутеры на старте.")
    pau_fee_val = st.number_input("Паушальный взнос (с НДС), ₽", min_value=150000, max_value=1500000, value=preset["pau_fee"], step=10000)

# 6. Выплаты ЦО (Франшиза)
with st.sidebar.expander("🏢 6. Выплаты в ЦО", expanded=False):
    royalty_deal_custom = st.number_input(
        "Роялти за каждую сделку, ₽", min_value=1000, max_value=5000, value=preset["royalty_deal"], step=100,
        help="Фиксированный платеж в пользу управляющей компании за факт проведения транзакции."
    )

# ----------------------------------------------------
# ВЫЧИСЛИТЕЛЬНОЕ ЯДРО МОДЕЛИ (Декомпозиция Оптимальный)
# ----------------------------------------------------

base = optimal_baseline.copy()
scenario_name = "Оптимальный"
start_opex = 150000 # Базовый OPEX на старте

# Региональный коэффициент ФОТ и Роялти
salary_mult = preset["salary_mult"]
royalty_fix_mult = preset["royalty_fix_factor"]

# Инициализация всех списков для декомпозиции P&L
adj_leads = []
adj_meetings = []
adj_contracts = []
adj_prepayments = []
adj_deals_secondary = []
adj_deals_primary = []
adj_deals_rent = []
adj_deals_suburban = []

# Доходы
rev_secondary_list = []
rev_primary_list = []
rev_rent_list = []
rev_suburban_list = []
rev_overseas_list = []
rev_other_p_list = []
rev_mortgage_list = []
rev_insurance_list = []
rev_legal_list = []
rev_total_list = []

# Расходы - ФОТ (Декомпозированные по категориям и ролям)
agent_D_list = []
agent_C_list = []
agent_B_list = []
agent_A_list = []
total_agent_payouts_list = []

rop_C_list = []
rop_B_list = []
admin_list = []
hr_list = []
ros_list = []
jurist_list = []
mort_broker_list = []
listing_list = []
photographer_list = []
smm_list = []
total_backoffice_salaries_list = []

taxes_payroll_list = []
total_payroll_list = []

# Расходы - OPEX
rent_list = []
internet_list = []
mobile_list = []
kanc_list = []
reklama_list = []
hh_list = []
buh_list = []
bank_list = []
cleaning_list = []
gsm_list = []
courier_list = []
events_list = []
total_opex_list = []

# Расходы - Роялти и ЦО
hq_royalty_fix_list = []
hq_royalty_deal_list = []
hq_crm_list = []
hq_kc_list = []
total_hq_payments_list = []

# Налоги
taxes_usn_list = []

# Капитальные затраты (CAPEX)
capex_franchise_list = [0]*24
capex_renovation_list = [0]*24
capex_equipment_list = [0]*24
total_capex_list = []

# Итоговая чистая прибыль
adj_net_profit = []

# Календарное расширение CAPEX на 15 и 20 месяцы (из листа Оптимальный)
capex_renovation_list[14] = capex_renovation_val
capex_equipment_list[14] = capex_equipment_val
capex_equipment_list[19] = 790000

leads_mult = scaling_leads / 100.0
conv_meet_rate = conversion_meeting / 10.0 # Относительно базовых 10%
conv_deal_rate = conversion_deal / 5.0 # Относительно базовых 5% для Оптимального
overall_voeronka_factor = leads_mult * conv_meet_rate * conv_deal_rate

# Подготовка OPEX календаря
base_internet = [5000]*12 + [950]*12
base_mobile = [7800]*2 + [8400] + [9000] + [9600] + [10200] + [10800] + [11400] + [12000] + [12600] + [13200] + [13800] + [18000] + [18480] + [19800] + [21120] + [22440] + [23760] + [25080] + [26400] + [27720] + [29040] + [30360] + [31680]
base_rent_list = [150000]*12 + [150000]*2 + [300000]*10
base_kanc = [6500]*2 + [7000] + [7500] + [8000] + [8500] + [9000] + [9500] + [10000] + [10500] + [11000] + [11500] + [15000] + [15500] + [16500] + [17500] + [18500] + [20000] + [21500] + [22500] + [24000] + [24500] + [25500] + [26500]
base_reklama = [35000] + [70000] + [105000] + [140000] + [210000] + [280000] + [350000] + [420000] + [490000] + [560000] + [630000] + [700000] + [184800] + [215600] + [231000] + [246400] + [261800] + [277200] + [292600] + [308000] + [323400] + [338800] + [354200] + [369600]
base_hh = [45000]*12 + [90000]*12
base_buh = [20000]*12 + [40000]*12
base_bank = [2000]*12 + [6000]*12
base_cleaning = [15000]*12 + [45000]*12
base_gsm = [0]*10 + [5000]*2 + [10000]*4 + [15000]*2 + [20000]*2 + [25000]*4
base_courier = [0]*11 + [10000]*13

base_events = [0]*24
base_events[11] = 100000
base_events[15] = 350000
base_events[21] = 490000

for i in range(24):
    # 1. Лиды и воронка
    L = int(base['leads'][i] * leads_mult)
    adj_leads.append(L)
    adj_meetings.append(int(base['meetings'][i] * leads_mult * conv_meet_rate))
    adj_contracts.append(int(base['contracts'][i] * overall_voeronka_factor))
    adj_prepayments.append(base['prepayments'][i] * overall_voeronka_factor)
    
    # 2. Сделки
    deals_sec = base['prepayments'][i] * overall_voeronka_factor * 0.90
    deals_prim = base['prepayments'][i] * overall_voeronka_factor * 0.10
    adj_deals_secondary.append(deals_sec)
    adj_deals_primary.append(deals_prim)
    
    # 3. ДОХОДЫ с учетом 1-3 месяцев строго равных нулю + КОРРЕКТИРОВКИ
    if i < 3: # 1, 2, 3 месяцы нет доходов
        rev_sec = get_override('  ├─ Вторичный рынок', str(i+1), 0.0)
        rev_prim = get_override('  ├─ Первичный рынок', str(i+1), 0.0)
        rev_rent = get_override('  ├─ Аренда (жилая/коммерческая)', str(i+1), 0.0)
        rev_sub = get_override('  ├─ Загородная недвижимость', str(i+1), 0.0)
        rev_overseas = get_override('  ├─ Зарубежная недвижимость', str(i+1), 0.0)
        rev_other_p = get_override('  ├─ Сделки: прочее (МЛС, срочновыкуп)', str(i+1), 0.0)
        rev_mort = get_override('  ├─ Сервисы: ипотека', str(i+1), 0.0)
        rev_ins = get_override('  ├─ Сервисы: страхование', str(i+1), 0.0)
        rev_legal = get_override('  └─ Сервисы: юр. сопровождение', str(i+1), 0.0)
    else:
        rev_sec = get_override('  ├─ Вторичный рынок', str(i+1), base['rev_sec'][i] * overall_voeronka_factor * (comm_secondary / 360000.0))
        rev_prim = get_override('  ├─ Первичный рынок', str(i+1), base['rev_prim'][i] * overall_voeronka_factor * (comm_primary / 440000.0))
        rev_rent = get_override('  ├─ Аренда (жилая/коммерческая)', str(i+1), base['rev_rent'][i] * (comm_rent_val / 80000.0))
        rev_sub = get_override('  ├─ Загородная недвижимость', str(i+1), base['rev_sub'][i] * (comm_suburban / 500000.0))
        rev_overseas = get_override('  ├─ Зарубежная недвижимость', str(i+1), base['rev_overseas'][i])
        rev_other_p = get_override('  ├─ Сделки: прочее (МЛС, срочновыкуп)', str(i+1), base['rev_other_p'][i])
        rev_mort = get_override('  ├─ Сервисы: ипотека', str(i+1), base['rev_mort'][i] * overall_voeronka_factor * (comm_secondary / 360000.0))
        rev_ins = get_override('  ├─ Сервисы: страхование', str(i+1), base['rev_ins'][i] * overall_voeronka_factor * (comm_secondary / 360000.0))
        rev_legal = get_override('  └─ Сервисы: юр. сопровождение', str(i+1), base['rev_legal'][i] * overall_voeronka_factor * (comm_secondary / 360000.0))
        
    rev_secondary_list.append(rev_sec)
    rev_primary_list.append(rev_prim)
    rev_rent_list.append(rev_rent)
    rev_suburban_list.append(rev_sub)
    rev_overseas_list.append(rev_overseas)
    rev_other_p_list.append(rev_other_p)
    rev_mortgage_list.append(rev_mort)
    rev_insurance_list.append(rev_ins)
    rev_legal_list.append(rev_legal)
    
    total_revenue_month = rev_sec + rev_prim + rev_rent + rev_sub + rev_overseas + rev_other_p + rev_mort + rev_ins + rev_legal
    rev_total_list.append(total_revenue_month)
    
    # 4. РАСХОДЫ - ФОТ + ДЕКОМПОЗИЦИЯ И КОРРЕКТИРОВКИ
    agent_payouts_base = base['agent_comm'][i] * (agent_commission_pct / 38.0) * overall_voeronka_factor
    
    # Check if compact row override exists
    compact_agent_override = get_override('  ├─ Выплаты агентам (% комиссионных)', str(i+1), None)
    if compact_agent_override is not None:
        agent_payouts_base = compact_agent_override
        
    agent_D_val = get_override('  ├─ Агент: категория D', str(i+1), agent_payouts_base * 0.30)
    agent_C_val = get_override('  ├─ Агент: категория C', str(i+1), agent_payouts_base * 0.35)
    agent_B_val = get_override('  ├─ Агент: категория B', str(i+1), agent_payouts_base * 0.25)
    agent_A_val = get_override('  ├─ Агент: категория A', str(i+1), agent_payouts_base * 0.10)
    
    agent_payouts_total_month = agent_D_val + agent_C_val + agent_B_val + agent_A_val
    payouts_agents_list.append(agent_payouts_total_month)
    agent_D_list.append(agent_D_val)
    agent_C_list.append(agent_C_val)
    agent_B_list.append(agent_B_val)
    agent_A_list.append(agent_A_val)
    
    # Backoffice salaries base
    salaries_back_base = base['salaries'][i] * (backoffice_salary_mult / 100.0) * salary_mult
    compact_backoffice_override = get_override('  ├─ Оклады бэк-офиса (оклады)', str(i+1), None)
    if compact_backoffice_override is not None:
        salaries_back_base = compact_backoffice_override
        
    v_rop_C = get_override('  ├─ РОП: категория C', str(i+1), salaries_back_base * 0.15)
    v_rop_B = get_override('  ├─ РОП: категория B', str(i+1), salaries_back_base * 0.15)
    v_admin = get_override('  ├─ Администратор офиса', str(i+1), salaries_back_base * 0.10)
    v_hr = get_override('  ├─ HR/рекрутер', str(i+1), salaries_back_base * 0.10)
    v_ros = get_override('  ├─ РОС/тренер', str(i+1), salaries_back_base * 0.10)
    v_jurist = get_override('  ├─ Юрист', str(i+1), salaries_back_base * 0.10)
    v_mort_broker = get_override('  ├─ Ипотечный Брокер', str(i+1), salaries_back_base * 0.10)
    v_listing = get_override('  ├─ Листинг-менеджер', str(i+1), salaries_back_base * 0.08)
    v_photographer = get_override('  ├─ Фотограф', str(i+1), salaries_back_base * 0.04)
    v_smm = get_override('  ├─ Маркетолог/SMM', str(i+1), salaries_back_base * 0.08)
    
    salaries_back_total_month = v_rop_C + v_rop_B + v_admin + v_hr + v_ros + v_jurist + v_mort_broker + v_listing + v_photographer + v_smm
    salaries_backoffice_list.append(salaries_back_total_month)
    
    rop_C_list.append(v_rop_C)
    rop_B_list.append(v_rop_B)
    admin_list.append(v_admin)
    hr_list.append(v_hr)
    ros_list.append(v_ros)
    jurist_list.append(v_jurist)
    mort_broker_list.append(v_mort_broker)
    listing_list.append(v_listing)
    photographer_list.append(v_photographer)
    smm_list.append(v_smm)
    
    # Taxes
    taxes_payroll_base = base['salary_taxes'][i] * (backoffice_salary_mult / 100.0) * salary_mult
    taxes_payroll = get_override('  └─ Налоги на ФОТ оклады', str(i+1), taxes_payroll_base)
    taxes_payroll_list.append(taxes_payroll)
    
    total_payroll_list.append(agent_payouts_total_month + salaries_back_total_month + taxes_payroll)
    
    # 5. РАСХОДЫ - OPEX + КОРРЕКТИРОВКИ
    rent_val = get_override('  ├─ Аренда офиса', str(i+1), office_rent_custom * (base_rent_list[i] / 150000.0))
    rent_list.append(rent_val)
    
    internet_val = get_override('  ├─ Интернет', str(i+1), base_internet[i])
    internet_list.append(internet_val)
    
    mobile_val = get_override('  ├─ Сотовая связь', str(i+1), base_mobile[i])
    mobile_list.append(mobile_val)
    
    kanc_val = get_override('  ├─ Канцелярия', str(i+1), base_kanc[i])
    kanc_list.append(kanc_val)
    
    reklama_val = get_override('  ├─ Реклама объектов', str(i+1), base_reklama[i] * leads_mult * (marketing_mult / 100.0))
    reklama_list.append(reklama_val)
    
    hh_val = get_override('  ├─ HeadHunter.ru', str(i+1), base_hh[i])
    hh_list.append(hh_val)
    
    buh_val = get_override('  ├─ Бухгалтерия: аутсорс', str(i+1), base_buh[i])
    buh_list.append(buh_val)
    
    bank_val = get_override('  ├─ Услуги банка', str(i+1), base_bank[i])
    bank_list.append(bank_val)
    
    cleaning_val = get_override('  ├─ Уборка офиса', str(i+1), base_cleaning[i])
    cleaning_list.append(cleaning_val)
    
    gsm_val = get_override('  ├─ ГСМ', str(i+1), base_gsm[i])
    gsm_list.append(gsm_val)
    
    courier_val = get_override('  ├─ Доставка/курьер', str(i+1), base_courier[i])
    courier_list.append(courier_val)
    
    events_val = get_override('  └─ Корпоративные мероприятия', str(i+1), base_events[i])
    events_list.append(events_val)
    
    total_opex = rent_val + internet_val + mobile_val + kanc_val + reklama_val + hh_val + buh_val + bank_val + cleaning_val + gsm_val + courier_val + events_val
    total_opex_list.append(total_opex)
    
    # 6. РАСХОДЫ - ВЫПЛАТЫ ЦО + КОРРЕКТИРОВКИ
    hq_fix = get_override('  ├─ Роялти FIX (вкл. НРФ)', str(i+1), base['hq_royalty_fix'][i] * royalty_fix_mult)
    hq_royalty_fix_list.append(hq_fix)
    
    hq_deal = get_override('  ├─ Роялти со сделок', str(i+1), base['hq_royalty_deal'][i] * (royalty_deal_custom / 2100.0) * overall_voeronka_factor)
    hq_royalty_deal_list.append(hq_deal)
    
    hq_crm = get_override('  ├─ CRM-система', str(i+1), 5500 if i < 12 else 11000)
    hq_crm_list.append(hq_crm)
    
    hq_kc_val = get_override('  └─ Колл-центр и прочие сервисы', str(i+1), max(0, base['hq_crm_and_others'][i] - (5500 if i < 12 else 11000)))
    hq_kc_list.append(hq_kc_val)
    
    total_hq_payments = hq_fix + hq_deal + hq_crm + hq_kc_val
    total_hq_payments_list.append(total_hq_payments)
    
    # 7. БИЗНЕС-НАЛОГИ + КОРРЕКТИРОВКИ
    tax_month = get_override('★ НАЛОГ УСН', str(i+1), total_revenue_month * (tax_rate / 100.0))
    taxes_usn_list.append(tax_month)
    
    # 8. КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX) Месячный тотал + КОРРЕКТИРОВКИ
    capex_fran = get_override('  ├─ Франшиза (Паушальный взнос + Роспатент)', str(i+1), capex_franchise_list[i])
    capex_renov = get_override('  ├─ Ремонт и Брендирование офиса', str(i+1), capex_renovation_list[i])
    capex_equip = get_override('  └─ Мебель, компьютеры и оборудование', str(i+1), capex_equipment_list[i])
    
    # Update lists so they are printed correctly in Excel
    capex_franchise_list[i] = capex_fran
    capex_renovation_list[i] = capex_renov
    capex_equipment_list[i] = capex_equip
    
    tot_capex_month = capex_fran + capex_renov + capex_equip
    total_capex_list.append(tot_capex_month)
    
    # ИТОГО РАСХОДЫ
    total_expenses_month = agent_payouts + salaries_back + taxes_payroll + total_opex + total_hq_payments + tax_month
    
    net_profit_month = total_revenue_month - total_expenses_month - tot_capex_month
    adj_net_profit.append(net_profit_month)

# Расчет накопленного денежного потока начиная с Месяца 0 (CAPEX + OPEX старт)
capex_m0_franchise = get_override('  ├─ Франшиза (Паушальный взнос + Роспатент)', 'Старт', pau_fee_val + 17000)
capex_m0_renovation = get_override('  ├─ Ремонт и Брендирование офиса', 'Старт', 520000)
capex_m0_equipment = get_override('  └─ Мебель, компьютеры и оборудование', 'Старт', 815000)
total_capex_m0 = capex_m0_franchise + capex_m0_renovation + capex_m0_equipment

start_rent = get_override('  ├─ Аренда офиса', 'Старт', office_rent_custom)
start_opex_total = start_rent
for metric in ['  ├─ Интернет', '  ├─ Сотовая связь', '  ├─ Канцелярия', '  ├─ Реклама объектов', '  ├─ HeadHunter.ru', '  ├─ Бухгалтерия: аутсорс', '  ├─ Услуги банка', '  ├─ Уборка офиса', '  ├─ ГСМ', '  ├─ Доставка/курьер', '  └─ OPEX: Корпоративы']:
    start_opex_total += get_override(metric, 'Старт', 0.0)

cum_flow = []
running_sum = -total_capex_m0 - start_opex_total
cum_flow.append(running_sum) # месяц 0

for i in range(24):
    running_sum += adj_net_profit[i]
    cum_flow.append(running_sum)

# ----------------------------------------------------
# KPI И ПОКАЗАТЕЛИ ЭФФЕКТИВНОСТИ
# ----------------------------------------------------
total_revenue_2years = sum(rev_total_list)
total_profit_2years = sum(adj_net_profit) - total_capex_m0 - office_rent_custom
max_investment = abs(min(cum_flow))

payback_month = "Н/Д"
for idx, val in enumerate(cum_flow):
    if val >= 0:
        payback_month = f"{idx} мес."
        break

breakeven_month = "Н/Д"
for idx, val in enumerate(adj_net_profit):
    if val > 0:
        breakeven_month = f"{idx + 1} мес."
        break

# Отображение KPI Карточек
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("💼 Общая выручка (2 года)", f"{total_revenue_2years:,.0f} ₽".replace(",", " "))
with col2:
    st.metric("📈 Итоговая чистая прибыль", f"{total_profit_2years:,.0f} ₽".replace(",", " "))
with col3:
    st.metric("📉 Пик инвестиций (CAPEX+OPEX)", f"{max_investment:,.0f} ₽".replace(",", " "))
with col4:
    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("🎯 Без-ость", breakeven_month)
    with col_b:
        st.metric("⏳ Окупаемость", payback_month)

# ----------------------------------------------------
# ТАБЫ: ГРАФИКИ, ДЕТАЛИЗАЦИЯ И СРАВНЕНИЕ
# ----------------------------------------------------
tab1, tab2, tab3, tab4 = st.tabs(["📊 Интерактивные Графики", "📋 Декомпозиция P&L (Все статьи)", "🌍 Сравнение Регионов", "📝 Описание Модели"])

with tab1:
    st.markdown("### Визуализация финансового развития")
    
    # 1. График накопленного денежного потока
    fig_cum = go.Figure()
    fig_cum.add_trace(go.Scatter(
        x=list(range(25)),
        y=cum_flow,
        mode='lines+markers',
        name='Накопленный кэш-флоу',
        line=dict(color='#BEAF87', width=3),
        marker=dict(size=6, color='#FFFFFF', line=dict(color='#BEAF87', width=1.5))
    ))
    fig_cum.add_shape(
        type="line", x0=0, y0=0, x1=24, y1=0,
        line=dict(color="#777779", width=1.5, dash="dash")
    )
    fig_cum.update_layout(
        title="Кривая накопленного кэш-флоу (J-Curve)",
        xaxis_title="Месяцы проекта (0 - старт)",
        yaxis_title="Баланс кэша, ₽",
        template="plotly_white",
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        font=dict(color="#252526")
    )
    st.plotly_chart(fig_cum, use_container_width=True)
    
    # 2. Столбчатый график Доходы vs Расходы
    fig_rev_exp = go.Figure()
    fig_rev_exp.add_trace(go.Bar(
        x=months,
        y=rev_total_list,
        name='Доходы (Комиссия ВКД)',
        marker_color='#BEAF87'
    ))
    
    expenses_no_capex = [rev_total_list[idx] - adj_net_profit[idx] - total_capex_list[idx] for idx in range(24)]
    fig_rev_exp.add_trace(go.Bar(
        x=months,
        y=expenses_no_capex,
        name='Операционные расходы + Налоги',
        marker_color='#252526'
    ))
    fig_rev_exp.update_layout(
        title="Ежемесячные доходы и операционные расходы",
        xaxis_title="Месяцы",
        yaxis_title="Сумма, ₽",
        barmode='group',
        template="plotly_white",
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        font=dict(color="#252526")
    )
    st.plotly_chart(fig_rev_exp, use_container_width=True)

    # 3. Круговая диаграмма структуры расходов
    avg_expenses = {
        'Выплаты агентам': sum(payouts_agents_list),
        'Оклады бэк-офиса': sum(salaries_backoffice_list),
        'Налоги на ФОТ оклады': sum(taxes_payroll_list),
        'Аренда офиса': sum(rent_list),
        'Реклама и маркетинг': sum(reklama_list),
        'HH.ru и подбор': sum(hh_list),
        'Роялти и ЦО': sum(total_hq_payments_list),
        'УСН налоги': sum(taxes_usn_list),
        'Прочие OPEX': sum(internet_list) + sum(mobile_list) + sum(kanc_list) + sum(buh_list) + sum(bank_list) + sum(cleaning_list) + sum(gsm_list) + sum(courier_list) + sum(events_list)
    }
    fig_pie = px.pie(
        names=list(avg_expenses.keys()),
        values=list(avg_expenses.values()),
        color_discrete_sequence=['#BEAF87', '#A19276', '#252526', '#777779', '#CCCCCC', '#E6E7E8', '#999999', '#BEAF87', '#EAEAEA'],
        title="Детализированная структура расходов за 2 года"
    )
    fig_pie.update_layout(
        template="plotly_white",
        paper_bgcolor="#FFFFFF",
        font=dict(color="#252526")
    )
    st.plotly_chart(fig_pie, use_container_width=True)

with tab2:
    st.markdown("### 📋 Детализированный отчет о прибылях и убытках (P&L)")
    
    # UI Toggle for detailed FOT decomposition according to global UX/UI best practices
    col_toggle_1, col_toggle_2 = st.columns([3, 1])
    with col_toggle_1:
        show_detailed_fot = st.checkbox(
            "🔍 **Показать полную декомпозицию ФОТ (по должностям и категориям)**", 
            value=False,
            help="Показывает подробный разрез выплат агентам по категориям (D, C, B, A) и окладов бэк-офиса по каждой штатной единице отдельно."
        )
    
    # Define labels lists
    compact_labels = [
        '► ДОХОДЫ (ВКД)',
        '  ├─ Вторичный рынок',
        '  ├─ Первичный рынок',
        '  ├─ Аренда (жилая/коммерческая)',
        '  ├─ Загородная недвижимость',
        '  ├─ Зарубежная недвижимость',
        '  ├─ Сделки: прочее (МЛС, срочновыкуп)',
        '  ├─ Сервисы: ипотека',
        '  ├─ Сервисы: страхование',
        '  └─ Сервисы: юр. сопровождение',
        '★ ИТОГО ДОХОДЫ',
        
        '► РАСХОДЫ: ФОТ',
        '  ├─ Выплаты агентам (% комиссионных)',
        '  ├─ Оклады бэк-офиса (оклады)',
        '  └─ Налоги на ФОТ оклады',
        '★ ИТОГО ФОТ',
        
        '► РАСХОДЫ: OPEX',
        '  ├─ Аренда офиса',
        '  ├─ Интернет',
        '  ├─ Сотовая связь',
        '  ├─ Канцелярия',
        '  ├─ Реклама объектов',
        '  ├─ HeadHunter.ru',
        '  ├─ Бухгалтерия: аутсорс',
        '  ├─ Услуги банка',
        '  ├─ Уборка офиса',
        '  ├─ ГСМ',
        '  ├─ Доставка/курьер',
        '  └─ Корпоративные мероприятия',
        '★ ИТОГО OPEX',
        
        '► ВЫПЛАТЫ ЦО (ФРАНШИЗА)',
        '  ├─ Роялти FIX (вкл. НРФ)',
        '  ├─ Роялти со сделок',
        '  ├─ CRM-система',
        '  └─ Колл-центр и прочие сервисы',
        '★ ИТОГО ВЫПЛАТЫ ЦО',
        
        '★ НАЛОГ УСН',
        '★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ',
        
        '► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)',
        '  ├─ Франшиза (Паушальный взнос + Роспатент)',
        '  ├─ Ремонт и Брендирование офиса',
        '  └─ Мебель, компьютеры и оборудование',
        '★ ИТОГО CAPEX',
        
        '★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ',
        '★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК'
    ]
    
    detailed_labels = [
        '► ДОХОДЫ (ВКД)',
        '  ├─ Вторичный рынок',
        '  ├─ Первичный рынок',
        '  ├─ Аренда (жилая/коммерческая)',
        '  ├─ Загородная недвижимость',
        '  ├─ Зарубежная недвижимость',
        '  ├─ Сделки: прочее (МЛС, срочновыкуп)',
        '  ├─ Сервисы: ипотека',
        '  ├─ Сервисы: страхование',
        '  └─ Сервисы: юр. сопровождение',
        '★ ИТОГО ДОХОДЫ',
        
        '► РАСХОДЫ: ФОТ',
        '  ├─ Агент: категория D',
        '  ├─ Агент: категория C',
        '  ├─ Агент: категория B',
        '  ├─ Агент: категория A',
        '  ├─ РОП: категория C',
        '  ├─ РОП: категория B',
        '  ├─ Администратор офиса',
        '  ├─ HR/рекрутер',
        '  ├─ РОС/тренер',
        '  ├─ Юрист',
        '  ├─ Ипотечный Брокер',
        '  ├─ Листинг-менеджер',
        '  ├─ Фотограф',
        '  ├─ Маркетолог/SMM',
        '  └─ Налоги на ФОТ оклады',
        '★ ИТОГО ФОТ',
        
        '► РАСХОДЫ: OPEX',
        '  ├─ Аренда офиса',
        '  ├─ Интернет',
        '  ├─ Сотовая связь',
        '  ├─ Канцелярия',
        '  ├─ Реклама объектов',
        '  ├─ HeadHunter.ru',
        '  ├─ Бухгалтерия: аутсорс',
        '  ├─ Услуги банка',
        '  ├─ Уборка офиса',
        '  ├─ ГСМ',
        '  ├─ Доставка/курьер',
        '  └─ Корпоративные мероприятия',
        '★ ИТОГО OPEX',
        
        '► ВЫПЛАТЫ ЦО (ФРАНШИЗА)',
        '  ├─ Роялти FIX (вкл. НРФ)',
        '  ├─ Роялти со сделок',
        '  ├─ CRM-система',
        '  └─ Колл-центр и прочие сервисы',
        '★ ИТОГО ВЫПЛАТЫ ЦО',
        
        '★ НАЛОГ УСН',
        '★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ',
        
        '► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)',
        '  ├─ Франшиза (Паушальный взнос + Роспатент)',
        '  ├─ Ремонт и Брендирование офиса',
        '  └─ Мебель, компьютеры и оборудование',
        '★ ИТОГО CAPEX',
        
        '★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ',
        '★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК'
    ]
    
    # Helper function to compile row values for any column
    def build_rows_for_column(col_label):
        if col_label == 'Старт':
            r_sec = v_m0_rev_sec
            r_prim = v_m0_rev_prim
            r_rent = v_m0_rev_rent
            r_sub = v_m0_rev_sub
            r_overseas = v_m0_rev_overseas
            r_other_p = v_m0_rev_other_p
            r_mort = v_m0_rev_mort
            r_ins = v_m0_rev_ins
            r_legal = v_m0_rev_legal
            r_total = v_m0_rev_total
            
            f_D = v_m0_agent_D
            f_C = v_m0_agent_C
            f_B = v_m0_agent_B
            f_A = v_m0_agent_A
            f_agent_total = v_m0_agent_total
            
            f_rop_C = v_m0_rop_C
            f_rop_B = v_m0_rop_B
            f_admin = v_m0_admin
            f_hr = v_m0_hr
            f_ros = v_m0_ros
            f_jurist = v_m0_jurist
            f_mort_broker = v_m0_mort_broker
            f_listing = v_m0_listing
            f_photographer = v_m0_photographer
            f_smm = v_m0_smm
            f_back_total = v_m0_back_total
            
            f_taxes = v_m0_salary_taxes
            f_total = v_m0_fot_total
            
            op_rent = office_rent_custom
            op_internet = 0.0
            op_mobile = 0.0
            op_kanc = 0.0
            op_reklama = 0.0
            op_hh = 0.0
            op_buh = 0.0
            op_bank = 0.0
            op_cleaning = 0.0
            op_gsm = 0.0
            op_courier = 0.0
            op_events = 0.0
            op_total = office_rent_custom
            
            hq_fix_val = 0.0
            hq_deal_val = 0.0
            hq_crm_val = 0.0
            hq_kc_val = 0.0
            hq_total_val = 0.0
            
            tx_val = 0.0
            op_tx_total = f_total + op_total + hq_total_val + tx_val
            
            cap_franchise = capex_m0_franchise
            cap_renovation = capex_m0_renovation
            cap_equipment = capex_m0_equipment
            cap_total = total_capex_m0
            
            np_val = -total_capex_m0 - op_tx_total
            cf_val = cum_flow[0]
        else:
            idx = months.index(col_label)
            r_sec = rev_secondary_list[idx]
            r_prim = rev_primary_list[idx]
            r_rent = rev_rent_list[idx]
            r_sub = rev_suburban_list[idx]
            r_overseas = rev_overseas_list[idx]
            r_other_p = rev_other_p_list[idx]
            r_mort = rev_mortgage_list[idx]
            r_ins = rev_insurance_list[idx]
            r_legal = rev_legal_list[idx]
            r_total = rev_total_list[idx]
            
            f_D = agent_D_list[idx]
            f_C = agent_C_list[idx]
            f_B = agent_B_list[idx]
            f_A = agent_A_list[idx]
            f_agent_total = total_agent_payouts_list[idx]
            
            f_rop_C = rop_C_list[idx]
            f_rop_B = rop_B_list[idx]
            f_admin = admin_list[idx]
            f_hr = hr_list[idx]
            f_ros = ros_list[idx]
            f_jurist = jurist_list[idx]
            f_mort_broker = mort_broker_list[idx]
            f_listing = listing_list[idx]
            f_photographer = photographer_list[idx]
            f_smm = smm_list[idx]
            f_back_total = total_backoffice_salaries_list[idx]
            
            f_taxes = taxes_payroll_list[idx]
            f_total = total_payroll_list[idx]
            
            op_rent = rent_list[idx]
            op_internet = internet_list[idx]
            op_mobile = mobile_list[idx]
            op_kanc = kanc_list[idx]
            op_reklama = reklama_list[idx]
            op_hh = hh_list[idx]
            op_buh = buh_list[idx]
            op_bank = bank_list[idx]
            op_cleaning = cleaning_list[idx]
            op_gsm = gsm_list[idx]
            op_courier = courier_list[idx]
            op_events = events_list[idx]
            op_total = total_opex_list[idx]
            
            hq_fix_val = hq_royalty_fix_list[idx]
            hq_deal_val = hq_royalty_deal_list[idx]
            hq_crm_val = hq_crm_list[idx]
            hq_kc_val = hq_kc_list[idx]
            hq_total_val = total_hq_payments_list[idx]
            
            tx_val = taxes_usn_list[idx]
            op_tx_total = f_total + op_total + hq_total_val + tx_val
            
            cap_franchise = capex_franchise_list[idx]
            cap_renovation = capex_renovation_list[idx]
            cap_equipment = capex_equipment_list[idx]
            cap_total = total_capex_list[idx]
            
            np_val = adj_net_profit[idx]
            cf_val = cum_flow[idx+1]

        # Format values helper
        def fmt(v, is_header_row=False):
            if is_header_row:
                return ""
            return f"{v:,.0f} ₽".replace(",", " ")

        if show_detailed_fot:
            return [
                "", # ► ДОХОДЫ (ВКД)
                fmt(r_sec),
                fmt(r_prim),
                fmt(r_rent),
                fmt(r_sub),
                fmt(r_overseas),
                fmt(r_other_p),
                fmt(r_mort),
                fmt(r_ins),
                fmt(r_legal),
                fmt(r_total), # ★ ИТОГО ДОХОДЫ
                
                "", # ► РАСХОДЫ: ФОТ
                fmt(f_D),
                fmt(f_C),
                fmt(f_B),
                fmt(f_A),
                fmt(f_rop_C),
                fmt(f_rop_B),
                fmt(f_admin),
                fmt(f_hr),
                fmt(f_ros),
                fmt(f_jurist),
                fmt(f_mort_broker),
                fmt(f_listing),
                fmt(f_photographer),
                fmt(f_smm),
                fmt(f_taxes),
                fmt(f_total), # ★ ИТОГО ФОТ
                
                "", # ► РАСХОДЫ: OPEX
                fmt(op_rent),
                fmt(op_internet),
                fmt(op_mobile),
                fmt(op_kanc),
                fmt(op_reklama),
                fmt(op_hh),
                fmt(op_buh),
                fmt(op_bank),
                fmt(op_cleaning),
                fmt(op_gsm),
                fmt(op_courier),
                fmt(op_events),
                fmt(op_total), # ★ ИТОГО OPEX
                
                "", # ► ВЫПЛАТЫ ЦО (ФРАНШИЗА)
                fmt(hq_fix_val),
                fmt(hq_deal_val),
                fmt(hq_crm_val),
                fmt(hq_kc_val),
                fmt(hq_total_val), # ★ ИТОГО ВЫПЛАТЫ ЦО
                
                fmt(tx_val), # ★ НАЛОГ УСН
                fmt(op_tx_total), # ★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ
                
                "", # ► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)
                fmt(cap_franchise),
                fmt(cap_renovation),
                fmt(cap_equipment),
                fmt(cap_total), # ★ ИТОГО CAPEX
                
                fmt(np_val), # ★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ
                fmt(cf_val)  # ★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК
            ]
        else:
            return [
                "", # ► ДОХОДЫ (ВКД)
                fmt(r_sec),
                fmt(r_prim),
                fmt(r_rent),
                fmt(r_sub),
                fmt(r_overseas),
                fmt(r_other_p),
                fmt(r_mort),
                fmt(r_ins),
                fmt(r_legal),
                fmt(r_total), # ★ ИТОГО ДОХОДЫ
                
                "", # ► РАСХОДЫ: ФОТ
                fmt(f_agent_total), #  ├─ Выплаты агентам (% комиссионных)
                fmt(f_back_total),  #  ├─ Оклады бэк-офиса (оклады)
                fmt(f_taxes),       #  └─ Налоги на ФОТ оклады
                fmt(f_total),       # ★ ИТОГО ФОТ
                
                "", # ► РАСХОДЫ: OPEX
                fmt(op_rent),
                fmt(op_internet),
                fmt(op_mobile),
                fmt(op_kanc),
                fmt(op_reklama),
                fmt(op_hh),
                fmt(op_buh),
                fmt(op_bank),
                fmt(op_cleaning),
                fmt(op_gsm),
                fmt(op_courier),
                fmt(op_events),
                fmt(op_total), # ★ ИТОГО OPEX
                
                "", # ► ВЫПЛАТЫ ЦО (ФРАНШИЗА)
                fmt(hq_fix_val),
                fmt(hq_deal_val),
                fmt(hq_crm_val),
                fmt(hq_kc_val),
                fmt(hq_total_val), # ★ ИТОГО ВЫПЛАТЫ ЦО
                
                fmt(tx_val), # ★ НАЛОГ УСН
                fmt(op_tx_total), # ★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ
                
                "", # ► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)
                fmt(cap_franchise),
                fmt(cap_renovation),
                fmt(cap_equipment),
                fmt(cap_total), # ★ ИТОГО CAPEX
                
                fmt(np_val), # ★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ
                fmt(cf_val)  # ★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК
            ]

    # Построение таблицы данных
    p_and_l_decomposed = {}
    if show_detailed_fot:
        p_and_l_decomposed['Категория / Статья'] = detailed_labels
    else:
        p_and_l_decomposed['Категория / Статья'] = compact_labels
        
    p_and_l_decomposed['Старт'] = build_rows_for_column('Старт')
    for m in months:
        p_and_l_decomposed[m] = build_rows_for_column(m)
        
    # Пре-расчет значений для колонки Старт (Месяц 0)
    v_m0_rev_sec = get_override('  ├─ Вторичный рынок', 'Старт', 0.0)
    v_m0_rev_prim = get_override('  ├─ Первичный рынок', 'Старт', 0.0)
    v_m0_rev_rent = get_override('  ├─ Аренда (жилая/коммерческая)', 'Старт', 0.0)
    v_m0_rev_sub = get_override('  ├─ Загородная недвижимость', 'Старт', 0.0)
    v_m0_rev_overseas = get_override('  ├─ Зарубежная недвижимость', 'Старт', 0.0)
    v_m0_rev_other_p = get_override('  ├─ Сделки: прочее (МЛС, срочновыкуп)', 'Старт', 0.0)
    v_m0_rev_mort = get_override('  ├─ Сервисы: ипотека', 'Старт', 0.0)
    v_m0_rev_ins = get_override('  ├─ Сервисы: страхование', 'Старт', 0.0)
    v_m0_rev_legal = get_override('  └─ Сервисы: юр. сопровождение', 'Старт', 0.0)
    v_m0_rev_total = v_m0_rev_sec + v_m0_rev_prim + v_m0_rev_rent + v_m0_rev_sub + v_m0_rev_overseas + v_m0_rev_other_p + v_m0_rev_mort + v_m0_rev_ins + v_m0_rev_legal

    v_m0_agent_D = get_override('  ├─ Агент: категория D', 'Старт', 0.0)
    v_m0_agent_C = get_override('  ├─ Агент: категория C', 'Старт', 0.0)
    v_m0_agent_B = get_override('  ├─ Агент: категория B', 'Старт', 0.0)
    v_m0_agent_A = get_override('  ├─ Агент: категория A', 'Старт', 0.0)
    
    v_m0_agent_compact = get_override('  ├─ Выплаты агентам (% комиссионных)', 'Старт', 0.0)
    if v_m0_agent_compact > 0.0 and (v_m0_agent_D + v_m0_agent_C + v_m0_agent_B + v_m0_agent_A == 0.0):
        v_m0_agent_D = v_m0_agent_compact * 0.30
        v_m0_agent_C = v_m0_agent_compact * 0.35
        v_m0_agent_B = v_m0_agent_compact * 0.25
        v_m0_agent_A = v_m0_agent_compact * 0.10
        
    v_m0_agent_total = v_m0_agent_D + v_m0_agent_C + v_m0_agent_B + v_m0_agent_A

    v_m0_rop_C = get_override('  ├─ РОП: категория C', 'Старт', 0.0)
    v_m0_rop_B = get_override('  ├─ РОП: категория B', 'Старт', 0.0)
    v_m0_admin = get_override('  ├─ Администратор офиса', 'Старт', 0.0)
    v_m0_hr = get_override('  ├─ HR/рекрутер', 'Старт', 0.0)
    v_m0_ros = get_override('  ├─ РОС/тренер', 'Старт', 0.0)
    v_m0_jurist = get_override('  ├─ Юрист', 'Старт', 0.0)
    v_m0_mort_broker = get_override('  ├─ Ипотечный Брокер', 'Старт', 0.0)
    v_m0_listing = get_override('  ├─ Листинг-менеджер', 'Старт', 0.0)
    v_m0_photographer = get_override('  ├─ Фотограф', 'Старт', 0.0)
    v_m0_smm = get_override('  ├─ Маркетолог/SMM', 'Старт', 0.0)
    
    v_m0_back_compact = get_override('  ├─ Оклады бэк-офиса (оклады)', 'Старт', 0.0)
    if v_m0_back_compact > 0.0 and (v_m0_rop_C + v_m0_rop_B + v_m0_admin + v_m0_hr + v_m0_ros + v_m0_jurist + v_m0_mort_broker + v_m0_listing + v_m0_photographer + v_m0_smm == 0.0):
        v_m0_rop_C = v_m0_back_compact * 0.15
        v_m0_rop_B = v_m0_back_compact * 0.15
        v_m0_admin = v_m0_back_compact * 0.10
        v_m0_hr = v_m0_back_compact * 0.10
        v_m0_ros = v_m0_back_compact * 0.10
        v_m0_jurist = v_m0_back_compact * 0.10
        v_m0_mort_broker = v_m0_back_compact * 0.10
        v_m0_listing = v_m0_back_compact * 0.08
        v_m0_photographer = v_m0_back_compact * 0.04
        v_m0_smm = v_m0_back_compact * 0.08
        
    v_m0_back_total = v_m0_rop_C + v_m0_rop_B + v_m0_admin + v_m0_hr + v_m0_ros + v_m0_jurist + v_m0_mort_broker + v_m0_listing + v_m0_photographer + v_m0_smm
    
    v_m0_salary_taxes = get_override('  └─ Налоги на ФОТ оклады', 'Старт', 0.0)
    v_m0_fot_total = v_m0_agent_total + v_m0_back_total + v_m0_salary_taxes

    v_m0_rent = start_rent
    v_m0_internet = get_override('  ├─ Интернет', 'Старт', 0.0)
    v_m0_mobile = get_override('  ├─ Сотовая связь', 'Старт', 0.0)
    v_m0_kanc = get_override('  ├─ Канцелярия', 'Старт', 0.0)
    v_m0_reklama = get_override('  ├─ Реклама объектов', 'Старт', 0.0)
    v_m0_hh = get_override('  ├─ HeadHunter.ru', 'Старт', 0.0)
    v_m0_buh = get_override('  ├─ Бухгалтерия: аутсорс', 'Старт', 0.0)
    v_m0_bank = get_override('  ├─ Услуги банка', 'Старт', 0.0)
    v_m0_cleaning = get_override('  ├─ Уборка офиса', 'Старт', 0.0)
    v_m0_gsm = get_override('  ├─ ГСМ', 'Старт', 0.0)
    v_m0_courier = get_override('  ├─ Доставка/курьер', 'Старт', 0.0)
    v_m0_events = get_override('  └─ Корпоративные мероприятия', 'Старт', 0.0)
    v_m0_opex_total = v_m0_rent + v_m0_internet + v_m0_mobile + v_m0_kanc + v_m0_reklama + v_m0_hh + v_m0_buh + v_m0_bank + v_m0_cleaning + v_m0_gsm + v_m0_courier + v_m0_events

    v_m0_hq_fix = get_override('  ├─ Роялти FIX (вкл. НРФ)', 'Старт', 0.0)
    v_m0_hq_deal = get_override('  ├─ Роялти со сделок', 'Старт', 0.0)
    v_m0_hq_crm = get_override('  ├─ CRM-система', 'Старт', 0.0)
    v_m0_hq_kc = get_override('  └─ Колл-центр и прочие сервисы', 'Старт', 0.0)
    v_m0_hq_total = v_m0_hq_fix + v_m0_hq_deal + v_m0_hq_crm + v_m0_hq_kc

    v_m0_tax = get_override('★ НАЛОГ УСН', 'Старт', 0.0)
    v_m0_opex_tax_total = v_m0_fot_total + v_m0_opex_total + v_m0_hq_total + v_m0_tax

    df_p_and_l_decomposed = pd.DataFrame(p_and_l_decomposed)
    
    # ----------------------------------------------------
    # ИНТЕРАКТИВНОЕ РЕДАКТИРОВАНИЕ ТАБЛИЦЫ P&L (OVERRIDES)
    # ----------------------------------------------------
    st.markdown("ℹ️ **Вы можете отредактировать любое расчетное значение в таблице ниже (кроме заголовков и итогов), и вся модель мгновенно пересчитает зависимые показатели.**")
    
    edited_df = st.data_editor(
        df_p_and_l_decomposed,
        use_container_width=True,
        height=600,
        disabled=["Категория / Статья"],
        key="pl_editor"
    )
    
    # Добавляем кнопку сброса ручных корректировок и глоссарий
    col_reset_1, col_reset_2 = st.columns([3, 1])
    with col_reset_2:
        if st.button("🔄 Сбросить ручные корректировки", use_container_width=True, help="Очищает все ручные изменения ячеек и возвращает модель к базовым расчетным формулам."):
            st.session_state['pl_overrides'] = {}
            st.success("Все ручные изменения успешно сброшены!")
            st.rerun()
            
    # ----------------------------------------------------
    # ИНТЕРАКТИВНЫЙ БИЗНЕС-СПРАВОЧНИК СТАТЕЙ (ГЛОССАРИЙ)
    # ----------------------------------------------------
    with st.expander("💡 Интерактивный бизнес-справочник статей P&L и подсказки", expanded=False):
        st.write("Выберите любую статью расходов или доходов ниже, чтобы увидеть подробную формулу, логику расчета и рекомендации управляющей компании Century 21:")
        
        glossary = {
            "Вторичный рынок (Доходы)": {
                "описание": "Комиссионные доходы агентства с продажи квартир и коммерческих объектов на вторичном рынке недвижимости.",
                "формула": "Количество сделок * Средняя комиссия за сделку (базово 360 000 ₽).",
                "совет": "Это основное ядро выручки вашего офиса (около 70-80% всех доходов). Контролируйте конверсию встреч в эксклюзивные договоры — она должна быть не ниже 10-15%."
            },
            "Первичный рынок (Доходы)": {
                "описание": "Комиссия, выплачиваемая застройщиками за продажу квартир в новостройках агентами вашего офиса.",
                "формула": "Количество сделок * Средняя комиссия от девелоперов (базово 440 000 ₽).",
                "совет": "Главное преимущество первичного рынка — комиссию выплачивает застройщик, а не покупатель. Обучайте агентов работе с базами новостроек для максимизации кросс-продаж."
            },
            "Аренда (Доходы)": {
                "описание": "Быстрые комиссионные доходы со сделок аренды жилой и коммерческой недвижимости.",
                "формула": "Количество сделок * Средняя комиссия (базово 80 000 ₽).",
                "совет": "Идеально подходит для старта молодых агентов, помогая им быстро получить первый доход и поверить в свои силы в течение первых 1-2 месяцев работы."
            },
            "Загородная недвижимость (Доходы)": {
                "описание": "Комиссионные сборы с продажи земельных участков, дач, загородных домов и коттеджей.",
                "формула": "Количество сделок * Средняя комиссия (базово 500 000 ₽).",
                "совет": "Загородный рынок имеет высокую сезонность (пик весной-летом), но радует очень крупным средним чеком."
            },
            "Доп. сервисы: ипотека / страхование / юр. сопровождение (Доходы)": {
                "описание": "Дополнительные кросс-доходы, получаемые от банков за ипотечный сплит, страховых компаний за оформление полисов и за юридическое сопровождение сторонних сделок.",
                "формула": "Доля сделок с услугой * Фиксированная ставка вознаграждения за услугу.",
                "совет": "Оформление ипотеки и страховки прямо в офисе — мощнейший рычаг повышения маржинальности агентства без роста OPEX."
            },
            "Выплаты агентам (Расходы: ФОТ)": {
                "описание": "Сдельная часть фонда оплаты труда. Выплачивается агентам в виде процента от принесенной комиссии.",
                "формула": "Совокупная выручка ВКД * Средний процент выплат (настраивается слайдером, базово 38%).",
                "совет": "Для мотивации агентов используется шкала Century 21 (от 35% для новичков категории D до 50% для звезд категории A). Переменная структура ФОТ защищает бизнес от убытков при снижении объемов сделок."
            },
            "Оклады бэк-офиса (Расходы: ФОТ)": {
                "описание": "Постоянная часть фонда оплаты труда. Оклады административного персонала офиса (РОП, рекрутер, юрист, ипотечный брокер, листинг-менеджер).",
                "формула": "Сумма окладов штатного расписания согласно календарю найма (масштабируется региональным коэффициентом).",
                "совет": "На старте (месяцы 1-5) минимизируйте фиксированные оклады. Нанимайте РОПа и рекрутеров в первую очередь, а узких специалистов (юрист, брокер) подключайте по мере роста штата агентов."
            },
            "Реклама объектов (Расходы: OPEX)": {
                "описание": "Ежемесячный рекламный бюджет на лидогенерацию, выгрузку объявлений на классифайды (Циан, Авито, Яндекс.Недвижимость) и локальное продвижение.",
                "формула": "Базовый норматив рекламы на одного агента * Количество активных агентов.",
                "совет": "Это ваш главный инвестиционный топливный бак. Снижение расходов на рекламу моментально бьет по воронке сделок через 1-2 месяца. Оптимизируйте выгрузки с помощью CRM-системы Century 21."
            },
            "Роялти FIX и со сделок (Выплаты ЦО)": {
                "описание": "Лицензионные франчайзинговые отчисления в пользу Центрального Офиса за использование бренда, бизнес-технологий и федерального маркетинга.",
                "формула": "Фиксированная ставка (по шкале месяцев) + Сдельная роялти (базово 2 100 ₽ за транзакцию).",
                "совет": "Фиксированный роялти имеет каникулярный период на старте проекта, снижая финансовую нагрузку во время раскрутки офиса."
            }
        }
        
        selected_metric_help = st.selectbox(
            "📚 Выберите статью P&L для вывода бизнес-справки и формулы:",
            options=list(glossary.keys()),
            key="glossary_selector"
        )
        
        if selected_metric_help:
            help_info = glossary[selected_metric_help]
            st.markdown(f"**📖 Суть статьи:** {help_info['описание']}")
            st.markdown(f"**🧮 Как рассчитывается:** `{help_info['формула']}`")
            st.markdown(f"**🎯 Совет Управляющей Компании:** *{help_info['совет']}*")
            
    # Детекция изменений ячеек
    changes_detected = False
    for idx_row in range(len(df_p_and_l_decomposed)):
        row_name = df_p_and_l_decomposed.at[idx_row, 'Категория / Статья']
        if "►" in row_name or "★" in row_name:
            continue
        for col in df_p_and_l_decomposed.columns[1:]: # Старт, 1, 2, ...
            val_base = df_p_and_l_decomposed.at[idx_row, col]
            val_edited = edited_df.at[idx_row, col]
            if val_base != val_edited:
                parsed_val = parse_value(val_edited)
                st.session_state['pl_overrides'][(row_name, col)] = parsed_val
                changes_detected = True
                
    if changes_detected:
        st.rerun()
    
    # ----------------------------------------------------
    # EXCEL EXPORT С ПОЛНОЙ ДЕКОМПОЗИЦИЕЙ
    # ----------------------------------------------------
    if openpyxl_available:
        def generate_excel():
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "C21 Decomposed P&L"
            
            ws.views.sheetView[0].showGridLines = True
            
            # Цветовые схемы Century 21
            gold_fill = PatternFill(start_color="BEAF87", end_color="BEAF87", fill_type="solid")
            dark_fill = PatternFill(start_color="252526", end_color="252526", fill_type="solid")
            light_gold_fill = PatternFill(start_color="F9F6EE", end_color="F9F6EE", fill_type="solid")
            category_fill = PatternFill(start_color="F1F1F3", end_color="F1F1F3", fill_type="solid")
            
            font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_bold = Font(name="Calibri", size=11, bold=True)
            font_regular = Font(name="Calibri", size=11)
            
            thin_border_side = Side(border_style="thin", color="#D3D3D3")
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            double_bottom_border = Border(bottom=Side(border_style="double", color="111111"), top=thin_border_side)
            
            # Заголовок листа
            ws.merge_cells("A1:Z1")
            ws["A1"] = f"ДЕТАЛИЗИРОВАННАЯ ФИНАНСОВАЯ МОДЕЛЬ CENTURY 21 - РЕГИОН: {region_select.upper()} (КЕЙС: БЮДЖЕТ ОПТИМАЛЬНЫЙ)"
            ws["A1"].font = font_title
            ws["A1"].fill = dark_fill
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40
            
            # Столбцы
            headers = ["Показатель / Период", "Старт"] + months
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = gold_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[3].height = 28
            
            # Месячный экспорт строк
            metrics_rows = [
                ('► ДОХОДЫ (ВКД)', [0], [""], ""),
                ('  ├─ Вторичный рынок', [0], rev_secondary_list, "#,##0"),
                ('  ├─ Первичный рынок', [0], rev_primary_list, "#,##0"),
                ('  ├─ Аренда (жилая/коммерческая)', [0], rev_rent_list, "#,##0"),
                ('  ├─ Загородная недвижимость', [0], rev_suburban_list, "#,##0"),
                ('  ├─ Зарубежная недвижимость', [0], rev_overseas_list, "#,##0"),
                ('  ├─ Сделки: прочее (МЛС, срочновыкуп)', [0], rev_other_p_list, "#,##0"),
                ('  ├─ Сервисы: ипотека', [0], rev_mortgage_list, "#,##0"),
                ('  ├─ Сервисы: страхование', [0], rev_insurance_list, "#,##0"),
                ('  └─ Сервисы: юр. сопровождение', [0], rev_legal_list, "#,##0"),
                ('★ ИТОГО ДОХОДЫ', [0], rev_total_list, "#,##0"),
                
                ('► РАСХОДЫ: ФОТ', [0], [""], ""),
                ('  ├─ Выплаты агентам (% комиссионных)', [0], payouts_agents_list, "#,##0"),
                ('  ├─ Оклады бэк-офиса (оклады)', [0], salaries_backoffice_list, "#,##0"),
                ('  └─ Налоги на ФОТ оклады', [0], taxes_payroll_list, "#,##0"),
                ('★ ИТОГО ФОТ', [0], total_payroll_list, "#,##0"),
                
                ('► РАСХОДЫ: OPEX', [0], [""], ""),
                ('  ├─ Аренда офиса', [office_rent_custom], rent_list, "#,##0"),
                ('  ├─ Интернет', [0], internet_list, "#,##0"),
                ('  ├─ Сотовая связь', [0], mobile_list, "#,##0"),
                ('  ├─ Канцелярия', [0], kanc_list, "#,##0"),
                ('  ├─ Реклама объектов', [0], reklama_list, "#,##0"),
                ('  ├─ HeadHunter.ru', [0], hh_list, "#,##0"),
                ('  ├─ Бухгалтерия: аутсорс', [0], buh_list, "#,##0"),
                ('  ├─ Услуги банка', [0], bank_list, "#,##0"),
                ('  ├─ Уборка офиса', [0], cleaning_list, "#,##0"),
                ('  ├─ ГСМ', [0], gsm_list, "#,##0"),
                ('  ├─ Доставка/курьер', [0], courier_list, "#,##0"),
                ('  └─ OPEX: Корпоративы', [0], events_list, "#,##0"),
                ('★ ИТОГО OPEX', [office_rent_custom], total_opex_list, "#,##0"),
                
                ('► ВЫПЛАТЫ ЦО (ФРАНШИЗА)', [0], [""], ""),
                ('  ├─ Роялти FIX (вкл. НРФ)', [0], hq_royalty_fix_list, "#,##0"),
                ('  ├─ Роялти со сделок', [0], hq_royalty_deal_list, "#,##0"),
                ('  ├─ CRM-система', [0], hq_crm_list, "#,##0"),
                ('  └─ Колл-центр и прочие сервисы', [0], hq_kc_list, "#,##0"),
                ('★ ИТОГО ВЫПЛАТЫ ЦО', [0], total_hq_payments_list, "#,##0"),
                
                ('★ НАЛОГ УСН', [0], taxes_usn_list, "#,##0"),
                ('★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ', [office_rent_custom], [total_payroll_list[idx]+total_opex_list[idx]+total_hq_payments_list[idx]+taxes_usn_list[idx] for idx in range(24)], "#,##0"),
                
                ('► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)', [0], [""], ""),
                ('  ├─ Франшиза (Паушальный взнос + Роспатент)', [capex_m0_franchise], capex_franchise_list, "#,##0"),
                ('  ├─ Ремонт и Брендирование офиса', [capex_renovation_val], capex_renovation_list, "#,##0"),
                ('  └─ Мебель, компьютеры и оборудование', [capex_equipment_val], capex_equipment_list, "#,##0"),
                ('★ ИТОГО CAPEX', [total_capex_m0], total_capex_list, "#,##0"),
                
                ('★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ', [-total_capex_m0-office_rent_custom], adj_net_profit, "#,##0"),
                ('★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК', [cum_flow[0]], cum_flow[1:], "#,##0")
            ]
            
            for r_idx, (m_name, m0_list, data_list, num_format) in enumerate(metrics_rows, 4):
                is_header = "►" in m_name
                is_total = "★" in m_name
                
                cell_a = ws.cell(row=r_idx, column=1, value=m_name)
                cell_a.font = font_bold if (is_header or is_total) else font_regular
                cell_a.border = thin_border
                
                if is_header:
                    cell_a.fill = category_fill
                elif is_total:
                    cell_a.fill = light_gold_fill
                
                # Заполнение столбца Старт (Месяц 0)
                cell_m0 = ws.cell(row=r_idx, column=2)
                if is_header:
                    cell_m0.value = ""
                    cell_m0.fill = category_fill
                else:
                    cell_m0.value = m0_list[0]
                    if num_format != "":
                        cell_m0.number_format = num_format
                    cell_m0.alignment = Alignment(horizontal="right")
                    if is_total:
                        cell_m0.font = font_bold
                        cell_m0.fill = light_gold_fill
                    else:
                        cell_m0.font = font_regular
                cell_m0.border = thin_border
                
                # Заполнение месяцев 1-24
                for c_idx in range(1, 25):
                    if is_header:
                        cell_val = ws.cell(row=r_idx, column=c_idx+2, value="")
                        cell_val.fill = category_fill
                        cell_val.border = thin_border
                        continue
                        
                    val = data_list[c_idx-1] if len(data_list) >= c_idx else 0
                    cell_val = ws.cell(row=r_idx, column=c_idx+2, value=val)
                    if num_format != "":
                        cell_val.number_format = num_format
                    cell_val.alignment = Alignment(horizontal="right")
                    cell_val.border = thin_border
                    
                    if is_total:
                        cell_val.font = font_bold
                        cell_val.fill = light_gold_fill
                        if "ЧИСТАЯ ПРИБЫЛЬ" in m_name:
                            cell_val.border = double_bottom_border
                    else:
                        cell_val.font = font_regular
                
                ws.row_dimensions[r_idx].height = 20
                
            ws.column_dimensions["A"].width = 45
            ws.column_dimensions["B"].width = 15
            for col in range(3, 27):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 15
                
            wb.save(output)
            return output.getvalue()
            
        excel_data = generate_excel()
        st.download_button(
            label="📥 Скачать детализированный P&L в Excel",
            data=excel_data,
            file_name=f"C21_P_and_L_Optimal_{region_select}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("⚠️ **Экспорт в Excel временно недоступен.**")

with tab3:
    st.markdown("### 🌍 Сравнение ключевых показателей регионов")
    st.write("Сравнение базовых финансово-экономических параметров по региональным сегментам Century 21:")
    
    comp_data = {
        'Региональный сегмент': list(REGIONAL_PRESETS.keys()),
        'Паушальный взнос (база)': [f"{v['pau_fee']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Вторичка': [f"{v['comm_sec']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Первичка': [f"{v['comm_prim']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Загородная': [f"{v['comm_sub']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Оклады бэк-офиса': [f"{v['salary_mult']*100:.0f}% к Москве" for v in REGIONAL_PRESETS.values()],
        'Аренда офиса (база)': [f"{v['rent_base']:,} ₽/мес.".replace(",", " ") for v in REGIONAL_PRESETS.values()],
    }
    df_comp = pd.DataFrame(comp_data)
    st.table(df_comp)

with tab4:
    st.markdown("""
    ### 📖 Справочник и методология декомпозиции (Кейс: Бюджет Оптимальный)
    
    Эта интерактивная модель разработана специально для симуляции листа **«Бюджет Оптимальный»** финансового плана Century 21 за 24 месяца:
    
    1. **Доходы (Комиссия ВКД):** 
       * Рассчитывается динамически на основе воронки лидов. 
       * Месяцы 1–3 полностью освобождены от доходов в связи со стартовым лагом набора базы и закрытия первых сделок (0 ₽).
       * Включает доп. сервисы — **комиссионные вознаграждения от банков по ипотеке, страховых компаний и юр. услуг**.
       
    2. **ФОТ (Зарплатный фонд):**
       * **Переменная часть (агенты):** Привязана к объему сделок и настраивается слайдером (базово 38%).
       * **Постоянная часть (оклады):** Включает фиксированные оклады РОПа, HR/рекрутера, юриста, ипотечного брокера, листинг-менеджера и администратора, скорректированные под выбранный регион.
       
    3. **OPEX (Операционные расходы):**
       * Разделен на 12 детализированных статей, включая затраты на интернет, сотовую связь, рекламу объектов, подбор персонала, IT-лицензии и уборку.
       
    4. **Выплаты ЦО (Франшиза):**
       * Фиксированные роялти (шкала по месяцам), переменные роялти с каждой сделки (настраиваются слайдером) и IT-сервисы ЦО.
       
    5. **CAPEX (Капитальные вложения):**
       * Рассчитывается как агрессивный старт в Месяце 0: Паушальный взнос + Роспатент + Полная закупка оборудования и ремонт офиса (**1 335 000 ₽**).
    """)
