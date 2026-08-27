import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import io

# Проверяем доступность openpyxl для экспорта в Excel без падения приложения
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    openpyxl_available = True
except ImportError:
    openpyxl_available = False

# Установка конфигурации страницы
st.set_page_config(
    page_title="CENTURY 21 Financial Model MVP v8",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Пользовательский CSS-стилинг в корпоративных цветах Century 21 (Светлая тема: Золотой #BEAF87, Угольный серый #252526, Светло-серый #E6E7E8)
st.markdown("""
<style>
    /* Установка фоновых цветов всего приложения */
    .stApp {
        background-color: #FFFFFF !important;
        color: #252526 !important;
        font-family: 'Segoe UI', Arial, sans-serif;
    }
    
    /* Оформление карточек показателей (Metrics) */
    div[data-testid="stMetric"] {
        background-color: #F8F9FA !important;
        padding: 15px !important;
        border-radius: 8px !important;
        border-left: 5px solid #BEAF87 !important;
        border-top: 1px solid #E6E7E8 !important;
        border-right: 1px solid #E6E7E8 !important;
        border-bottom: 1px solid #E6E7E8 !important;
        box-shadow: 0 4px 6px rgba(0,0,0,0.04) !important;
    }
    div[data-testid="stMetric"] label {
        color: #252526 !important;
        font-weight: 700 !important;
        font-size: 14px !important;
    }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        color: #A19276 !important;
        font-size: 22px !important;
        font-weight: bold !important;
    }
    
    /* Заголовки */
    .main-title {
        color: #252526 !important;
        font-size: 32px !important;
        font-weight: 800 !important;
        text-align: center !important;
        margin-bottom: 5px !important;
        text-transform: uppercase !important;
        letter-spacing: 2px !important;
        border-bottom: 3px solid #BEAF87 !important;
        padding-bottom: 15px !important;
        display: inline-block !important;
        width: 100% !important;
    }
    .sub-title {
        color: #808285 !important;
        font-size: 16px !important;
        text-align: center !important;
        margin-top: 10px !important;
        margin-bottom: 30px !important;
        font-style: italic !important;
    }
    
    /* Оформление боковой панели (Sidebar) */
    [data-testid="stSidebar"] {
        background-color: #E6E7E8 !important;
        color: #252526 !important;
        border-right: 1px solid #D1D2D4 !important;
    }
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 {
        color: #252526 !important;
    }
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] label {
        color: #252526 !important;
        font-weight: 600 !important;
    }
    
    /* Стилизация кнопок */
    div.stButton > button:first-child {
        background-color: #BEAF87 !important;
        color: #FFFFFF !important;
        font-weight: bold !important;
        border: none !important;
        border-radius: 5px !important;
        transition: all 0.3s ease !important;
        width: 100% !important;
        padding: 10px !important;
        border-bottom: 3px solid #A19276 !important;
    }
    div.stButton > button:first-child:hover {
        background-color: #A19276 !important;
        color: #FFFFFF !important;
        box-shadow: 0 4px 15px rgba(161, 146, 118, 0.4) !important;
    }
    
    /* Стилизация вкладок (Tabs) */
    .stTabs [data-baseweb="tab-list"] {
        gap: 10px !important;
        background-color: #F8F9FA !important;
        padding: 8px 8px 0px 8px !important;
        border-radius: 8px 8px 0px 0px !important;
        border-bottom: 2px solid #E6E7E8 !important;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #E6E7E8 !important;
        border: 1px solid #D1D2D4 !important;
        border-radius: 5px 5px 0px 0px !important;
        padding: 10px 20px !important;
        color: #252526 !important;
    }
    .stTabs [aria-selected="true"] {
        background-color: #FFFFFF !important;
        color: #BEAF87 !important;
        border-bottom: 3px solid #BEAF87 !important;
        font-weight: bold !important;
    }
    
    /* Оформление таблиц данных */
    .stDataFrame {
        border: 1px solid #E6E7E8 !important;
        border-radius: 8px !important;
        background-color: #FFFFFF !important;
    }
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# ОПРЕДЕЛЕНИЕ БАЗОВЫХ ДАННЫХ (На основе источников C21)
# ----------------------------------------------------

months = [f"Месяц {i}" for i in range(1, 25)]

hyp1_baseline = {
    'leads': [60, 120, 180, 240, 300, 360, 420, 480, 540, 600, 660, 720,
              840, 900, 960, 1020, 1080, 1140, 1200, 1260, 1320, 1380, 1440, 1500],
    'deals_rent': [3, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1, 1,
                   3, 3, 4, 4, 4, 4, 5, 5, 6, 6, 6, 6],
    'deals_suburban': [0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1, 1,
                       2, 2, 2, 2, 3, 3, 4, 4, 5, 5, 5, 5],
    'revenue': [0, 0, 0, 1520000, 957400, 1096100, 1811500, 2619650, 2322329, 2559567, 3376806, 3614044,
                3974182, 4247600, 4551568, 5070055, 5874023, 6357991, 6941959, 7245928, 8049896, 8313864, 8837832, 8881800],
    'agent_comm': [0, 0, 0, 612100, 398312, 460152, 892040, 1274170, 1109098, 1215943, 1628083, 1734144,
                   2053406, 2216022, 2372104, 2625039, 3002856, 3369888, 3678751, 3834367, 4214813, 4349624, 4587930, 4640042],
    'salaries': [214935, 214935, 214935, 214935, 214935, 144935, 144935, 144935, 144935, 144935, 144935, 144935,
                 957423, 957814, 960140, 1118661, 1120987, 1453313, 1598638, 1600964, 1753290, 1755616, 1757942, 1760267],
    'salary_taxes': [44935]*12 + [137423, 137814, 140140, 148661, 150987, 153313, 198638, 200964, 203290, 205616, 207942, 210267],
    'opex': [168800, 186300, 204900, 223500, 359600, 395700, 431800, 467900, 504000, 540100, 581200, 727300,
             460550, 463210, 621050, 978890, 641730, 650070, 663410, 671250, 684590, 1181930, 699770, 707610],
    'hq_royalty_fix': [0, 0, 0, 45675, 45675, 45675, 45675, 78750, 78750, 78750, 78750, 78750,
                       111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825],
    'hq_royalty_deal': [0, 0, 0, 8400, 6752, 5927, 9303, 12679, 11855, 13130, 16506, 17782,
                        18863, 20267, 21670, 24477, 27980, 29383, 32887, 34290, 37793, 39197, 40600, 42003],
    'hq_crm_and_others': [19400, 5500, 5500, 18500, 18500, 18500, 18500, 18500, 45740, 50309, 54878, 59447,
                          40068, 33446, 33635, 33824, 34013, 100202, 34391, 34580, 34769, 34958, 101147, 35336],
    'capex': [0]*24
}

optimal_baseline = {
    'leads': [440, 880, 1320, 1760, 2200, 2640, 3080, 3520, 3960, 4400, 4840, 5280,
              6160, 6600, 7040, 7480, 7920, 8360, 8800, 9240, 9680, 10120, 10560, 11000],
    'deals_rent': [3, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1,
                   3, 3, 4, 4, 4, 4, 5, 5, 6, 6, 6, 6],
    'deals_suburban': [0, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1,
                       2, 2, 2, 2, 3, 3, 4, 4, 5, 5, 5, 5],
    'revenue': [0, 0, 0, 800000, 1520000, 2623800, 2717000, 3714700, 4985694, 4985609, 5565525, 6725440,
                7814356, 8482712, 9167967, 10435378, 11620634, 12485890, 13451145, 14136401, 15321656, 15966912, 16872168, 17297423],
    'agent_comm': [0, 0, 0, 338500, 612100, 1053409, 1314265, 1784222, 2332633, 2318694, 2629481, 3171647,
                   4103166, 4502709, 4855674, 5449734, 6035280, 6593359, 7120893, 7473269, 8061122, 8387655, 8822681, 9066321],
    'salaries': [214935, 214935, 214935, 214935, 214935, 144935, 144935, 144935, 144935, 144935, 144935, 144935,
                 963068, 964023, 966913, 1126564, 1129454, 1462344, 1608235, 1611125, 1764015, 1766905, 1769796, 1772686],
    'salary_taxes': [44935]*12 + [143068, 144023, 146913, 156564, 159454, 162344, 208235, 211125, 214015, 216905, 219796, 222686],
    'opex': [150000, 286300, 321300, 357400, 393500, 464600, 535700, 606800, 677900, 749000, 820100, 896200,
             569750, 601530, 769250, 1136970, 809690, 827910, 851130, 868850, 892070, 1399290, 927010, 944730],
    'hq_royalty_fix': [0, 0, 0, 45675, 45675, 45675, 45675, 78750, 78750, 78750, 78750, 78750,
                       111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825],
    'hq_royalty_deal': [0, 0, 0, 4200, 8400, 13556, 14574, 19793, 25011, 26030, 29148, 34367,
                        39134, 42564, 45994, 52855, 58385, 61816, 67346, 70776, 76307, 79737, 83167, 86598],
    'hq_crm_and_others': [19400, 5500, 5500, 18500, 18500, 18500, 18500, 18500, 85147, 96247, 107347, 118470,
                          43344, 37268, 37730, 38192, 38654, 105116, 39578, 40040, 40502, 40964, 107426, 41888],
    'capex': [0]*24
}

# ----------------------------------------------------
# РЕГИОНАЛЬНЫЕ ПРЕСЕТЫ И ПАРАМЕТРЫ ФРАНШИЗЫ
# ----------------------------------------------------
REGIONAL_PRESETS = {
    "Москва (Moscow)": {
        "pau_fee": 577500,
        "comm_sec": 360000,
        "comm_prim": 440000,
        "comm_sub": 500000,
        "comm_rent": 80000,
        "rent_base": 150000,
        "salary_mult": 1.0,
        "royalty_fix_factor": 1.0,
        "royalty_deal": 2100,
    },
    "Московская область (МО)": {
        "pau_fee": 500000,
        "comm_sec": 280000,
        "comm_prim": 330000,
        "comm_sub": 400000,
        "comm_rent": 60000,
        "rent_base": 100000,
        "salary_mult": 0.85,
        "royalty_fix_factor": 0.85,
        "royalty_deal": 2000,
    },
    "Крупные города (население > 800к)": {
        "pau_fee": 450000,
        "comm_sec": 220000,
        "comm_prim": 260000,
        "comm_sub": 320000,
        "comm_rent": 45000,
        "rent_base": 80000,
        "salary_mult": 0.75,
        "royalty_fix_factor": 0.75,
        "royalty_deal": 2000,
    },
    "Средние города (население 400к - 800к)": {
        "pau_fee": 350000,
        "comm_sec": 160000,
        "comm_prim": 200000,
        "comm_sub": 240000,
        "comm_rent": 35000,
        "rent_base": 55000,
        "salary_mult": 0.60,
        "royalty_fix_factor": 0.60,
        "royalty_deal": 1800,
    },
    "Малые города (население < 400к)": {
        "pau_fee": 270000,
        "comm_sec": 110000,
        "comm_prim": 140000,
        "comm_sub": 180000,
        "comm_rent": 25000,
        "rent_base": 38000,
        "salary_mult": 0.45,
        "royalty_fix_factor": 0.45,
        "royalty_deal": 1500,
    }
}

# ----------------------------------------------------
# ШАПКА ПРИЛОЖЕНИЯ
# ----------------------------------------------------
st.markdown("<div class='main-title'>CENTURY 21 Financial Application (v8)</div>", unsafe_allow_html=True)
st.markdown("<div class='sub-title'>Интерактивная финансовая модель и декомпозиция доходов/расходов по регионам РФ</div>", unsafe_allow_html=True)

# ----------------------------------------------------
# БОКОВАЯ ПАНЕЛЬ С УПРАВЛЕНИЕМ
# ----------------------------------------------------
st.sidebar.markdown("<h2 style='color:#252526; text-align:center; font-weight:800;'>Управление моделью</h2>", unsafe_allow_html=True)

# Выбор Регионального типа
region_select = st.sidebar.selectbox(
    "🌍 Выберите регион / тип города:",
    options=list(REGIONAL_PRESETS.keys())
)
preset = REGIONAL_PRESETS[region_select]

# 1. Сценарий
scenario = st.sidebar.selectbox(
    "📊 Выберите базовый сценарий проекта:",
    options=["Гипотеза 1 (Региональный рост)", "Бюджет Оптимальный (Интенсивный старт)"]
)

# Параметры подгружаются из регионального пресета, но пользователь может их скорректировать
st.sidebar.markdown("### 🎚️ Калибровка региональных параметров")

# Паушальный взнос (слайдер на базе пресета)
regional_pau_fee = st.sidebar.number_input("Паушальный взнос (с НДС), ₽", min_value=150000, max_value=1000000, value=preset["pau_fee"], step=10000)

# 2. Драйверы воронки продаж
st.sidebar.markdown("### 📈 Воронка продаж")
scaling_leads = st.sidebar.slider("Количество лидов (звонков), % к базе", min_value=50, max_value=200, value=100, step=10)
conversion_meeting = st.sidebar.slider("Конверсия лид -> встреча, %", min_value=2.0, max_value=20.0, value=10.0, step=0.5)

default_conversion_deal = 15.0 if "Гипотеза 1" in scenario else 5.0
conversion_deal = st.sidebar.slider("Конверсия встреча -> договор, %", min_value=1.0, max_value=30.0, value=default_conversion_deal, step=0.5)

# 3. Средние чеки
st.sidebar.markdown("### 💰 Средние комиссии (чеки сделок)")
comm_secondary = st.sidebar.number_input("Вторичный рынок, ₽", min_value=50000, max_value=1000000, value=preset["comm_sec"], step=10000)
comm_primary = st.sidebar.number_input("Первичный рынок, ₽", min_value=50000, max_value=1000000, value=preset["comm_prim"], step=10000)
comm_suburban = st.sidebar.number_input("Загородная недвижимость, ₽", min_value=50000, max_value=1000000, value=preset["comm_sub"], step=10000)
comm_rent_val = st.sidebar.number_input("Аренда (жилая/коммерческая), ₽", min_value=10000, max_value=200000, value=preset["comm_rent"], step=5000)

# 4. ФОТ и комиссия агентов
st.sidebar.markdown("### 👥 Персонал и Выплаты")
agent_commission_pct = st.sidebar.slider("Средний % выплат агентам", min_value=25, max_value=60, value=38, step=1)
backoffice_salary_mult = st.sidebar.slider("Индекс окладов бэк-офиса, % к базе", min_value=50, max_value=150, value=100, step=5)

# 5. OPEX и Роялти
st.sidebar.markdown("### 🏢 Аренда и Налоги")
office_rent_custom = st.sidebar.number_input("Аренда офиса в месяц, ₽", min_value=10000, max_value=500000, value=preset["rent_base"], step=5000)
royalty_deal_custom = st.sidebar.number_input("Роялти за каждую сделку, ₽", min_value=1000, max_value=5000, value=preset["royalty_deal"], step=100)
tax_rate = st.sidebar.slider("Ставка налога УСН, %", min_value=3, max_value=15, value=7, step=1)

# ----------------------------------------------------
# ВЫЧИСЛИТЕЛЬНОЕ ЯДРО МОДЕЛИ (Полная декомпозиция)
# ----------------------------------------------------

# Загрузка базовых данных под сценарий
if "Гипотеза 1" in scenario:
    base = hyp1_baseline.copy()
    scenario_name = "Гипотеза 1"
    start_opex = 50000 # Офисный OPEX Месяца 0
else:
    base = optimal_baseline.copy()
    scenario_name = "Оптимальный"
    start_opex = 150000

# Региональный коэффициент ФОТ и Роялти
salary_mult = preset["salary_mult"]
royalty_fix_mult = preset["royalty_fix_factor"]

# Инициализация всех списков для декомпозиции P&L
adj_leads = []
adj_meetings = []
adj_contracts = []
adj_prepayments = []
adj_deals_secondary = []
adj_deals_primary = []
adj_deals_rent = []
adj_deals_suburban = []

# Доходы
rev_secondary_list = []
rev_primary_list = []
rev_rent_list = []
rev_suburban_list = []
rev_services_list = []
rev_others_list = []
rev_total_list = []

# Расходы - ФОТ
payouts_agents_list = []
salaries_backoffice_list = []
taxes_payroll_list = []
total_payroll_list = []

# Расходы - OPEX
rent_list = []
marketing_list = []
recruiting_list = []
cleaning_list = []
communications_list = []
accounting_banking_list = []
opex_others_list = []
total_opex_list = []

# Расходы - Роялти и ЦО
hq_royalty_fix_list = []
hq_royalty_deal_list = []
hq_crm_list = []
hq_others_list = []
total_hq_payments_list = []

# Бизнес-налоги
taxes_usn_list = []

# Капитальные затраты (CAPEX)
capex_franchise_list = [0]*24
capex_renovation_list = [0]*24
capex_equipment_list = [0]*24
total_capex_list = []

# Итоговая чистая прибыль
adj_net_profit = []

# Заполнение CAPEX календаря (Месяц 1 - 24)
if scenario_name == "Гипотеза 1":
    # Месяц 5 (индекс 4)
    capex_renovation_list[4] = 520000 # ремонт + брендирование
    capex_equipment_list[4] = 815000 # оборудование + мебель
    # Месяц 15 (индекс 14)
    capex_renovation_list[14] = 520000
    capex_equipment_list[14] = 815000
    # Месяц 20 (индекс 19)
    capex_equipment_list[19] = 790000
else:
    # Оптимальный (Интенсивный старт) - CAPEX вынесен в Месяц 0, в операционных месяцах идет расширение:
    # Месяц 15 (индекс 14)
    capex_renovation_list[14] = 520000
    capex_equipment_list[14] = 815000
    # Месяц 20 (индекс 19)
    capex_equipment_list[19] = 790000

leads_mult = scaling_leads / 100.0
conv_meet_rate = conversion_meeting / 100.0
conv_deal_rate = conversion_deal / 100.0

for i in range(24):
    # 1. Лиды и воронка
    L = int(base['leads'][i] * leads_mult)
    adj_leads.append(L)
    
    meetings = int(L * conv_meet_rate)
    adj_meetings.append(meetings)
    
    contracts = int(meetings * conv_deal_rate)
    adj_contracts.append(contracts)
    
    prepayments = contracts * 0.75
    adj_prepayments.append(prepayments)
    
    # 2. Сделки по категориям
    deals_sec = prepayments * 0.90
    adj_deals_secondary.append(deals_sec)
    
    deals_prim = prepayments * 0.10
    adj_deals_primary.append(deals_prim)
    
    deals_r = base['deals_rent'][i]
    adj_deals_rent.append(deals_r)
    
    deals_sub = base['deals_suburban'][i]
    adj_deals_suburban.append(deals_sub)
    
    # 3. ДОХОДЫ - Декомпозиция
    rev_sec = deals_sec * comm_secondary
    rev_secondary_list.append(rev_sec)
    
    rev_prim = deals_prim * comm_primary
    rev_primary_list.append(rev_prim)
    
    rev_rent = deals_r * comm_rent_val
    rev_rent_list.append(rev_rent)
    
    rev_sub = deals_sub * comm_suburban
    rev_suburban_list.append(rev_sub)
    
    # Расчет доп. сервисов (Ипотека, Страхование, Юр. сопровождение) на базе воронки вторички
    # 17% сделок берут ипотеку (70к), 17% страхуют (18к), 7% юр.сопровождение (150к) - масштабируем пропорционально чеку вторички
    serv_coef = comm_secondary / 360000.0
    rev_mortgage = deals_sec * 0.17 * 70000 * serv_coef
    rev_insurance = deals_sec * 0.17 * 18000 * serv_coef
    rev_legal = deals_sec * 0.07 * 150000 * serv_coef
    rev_services = rev_mortgage + rev_insurance + rev_legal
    rev_services_list.append(rev_services)
    
    # Прочие доходы (оверсис, МЛС) - находим остаток в базовой выручке и масштабируем
    base_leads = base['leads'][i]
    if base_leads > 0:
        base_meet = base_leads * 0.10
        base_contract = base_meet * (0.15 if scenario_name == "Гипотеза 1" else 0.05)
        base_prep = base_contract * 0.75
        base_deals_s = base_prep * 0.90
        base_rev_s = base_deals_s * 360000
        base_rev_p = (base_prep * 0.10) * 440000
        base_rev_r = base['deals_rent'][i] * 80000
        base_rev_sub = base['deals_suburban'][i] * 500000
        base_rev_serv = base_deals_s * (0.17*70000 + 0.17*18000 + 0.07*150000)
        base_total_modeled = base_rev_s + base_rev_p + base_rev_r + base_rev_sub + base_rev_serv
        residual = base['revenue'][i] - base_total_modeled
        residual = max(0, residual)
    else:
        residual = 0
    rev_others = residual * (L / base_leads if base_leads > 0 else 0) * serv_coef
    rev_others_list.append(rev_others)
    
    # Общая выручка за месяц
    total_revenue_month = rev_sec + rev_prim + rev_rent + rev_sub + rev_services + rev_others
    rev_total_list.append(total_revenue_month)
    
    # 4. РАСХОДЫ - ФОТ Декомпозиция
    agent_payouts = total_revenue_month * (agent_commission_pct / 100.0)
    payouts_agents_list.append(agent_payouts)
    
    salaries_back = base['salaries'][i] * (backoffice_salary_mult / 100.0) * salary_mult
    salaries_backoffice_list.append(salaries_back)
    
    taxes_payroll = salaries_back * 0.209  # соотношение налогов на ФОТ в Century 21
    taxes_payroll_list.append(taxes_payroll)
    
    total_payroll_list.append(agent_payouts + salaries_back + taxes_payroll)
    
    # 5. РАСХОДЫ - OPEX Декомпозиция
    # Корректируем базовую аренду
    base_rent_in_opex = (50000 if i < 4 else 150000) if scenario_name == "Гипотеза 1" else 150000
    opex_without_rent = max(0, base['opex'][i] - base_rent_in_opex)
    
    rent_list.append(office_rent_custom)
    
    # Декомпозируем остальные затраты
    marketing_val = opex_without_rent * 0.50
    marketing_list.append(marketing_val)
    
    recruiting_val = 45000 if opex_without_rent > 50000 else opex_without_rent * 0.15
    recruiting_list.append(recruiting_val)
    
    cleaning_val = 15000 if opex_without_rent > 20000 else opex_without_rent * 0.05
    cleaning_list.append(cleaning_val)
    
    accounting_banking_val = 22000 if opex_without_rent > 30000 else opex_without_rent * 0.08
    accounting_banking_list.append(accounting_banking_val)
    
    communications_val = opex_without_rent * 0.10
    communications_list.append(communications_val)
    
    opex_others_val = max(0, opex_without_rent - (marketing_val + recruiting_val + cleaning_val + accounting_banking_val + communications_val))
    opex_others_list.append(opex_others_val)
    
    total_opex_list.append(office_rent_custom + marketing_val + recruiting_val + cleaning_val + accounting_banking_val + communications_val + opex_others_val)
    
    # 6. РАСХОДЫ - ВЫПЛАТЫ ЦО (Франшиза)
    hq_fix = base['hq_royalty_fix'][i] * royalty_fix_mult
    hq_royalty_fix_list.append(hq_fix)
    
    total_deals = deals_sec + deals_prim + deals_r + deals_sub
    hq_deal = total_deals * royalty_deal_custom
    hq_royalty_deal_list.append(hq_deal)
    
    hq_crm = base['hq_crm_and_others'][i]
    hq_crm_list.append(hq_crm)
    
    # Прочие комиссионные платежи в ЦО (ипотечный сплит, страхование и др.)
    hq_others_val = (rev_mortgage * 0.50) + (deals_sub * 2100) # residual
    hq_others_list.append(hq_others_val)
    
    total_hq_payments_list.append(hq_fix + hq_deal + hq_crm + hq_others_val)
    
    # 7. БИЗНЕС-НАЛОГИ
    tax_month = total_revenue_month * (tax_rate / 100.0)
    taxes_usn_list.append(tax_month)
    
    # 8. КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX) Месячный тотал
    tot_capex_month = capex_franchise_list[i] + capex_renovation_list[i] + capex_equipment_list[i]
    total_capex_list.append(tot_capex_month)
    
    # ИТОГО РАСХОДЫ Месяца
    total_expenses_month = (agent_payouts + salaries_back + taxes_payroll + 
                            office_rent_custom + marketing_val + recruiting_val + cleaning_val + accounting_banking_val + communications_val + opex_others_val +
                            hq_fix + hq_deal + hq_crm + hq_others_val + tax_month)
    
    # Чистая прибыль за месяц (Revenue - Expenses - CAPEX)
    net_profit_month = total_revenue_month - total_expenses_month - tot_capex_month
    adj_net_profit.append(net_profit_month)

# Расчет накопленного денежного потока начиная с Месяца 0 (CAPEX + OPEX старт)
capex_m0_franchise = regional_pau_fee + 17000 # паушальный + роспатент
capex_m0_renovation = 520000 if scenario_name == "Оптимальный" else 0
capex_m0_equipment = 815000 if scenario_name == "Оптимальный" else 0
total_capex_m0 = capex_m0_franchise + capex_m0_renovation + capex_m0_equipment

cum_flow = []
running_sum = -total_capex_m0 - start_opex
cum_flow.append(running_sum) # месяц 0

for i in range(24):
    running_sum += adj_net_profit[i]
    cum_flow.append(running_sum)

# ----------------------------------------------------
# KPI И ПОКАЗАТЕЛИ ЭФФЕКТИВНОСТИ
# ----------------------------------------------------
total_revenue_2years = sum(rev_total_list)
total_profit_2years = sum(adj_net_profit) - total_capex_m0 - start_opex
max_investment = abs(min(cum_flow))

# Месяц окупаемости (первый месяц, где cum_flow >= 0)
payback_month = "Н/Д"
for idx, val in enumerate(cum_flow):
    if val >= 0:
        payback_month = f"{idx} мес."
        break

# Точка безубыточности (первый месяц с положительной чистой прибылью)
breakeven_month = "Н/Д"
for idx, val in enumerate(adj_net_profit):
    if val > 0:
        breakeven_month = f"{idx + 1} мес."
        break

# Отображение KPI Карточек
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("💼 Общая выручка (2 года)", f"{total_revenue_2years:,.0f} ₽".replace(",", " "))
with col2:
    st.metric("📈 Итоговая чистая прибыль", f"{total_profit_2years:,.0f} ₽".replace(",", " "))
with col3:
    st.metric("📉 Пик инвестиций (CAPEX+OPEX)", f"{max_investment:,.0f} ₽".replace(",", " "))
with col4:
    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("🎯 Без-ость", breakeven_month)
    with col_b:
        st.metric("⏳ Окупаемость", payback_month)

# ----------------------------------------------------
# ТАБЫ: ГРАФИКИ, ДЕТАЛИЗАЦИЯ И СРАВНЕНИЕ
# ----------------------------------------------------
tab1, tab2, tab3, tab4 = st.tabs(["📊 Интерактивные Графики", "📋 Декомпозиция P&L (Все статьи)", "🌍 Сравнение Регионов", "📝 Описание Модели"])

with tab1:
    st.markdown("### Визуализация финансового развития")
    
    # 1. График накопленного денежного потока (J-Curve в светлых тонах бренда Century 21)
    fig_cum = go.Figure()
    fig_cum.add_trace(go.Scatter(
        x=list(range(25)),
        y=cum_flow,
        mode='lines+markers',
        name='Накопленный кэш-флоу',
        line=dict(color='#A19276', width=3), # Dark Gold
        marker=dict(size=6, color='#FFFFFF', line=dict(color='#252526', width=1.5)) # Obsessed Grey border
    ))
    fig_cum.add_shape(
        type="line", x0=0, y0=0, x1=24, y1=0,
        line=dict(color="#808285", width=1.5, dash="dash") # Medium Grey
    )
    fig_cum.update_layout(
        title="Кривая накопленного кэш-флоу (J-Curve)",
        xaxis_title="Месяцы проекта (0 - старт)",
        yaxis_title="Баланс кэша, ₽",
        template="plotly_white",
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        font=dict(color="#252526", family="Segoe UI, Arial"),
        xaxis=dict(gridcolor="#E6E7E8"),
        yaxis=dict(gridcolor="#E6E7E8")
    )
    st.plotly_chart(fig_cum, use_container_width=True)
    
    # 2. Столбчатый график Доходы vs Расходы (Светлая тема)
    fig_rev_exp = go.Figure()
    fig_rev_exp.add_trace(go.Bar(
        x=months,
        y=rev_total_list,
        name='Доходы (Комиссия ВКД)',
        marker_color='#BEAF87' # Relentless Gold
    ))
    
    expenses_no_capex = [rev_total_list[idx] - adj_net_profit[idx] - total_capex_list[idx] for idx in range(24)]
    fig_rev_exp.add_trace(go.Bar(
        x=months,
        y=expenses_no_capex,
        name='Операционные расходы + Налоги',
        marker_color='#252526' # Obsessed Grey
    ))
    fig_rev_exp.update_layout(
        title="Ежемесячные доходы и операционные расходы",
        xaxis_title="Месяцы",
        yaxis_title="Сумма, ₽",
        barmode='group',
        template="plotly_white",
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        font=dict(color="#252526", family="Segoe UI, Arial"),
        xaxis=dict(gridcolor="#E6E7E8"),
        yaxis=dict(gridcolor="#E6E7E8")
    )
    st.plotly_chart(fig_rev_exp, use_container_width=True)

    # 3. Круговая диаграмма структуры расходов
    avg_expenses = {
        'Выплаты агентам': sum(payouts_agents_list),
        'Оклады бэк-офиса': sum(salaries_backoffice_list),
        'Налоги на ФОТ оклады': sum(taxes_payroll_list),
        'Аренда офиса': sum(rent_list),
        'Реклама и маркетинг': sum(marketing_list),
        'HH.ru и подбор': sum(recruiting_list),
        'Роялти и CRM ЦО': sum(total_hq_payments_list),
        'УСН налоги': sum(taxes_usn_list),
        'Прочий OPEX': sum(cleaning_list) + sum(accounting_banking_list) + sum(communications_list) + sum(opex_others_list)
    }
    fig_pie = px.pie(
        names=list(avg_expenses.keys()),
        values=list(avg_expenses.values()),
        color_discrete_sequence=['#BEAF87', '#252526', '#A19276', '#808285', '#E6E7E8', '#CCCCCC', '#999999', '#D1D2D4', '#E1E1E1'],
        title="Детализированная структура расходов за 2 года"
    )
    fig_pie.update_layout(
        template="plotly_white",
        paper_bgcolor="#FFFFFF",
        font=dict(color="#252526", family="Segoe UI, Arial")
    )
    st.plotly_chart(fig_pie, use_container_width=True)

with tab2:
    st.markdown("### 📋 Детализированный отчет о прибылях и убытках (P&L)")
    
    # Построение иерархической таблицы данных
    p_and_l_decomposed = {
        'Категория / Статья': [
            '► ДОХОДЫ (ВКД)',
            '  ├─ Вторичный рынок',
            '  ├─ Первичный рынок',
            '  ├─ Аренда (жилая/коммерческая)',
            '  ├─ Загородная недвижимость',
            '  ├─ Доп. услуги (Ипотека/Страхование/Юр)',
            '  └─ Прочие доходы (МЛС и др.)',
            '★ ИТОГО ДОХОДЫ',
            
            '► РАСХОДЫ: ФОТ',
            '  ├─ Выплаты агентам (% комиссионных)',
            '  ├─ Оклады бэк-офиса (оклады)',
            '  └─ Налоги на ФОТ оклады',
            '★ ИТОГО ФОТ',
            
            '► РАСХОДЫ: OPEX',
            '  ├─ Аренда офиса',
            '  ├─ Реклама объектов и лидогенерация',
            '  ├─ Подбор персонала (HH.ru)',
            '  ├─ Уборка и жизнеобеспечение',
            '  ├─ Телефония, интернет, канцелярия',
            '  └─ Аутсорс бухгалтерии и банка',
            '  └─ Прочий OPEX (ГСМ, курьеры, праздники)',
            '★ ИТОГО OPEX',
            
            '► ВЫПЛАТЫ ЦО (ФРАНШИЗА)',
            '  ├─ Роялти FIX (вкл. НРФ)',
            '  ├─ Роялти со сделок',
            '  ├─ CRM и IT-платформа',
            '  └─ Прочие сервисы ЦО',
            '★ ИТОГО ВЫПЛАТЫ ЦО',
            
            '★ НАЛОГ УСН',
            '★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ',
            
            '► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)',
            '  ├─ Франшиза (Паушальный взнос + Роспатент)',
            '  ├─ Ремонт и Брендирование офиса',
            '  └─ Мебель, компьютеры и оборудование',
            '★ ИТОГО CAPEX',
            
            '★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ',
            '★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК'
        ]
    }
    
    # Заполняем таблицу по месяцам
    for idx, m in enumerate(months):
        p_and_l_decomposed[m] = [
            "", # Заголовок доходов
            f"{rev_secondary_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_primary_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_rent_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_suburban_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_services_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_others_list[idx]:,.0f} ₽".replace(",", " "),
            f"{rev_total_list[idx]:,.0f} ₽".replace(",", " "),
            
            "", # Заголовок ФОТ
            f"{payouts_agents_list[idx]:,.0f} ₽".replace(",", " "),
            f"{salaries_backoffice_list[idx]:,.0f} ₽".replace(",", " "),
            f"{taxes_payroll_list[idx]:,.0f} ₽".replace(",", " "),
            f"{total_payroll_list[idx]:,.0f} ₽".replace(",", " "),
            
            "", # Заголовок OPEX
            f"{rent_list[idx]:,.0f} ₽".replace(",", " "),
            f"{marketing_list[idx]:,.0f} ₽".replace(",", " "),
            f"{recruiting_list[idx]:,.0f} ₽".replace(",", " "),
            f"{cleaning_list[idx]:,.0f} ₽".replace(",", " "),
            f"{communications_list[idx]:,.0f} ₽".replace(",", " "),
            f"{accounting_banking_list[idx]:,.0f} ₽".replace(",", " "),
            f"{opex_others_list[idx]:,.0f} ₽".replace(",", " "),
            f"{total_opex_list[idx]:,.0f} ₽".replace(",", " "),
            
            "", # Заголовок франшизы
            f"{hq_royalty_fix_list[idx]:,.0f} ₽".replace(",", " "),
            f"{hq_royalty_deal_list[idx]:,.0f} ₽".replace(",", " "),
            f"{hq_crm_list[idx]:,.0f} ₽".replace(",", " "),
            f"{hq_others_list[idx]:,.0f} ₽".replace(",", " "),
            f"{total_hq_payments_list[idx]:,.0f} ₽".replace(",", " "),
            
            f"{taxes_usn_list[idx]:,.0f} ₽".replace(",", " "),
            f"{(total_payroll_list[idx] + total_opex_list[idx] + total_hq_payments_list[idx] + taxes_usn_list[idx]):,.0f} ₽".replace(",", " "),
            
            "", # Заголовок CAPEX
            f"{capex_franchise_list[idx]:,.0f} ₽".replace(",", " "),
            f"{capex_renovation_list[idx]:,.0f} ₽".replace(",", " "),
            f"{capex_equipment_list[idx]:,.0f} ₽".replace(",", " "),
            f"{total_capex_list[idx]:,.0f} ₽".replace(",", " "),
            
            f"{adj_net_profit[idx]:,.0f} ₽".replace(",", " "),
            f"{cum_flow[idx+1]:,.0f} ₽".replace(",", " ") # индекс + 1 т.к. в cum_flow есть Месяц 0
        ]
        
    df_p_and_l_decomposed = pd.DataFrame(p_and_l_decomposed)
    st.dataframe(df_p_and_l_decomposed, use_container_width=True, height=600)
    
    # ----------------------------------------------------
    # EXCEL EXPORT С ПОЛНОЙ ДЕКОМПОЗИЦИЕЙ
    # ----------------------------------------------------
    if openpyxl_available:
        def generate_excel():
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "C21 Decomposed P&L"
            
            ws.views.sheetView[0].showGridLines = True
            
            # Цветовые схемы Century 21 согласно BrandBook
            gold_fill = PatternFill(start_color="BEAF87", end_color="BEAF87", fill_type="solid") # Relentless Gold
            dark_fill = PatternFill(start_color="252526", end_color="252526", fill_type="solid") # Obsessed Grey
            light_gold_fill = PatternFill(start_color="F9F6EE", end_color="F9F6EE", fill_type="solid") # Warm Cream
            category_fill = PatternFill(start_color="E6E7E8", end_color="E6E7E8", fill_type="solid") # Light Grey
            
            font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            font_header = Font(name="Calibri", size=11, bold=True, color="111111")
            font_bold = Font(name="Calibri", size=11, bold=True)
            font_regular = Font(name="Calibri", size=11)
            
            thin_border_side = Side(border_style="thin", color="D3D3D3")
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            double_bottom_border = Border(bottom=Side(border_style="double", color="111111"), top=thin_border_side)
            
            # Заголовок листа
            ws.merge_cells("A1:Y1")
            ws["A1"] = f"ДЕТАЛИЗИРОВАННАЯ ФИНАНСОВАЯ МОДЕЛЬ CENTURY 21 - РЕГИОН: {region_select.upper()} ({scenario_name.upper()})"
            ws["A1"].font = font_title
            ws["A1"].fill = dark_fill
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40
            
            # Столбцы
            headers = ["Показатель / Месяцы"] + months
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = gold_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[3].height = 28
            
            # Месячный экспорт строк
            metrics_rows = [
                ('► ДОХОДЫ (ВКД)', [""], ""),
                ('  ├─ Вторичный рынок', rev_secondary_list, "#,##0"),
                ('  ├─ Первичный рынок', rev_primary_list, "#,##0"),
                ('  ├─ Аренда (жилая/коммерческая)', rev_rent_list, "#,##0"),
                ('  ├─ Загородная недвижимость', rev_suburban_list, "#,##0"),
                ('  ├─ Доп. услуги (Ипотека/Страхование/Юр)', rev_services_list, "#,##0"),
                ('  └─ Прочие доходы (МЛС и др.)', rev_others_list, "#,##0"),
                ('★ ИТОГО ДОХОДЫ', rev_total_list, "#,##0"),
                
                ('► РАСХОДЫ: ФОТ', [""], ""),
                ('  ├─ Выплаты агентам (% комиссионных)', payouts_agents_list, "#,##0"),
                ('  ├─ Оклады бэк-офиса (оклады)', salaries_backoffice_list, "#,##0"),
                ('  └─ Налоги на ФОТ оклады', taxes_payroll_list, "#,##0"),
                ('★ ИТОГО ФОТ', total_payroll_list, "#,##0"),
                
                ('► РАСХОДЫ: OPEX', [""], ""),
                ('  ├─ Аренда офиса', rent_list, "#,##0"),
                ('  ├─ Реклама объектов и лидогенерация', marketing_list, "#,##0"),
                ('  ├─ Подбор персонала (HH.ru)', recruiting_list, "#,##0"),
                ('  ├─ Уборка и жизнеобеспечение', cleaning_list, "#,##0"),
                ('  ├─ Телефония, internet, канцелярия', communications_list, "#,##0"),
                ('  ├─ Аутсорс бухгалтерии и банка', accounting_banking_list, "#,##0"),
                ('  └─ Прочий OPEX (ГСМ, курьеры, праздники)', opex_others_list, "#,##0"),
                ('★ ИТОГО OPEX', total_opex_list, "#,##0"),
                
                ('► ВЫПЛАТЫ ЦО (ФРАНШИЗА)', [""], ""),
                ('  ├─ Роялти FIX (вкл. НРФ)', hq_royalty_fix_list, "#,##0"),
                ('  ├─ Роялти со сделок', hq_royalty_deal_list, "#,##0"),
                ('  ├─ CRM и IT-платформа', hq_crm_list, "#,##0"),
                ('  └─ Прочие сервисы ЦО', hq_others_list, "#,##0"),
                ('★ ИТОГО ВЫПЛАТЫ ЦО', total_hq_payments_list, "#,##0"),
                
                ('★ НАЛОГ УСН', taxes_usn_list, "#,##0"),
                ('★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ', [total_payroll_list[idx]+total_opex_list[idx]+total_hq_payments_list[idx]+taxes_usn_list[idx] for idx in range(24)], "#,##0"),
                
                ('► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)', [""], ""),
                ('  ├─ Франшиза (Паушальный взнос + Роспатент)', capex_franchise_list, "#,##0"),
                ('  ├─ Ремонт и Брендирование офиса', capex_renovation_list, "#,##0"),
                ('  └─ Мебель, компьютеры и оборудование', capex_equipment_list, "#,##0"),
                ('★ ИТОГО CAPEX', total_capex_list, "#,##0"),
                
                ('★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ', adj_net_profit, "#,##0"),
                ('★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК', cum_flow[1:], "#,##0")
            ]
            
            for r_idx, (m_name, data_list, num_format) in enumerate(metrics_rows, 4):
                is_header = "►" in m_name
                is_total = "★" in m_name
                
                cell_a = ws.cell(row=r_idx, column=1, value=m_name)
                cell_a.font = font_bold if (is_header or is_total) else font_regular
                cell_a.border = thin_border
                
                if is_header:
                    cell_a.fill = category_fill
                elif is_total:
                    cell_a.fill = light_gold_fill
                
                for c_idx in range(1, 25):
                    # Если строка заголовка, пропускаем заполнение столбцов значениями
                    if is_header:
                        cell_val = ws.cell(row=r_idx, column=c_idx+1, value="")
                        cell_val.fill = category_fill
                        cell_val.border = thin_border
                        continue
                        
                    val = data_list[c_idx-1] if len(data_list) >= c_idx else 0
                    cell_val = ws.cell(row=r_idx, column=c_idx+1, value=val)
                    cell_val.number_format = num_format
                    cell_val.alignment = Alignment(horizontal="right")
                    cell_val.border = thin_border
                    
                    if is_total:
                        cell_val.font = font_bold
                        cell_val.fill = light_gold_fill
                        if "ЧИСТАЯ ПРИБЫЛЬ" in m_name:
                            cell_val.border = double_bottom_border
                    else:
                        cell_val.font = font_regular
                
                ws.row_dimensions[r_idx].height = 20
                
            # Ширина столбца A
            ws.column_dimensions["A"].width = 45
            for col in range(2, 26):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 15
                
            wb.save(output)
            return output.getvalue()
            
        excel_data = generate_excel()
        st.download_button(
            label="📥 Скачать детализированный P&L в Excel",
            data=excel_data,
            file_name=f"C21_P_and_L_Decomposed_{region_select.split(' ')[0]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("⚠️ **Экспорт в Excel временно недоступен.** Требуется библиотека `openpyxl`. "
                   "Скопируйте содержимое вашего файла `req_fin.txt` в стандартный файл `requirements.txt` в вашем репозитории на GitHub.")

with tab3:
    st.markdown("### 🌍 Сравнение ключевых показателей регионов")
    st.write("Сравнение базовых финансово-экономических параметров по региональным сегментам Century 21:")
    
    comp_data = {
        'Региональный сегмент': list(REGIONAL_PRESETS.keys()),
        'Паушальный взнос (база)': [f"{v['pau_fee']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Вторичка': [f"{v['comm_sec']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Первичка': [f"{v['comm_prim']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Загородная': [f"{v['comm_sub']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Оклады бэк-офиса': [f"{v['salary_mult']*100:.0f}% к Москве" for v in REGIONAL_PRESETS.values()],
        'Аренда офиса (база)': [f"{v['rent_base']:,} ₽/мес.".replace(",", " ") for v in REGIONAL_PRESETS.values()],
    }
    df_comp = pd.DataFrame(comp_data)
    st.table(df_comp)
    
    st.markdown("""
    💡 **Инсайт для клиента:** 
    Как видно из таблицы, хотя региональные рынки имеют меньший средний чек по сделкам, они значительно выигрывают за счет **низких операционных затрат (ФОТ оклады и Аренда офиса)**. 
    Благодаря этому региональные агентства сохраняют высокую операционную маржинальность и часто окупаются еще быстрее, чем столичные!
    """)

with tab4:
    st.markdown("""
    ### 📖 Справочник и методология декомпозиции
    
    Эта интерактивная модель разделяет финансовый план агентства недвижимости **Century 21** на 5 ключевых блоков:
    
    1. **Доходы (Комиссия ВКД):** 
       * Рассчитывается динамически на основе воронки лидов. 
       * Включает доп. сервисы — **комиссионные вознаграждения от банков по ипотеке, страховых компаний и юр. услуг**.
       
    2. **ФОТ (Зарплатный фонд):**
       * **Переменная часть (агенты):** Прямо привязана к объему сделок и настраивается слайдером (базово 38%).
       * **Постоянная часть (оклады):** Включает фиксированные оклады РОПа, HR, юристов, рекрутеров, скорректированные под выбранный регион.
       
    3. **OPEX (Операционные расходы):**
       * Включает аренду офиса, затраты на лидогенерацию (маркетинг), сервисы по набору персонала (HeadHunter), а также жизнеобеспечение (уборка, интернет, связь, канцелярия).
       
    4. **Выплаты ЦО (Франшиза):**
       * Фиксированные роялти (шкала по месяцам), переменные роялти с каждой сделки (настраиваются слайдером) и IT-сервисы ЦО.
       
    5. **CAPEX (Капитальные вложения):**
       * Регистрационные расходы, запуск офиса (ремонт, мебель, компьютеры) на старте и календарное расширение рабочих мест во второй год.
    """)
