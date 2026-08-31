import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import io

# Проверяем доступность openpyxl для экспорта в Excel без падения приложения
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    openpyxl_available = True
except ImportError:
    openpyxl_available = False

# Установка конфигурации страницы
st.set_page_config(
    page_title="CENTURY 21 Financial Model (Бюджет Оптимальный) v14",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ----------------------------------------------------
# ПОМОЩНИКИ ДЛЯ РУЧНЫХ КОРРЕКТИРОВОК P&L, ДРАЙВЕРОВ И КАДРОВ
# ----------------------------------------------------
if 'pl_overrides' not in st.session_state:
    st.session_state['pl_overrides'] = {}
if 'driver_overrides' not in st.session_state:
    st.session_state['driver_overrides'] = {}
if 'headcount_overrides' not in st.session_state:
    st.session_state['headcount_overrides'] = {}

def get_override(metric_name, col_name, default_val):
    key = (metric_name, col_name)
    if key in st.session_state['pl_overrides']:
        return st.session_state['pl_overrides'][key]
    return default_val

def get_driver_override(driver_name, col_name, default_val):
    key = (driver_name, col_name)
    if key in st.session_state['driver_overrides']:
        return st.session_state['driver_overrides'][key]
    return default_val

def get_headcount_override(role_name, col_name, default_val):
    key = (role_name, col_name)
    if key in st.session_state['headcount_overrides']:
        return st.session_state['headcount_overrides'][key]
    return default_val

def parse_value(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(' ', '').replace('₽', '').replace(',', '.').strip()
    if s == "" or s == "-" or s == "0":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


# ----------------------------------------------------
# АВТОРИЗАЦИЯ ПОЛЬЗОВАТЕЛЯ
# ----------------------------------------------------
if 'authenticated' not in st.session_state:
    st.session_state['authenticated'] = False

if not st.session_state['authenticated']:
    # Beautiful custom styling for login
    st.markdown("""
    <style>
        /* Импорт шрифта */
        @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@300;400;600;700;800&display=swap');
        
        html, body, [data-testid="stAppViewContainer"] {
            font-family: 'Segoe UI', Arial, sans-serif;
            background-color: #FFFFFF;
            color: #252526;
        }
        
        .login-title {
            color: #252526;
            font-size: 26px;
            font-weight: 800;
            text-align: center;
            margin-top: 10px;
            margin-bottom: 5px;
            text-transform: uppercase;
            letter-spacing: 2px;
        }
        .brand-accent {
            color: #BEAF87; /* Relentless Gold */
        }
        .login-sub {
            color: #808285; /* Medium Grey */
            font-size: 14px;
            text-align: center;
            margin-bottom: 30px;
        }
        div.stButton > button {
            background-color: #BEAF87 !important;
            color: #FFFFFF !important;
            font-weight: bold !important;
            border: none !important;
            border-radius: 5px !important;
            padding: 12px 24px !important;
            transition: all 0.3s ease !important;
            width: 100% !important;
            box-shadow: 0 2px 5px rgba(190, 175, 135, 0.3) !important;
        }
        div.stButton > button:hover {
            background-color: #A19276 !important; /* Dark Gold */
            box-shadow: 0 4px 12px rgba(161, 146, 118, 0.5) !important;
            transform: translateY(-1px);
        }
        
        /* Контейнер для монограммы с соблюдением охранного поля 0.5X */
        .brand-monogram-container {
            display: block;
            margin: 35px auto; /* Охранное поле 0.5X для монограммы высотой 70px */
            text-align: center;
        }
    </style>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1.4, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        # Встраиваем монограмму C21 согласно Brandbook (в один цвет Relentless Gold, с охранным полем 0.5X)
        st.markdown("""
        <div class="brand-monogram-container">
            <svg viewBox="0 0 100 100" width="70" height="70" style="display: block; margin: 0 auto;">
                <path d="M 75,25 A 35,35 0 1,0 75,75" fill="none" stroke="#BEAF87" stroke-width="8" stroke-linecap="round" />
                <text x="48" y="59" font-family="'Segoe UI', Arial, sans-serif" font-weight="bold" font-size="26" fill="#BEAF87" text-anchor="middle">21</text>
            </svg>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div class='login-title'>CENTURY 21 <span class='brand-accent'>Financial MVP</span></div>", unsafe_allow_html=True)
        st.markdown("<div class='login-sub'>Интерактивная финансовая модель | Кейс: Бюджет Оптимальный</div>", unsafe_allow_html=True)
        
        with st.container(border=True):
            st.markdown("<h4 style='text-align: center; color: #252526; font-weight: 700; margin-bottom: 20px;'>Вход в личный кабинет</h4>", unsafe_allow_html=True)
            username = st.text_input("Логин", placeholder="Введите имя пользователя")
            password = st.text_input("Пароль", type="password", placeholder="Введите ваш пароль")
            
            st.markdown("<div style='margin-top: 20px;'>", unsafe_allow_html=True)
            if st.button("Войти в систему"):
                if username == "admin" and password == "c21admin":
                    st.session_state['authenticated'] = True
                    st.success("Успешная авторизация! Загрузка модели...")
                    st.rerun()
                else:
                    st.error("Неверный логин или пароль. Попробуйте еще раз.")
            st.markdown("</div>", unsafe_allow_html=True)
        st.stop()

# Пользовательский CSS-стилинг в корпоративных цветах Century 21 (Premium Light Theme)
st.markdown("""
<style>
    /* Импорт шрифта */
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@300;400;600;700;800&display=swap');
    
    html, body, [data-testid="stAppViewContainer"] {
        font-family: 'Segoe UI', Arial, sans-serif;
        background-color: #FFFFFF;
        color: #252526;
    }
    
    /* Стилизация боковой панели */
    [data-testid="stSidebar"] {
        background-color: #E6E7E8 !important;
        border-right: 1px solid #D3D3D3;
        padding-top: 20px;
    }
    
    [data-testid="stSidebar"] * {
        color: #252526 !important;
    }
    
    /* Красивые карточки KPI */
    div[data-testid="stMetric"] {
        background-color: #F8F9FA;
        border: 1px solid #E6E7E8;
        border-left: 5px solid #BEAF87;
        padding: 15px 20px;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    
    div[data-testid="stMetric"] label {
        color: #777779 !important;
        font-size: 13px !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        color: #252526 !important;
        font-size: 22px !important;
        font-weight: 700 !important;
        margin-top: 5px;
    }
    
    /* Кнопки */
    div.stButton > button {
        background-color: #BEAF87 !important;
        color: #FFFFFF !important;
        font-weight: bold !important;
        border: none !important;
        border-radius: 5px !important;
        padding: 12px 24px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 2px 5px rgba(190, 175, 135, 0.3) !important;
        width: 100% !important;
    }
    
    div.stButton > button:hover {
        background-color: #A19276 !important;
        color: #FFFFFF !important;
        box-shadow: 0 4px 12px rgba(161, 146, 118, 0.5) !important;
        transform: translateY(-1px);
    }
    
    /* Заголовки */
    .main-title {
        color: #252526;
        font-size: 30px;
        font-weight: 800;
        text-align: center;
        margin-top: 10px;
        margin-bottom: 5px;
        letter-spacing: 0.5px;
    }
    
    .brand-accent {
        color: #A19276;
    }
    
    .sub-title {
        color: #777779;
        font-size: 15px;
        text-align: center;
        margin-bottom: 30px;
        font-weight: 400;
    }
    
    /* Вкладки */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        border-bottom: 2px solid #E6E7E8;
    }
    
    .stTabs [data-baseweb="tab"] {
        background-color: #F8F9FA;
        border: 1px solid #E6E7E8;
        border-bottom: none;
        border-radius: 6px 6px 0px 0px;
        padding: 10px 24px;
        color: #777779;
        font-weight: 600;
        transition: all 0.2s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #FFFFFF !important;
        color: #A19276 !important;
        border-top: 3px solid #BEAF87 !important;
        border-left: 1px solid #BEAF87 !important;
        border-right: 1px solid #BEAF87 !important;
    }
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# ОПРЕДЕЛЕНИЕ ДАННЫХ ЛИСТА "БЮДЖЕТ ОПТИМАЛЬНЫЙ" (C21)
# ----------------------------------------------------

months = [str(i) for i in range(1, 25)]

optimal_baseline = {
    'leads': [440, 880, 1320, 1760, 2200, 2640, 3080, 3520, 3960, 4400, 4840, 5280,
              6160, 6600, 7040, 7480, 7920, 8360, 8800, 9240, 9680, 10120, 10560, 11000],
    'meetings': [44, 88, 132, 176, 220, 264, 308, 352, 396, 440, 484, 528,
                 616, 660, 704, 748, 792, 836, 880, 924, 968, 1012, 1056, 1100],
    'contracts': [0, 2, 4, 7, 9, 11, 13, 15, 18, 20, 22, 24,
                  26, 31, 33, 35, 37, 40, 42, 44, 46, 48, 51, 53],
    'prepayments': [0.0, 1.5, 3.0, 5.3, 6.8, 8.3, 9.8, 11.3, 13.5, 15.0, 16.5, 18.0,
                    19.5, 23.3, 24.8, 26.3, 27.8, 30.0, 31.5, 33.0, 34.5, 36.0, 38.3, 39.8],
    
    # Декомпозиция ДОХОДОВ (ВКД)
    'rev_sec': [0, 0, 0, 360000, 1080000, 1603800, 2138400, 2673000, 3207600, 3742200, 4276800, 4811400,
                5346000, 5880600, 6415200, 7484400, 8019000, 8553600, 9088200, 9622800, 10157400, 10692000, 11226600, 11761200],
    'rev_prim': [0, 0, 0, 440000, 440000, 440000, 440000, 880000, 880000, 880000, 880000, 880000,
                 653400, 718740, 784080, 914760, 980100, 1045440, 1110780, 1176120, 1241460, 1306800, 1372140, 1437480],
    'rev_rent': [0, 0, 0, 0, 0, 80000, 0, 0, 80000, 0, 0, 80000,
                 240000, 240000, 280000, 280000, 320000, 320000, 360000, 400000, 440000, 440000, 480000, 480000],
    'rev_sub': [0, 0, 0, 0, 0, 500000, 0, 0, 500000, 0, 0, 500000,
                1000000, 1000000, 1000000, 1000000, 1500000, 1500000, 2000000, 2000000, 2500000, 2500000, 2500000, 2500000],
    'rev_overseas': [0]*17 + [220000] + [0]*4 + [220000] + [0],
    'rev_other_p': [0]*12 + [75600]*12,
    'rev_mort': [0]*8 + [106029, 123701, 141372, 159044, 176715, 194387, 212058, 247401, 265073, 282744, 300416, 318087, 335759, 353430, 371102, 388773],
    'rev_ins': [0]*8 + [27265, 31809, 36353, 40897, 45441, 49985, 54529, 63617, 68162, 72706, 77250, 81794, 86338, 90882, 95426, 99970],
    'rev_legal': [0]*6 + [138600, 161700, 184800, 207900, 231000, 254100, 277200, 323400, 346500, 369600, 392700, 415800, 438900, 462000, 485100, 508200, 531300, 554400],
    'revenue': [0, 0, 0, 800000, 1520000, 2623800, 2717000, 3714700, 4985694, 4985609, 5565525, 6725440,
                7814356, 8482712, 9167967, 10435378, 11620634, 12485890, 13451145, 14136401, 15321656, 15966912, 16872168, 17297423],

    # Декомпозиция РАСХОДОВ - ФОТ
    'salaries': [214935, 214935, 214935, 214935, 214935, 144935, 144935, 144935, 144935, 144935, 144935, 144935,
                 963068, 964023, 966913, 1126564, 1129454, 1462344, 1608235, 1611125, 1764015, 1766905, 1769796, 1772686],
    'salary_taxes': [44935]*12 + [143068, 144023, 146913, 156564, 159454, 162344, 208235, 211125, 214015, 216905, 219796, 222686],
    'agent_comm': [0, 0, 0, 338500, 612100, 1053409, 1314265, 1784222, 2332633, 2318694, 2629481, 3171647,
                   4103166, 4502709, 4855674, 5449734, 6035280, 6593359, 7120893, 7473269, 8061122, 8387655, 8822681, 9066321],

    # Декомпозиция OPEX
    'opex': [150000, 286300, 321300, 357400, 393500, 464600, 535700, 606800, 677900, 749000, 820100, 896200,
             569750, 601530, 769250, 1136970, 809690, 827910, 851130, 868850, 892070, 1399290, 927010, 944730],

    # Декомпозиция ВЫПЛАТ ЦО
    'hq_royalty_fix': [0, 0, 0, 45675, 45675, 45675, 45675, 78750, 78750, 78750, 78750, 78750,
                       111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825, 111825],
    'hq_royalty_deal': [0, 0, 0, 4200, 8400, 13556, 14574, 19793, 25011, 26030, 29148, 34367,
                        39134, 42564, 45994, 52855, 58385, 61816, 67346, 70776, 76307, 79737, 83167, 86598],
    'hq_crm_and_others': [19400, 5500, 5500, 68375, 72575, 77731, 78749, 117043, 188908, 201034, 215260, 231587,
                          194303, 191657, 195549, 202872, 208864, 278757, 218749, 222641, 228634, 232526, 302418, 240311],
    'capex': [0]*24
}

# Реальные базовые оклады backoffice специалистов для расчета при изменении кадровой структуры
BASE_SALARY_RATES = {
    'РОП: категория C': 70000.0,
    'РОП: категория B': 100000.0,
    'Администратор офиса': 70000.0,
    'HR/рекрутер': 100000.0,
    'РОС/тренер': 100000.0,
    'Юрист': 150000.0,
    'Ипотечный Брокер': 180000.0,
    'Листинг-менеджер': 50000.0,
    'Фотограф': 30000.0,
    'Маркетолог/SMM': 150000.0
}

# Дефолтная структура штатного расписания (чел.) по месяцам
DEFAULT_HEADCOUNTS = {
    'РОП: категория C': [1]*12 + [1]*12,
    'РОП: категория B': [0]*12 + [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 2],
    'Администратор офиса': [0]*12 + [1]*12,
    'HR/рекрутер': [1]*12 + [2, 2, 2, 2, 2, 2, 3, 3, 3, 3, 3, 3],
    'РОС/тренер': [0]*12 + [1]*12,
    'Юрист': [0]*12 + [1, 1, 1, 2, 2, 2, 2, 2, 3, 3, 3, 3],
    'Ипотечный Брокер': [0]*12 + [1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2, 2],
    'Листинг-менеджер': [0]*12 + [1]*12,
    'Фотограф': [1]*12 + [1]*12,
    'Маркетолог/SMM': [0]*5 + [1]*7 + [0]*5 + [1]*7,
    
    # Агенты
    'Агент: категория D': [1, 2, 3, 4, 5, 5, 6, 7, 8, 9, 9, 10] + [14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25],
    'Агент: категория C': [0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1, 1] + [4, 4, 5, 5, 5, 5, 5, 6, 6, 6, 7, 7],
    'Агент: категория B': [0]*10 + [1, 1] + [1, 1, 1, 1, 2, 2, 2, 2, 3, 3, 3, 3],
    'Агент: категория A': [0]*12 + [1, 1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2]
}

# Операционные драйверы (базовые массивы из Excel)
base_deals_rent = [3, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1, 3, 3, 4, 4, 4, 4, 5, 5, 6, 6, 6, 6]
base_deals_suburban = [0, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1, 2, 2, 2, 2, 3, 3, 4, 4, 5, 5, 5, 5]
base_deals_overseas = [0]*14 + [1] + [0]*4 + [1] + [0] + [2, 1, 0]
base_deals_other = [0]*12 + [0.3]*12
base_services_mortgage = [0]*8 + [1.5, 1.8, 2.0, 2.3] + [3, 3, 3, 4, 4, 4, 4, 5, 5, 5, 5, 6]
base_services_insurance = [0]*8 + [1.5, 1.8, 2.0, 2.3] + [3, 3, 3, 4, 4, 4, 4, 5, 5, 5, 5, 6]
base_services_legal = [0]*6 + [0.9, 1.1, 1.2, 1.4, 1.5, 1.7] + [2, 2, 2, 2, 3, 3, 3, 3, 3, 3, 4, 4]

# Региональные параметры и пресеты
REGIONAL_PRESETS = {
    "Москва": {
        "pau_fee": 577500,
        "comm_sec": 360000,
        "comm_prim": 440000,
        "comm_sub": 500000,
        "comm_rent": 80000,
        "rent_base": 150000,
        "salary_mult": 1.0,
        "royalty_fix_factor": 1.0,
        "royalty_deal": 2100,
    },
    "Московская область (МО)": {
        "pau_fee": 500000,
        "comm_sec": 280000,
        "comm_prim": 330000,
        "comm_sub": 400000,
        "comm_rent": 60000,
        "rent_base": 100000,
        "salary_mult": 0.85,
        "royalty_fix_factor": 0.85,
        "royalty_deal": 2000,
    },
    "Крупные города (население > 800к)": {
        "pau_fee": 450000,
        "comm_sec": 220000,
        "comm_prim": 260000,
        "comm_sub": 320000,
        "comm_rent": 45000,
        "rent_base": 80000,
        "salary_mult": 0.75,
        "royalty_fix_factor": 0.75,
        "royalty_deal": 2000,
    },
    "Средние города (население 400к - 800к)": {
        "pau_fee": 350000,
        "comm_sec": 160000,
        "comm_prim": 200000,
        "comm_sub": 240000,
        "comm_rent": 35000,
        "rent_base": 55000,
        "salary_mult": 0.60,
        "royalty_fix_factor": 0.60,
        "royalty_deal": 1800,
    },
    "Малые города (население < 400к)": {
        "pau_fee": 270000,
        "comm_sec": 110000,
        "comm_prim": 140000,
        "comm_sub": 180000,
        "comm_rent": 25000,
        "rent_base": 38000,
        "salary_mult": 0.45,
        "royalty_fix_factor": 0.45,
        "royalty_deal": 1500,
    }
}

# ----------------------------------------------------
# ШАПКА ПРИЛОЖЕНИЯ
# ----------------------------------------------------
st.markdown("<div class='main-title'>💼 CENTURY 21 <span class='brand-accent'>Optimal Budget Case</span></div>", unsafe_allow_html=True)
st.markdown("<div class='sub-title'>Интерактивная финансовая модель и декомпозиция доходов/расходов на основе листа «Бюджет Оптимальный»</div>", unsafe_allow_html=True)

# ----------------------------------------------------
# БОКОВАЯ ПАНЕЛЬ С УПРАВЛЕНИЕМ & ТУЛТИПАМИ
# ----------------------------------------------------
st.sidebar.markdown("""
<div style='text-align: center; margin: 30px auto; display: block;'>
    <svg viewBox="0 0 100 100" width="60" height="60" style="display: block; margin: 0 auto;">
        <path d="M 75,25 A 35,35 0 1,0 75,75" fill="none" stroke="#252526" stroke-width="8" stroke-linecap="round" />
        <text x="48" y="59" font-family="'Segoe UI', Arial, sans-serif" font-weight="bold" font-size="26" fill="#252526" text-anchor="middle">21</text>
    </svg>
</div>
""", unsafe_allow_html=True)

st.sidebar.markdown("<h2 style='color:#252526; text-align:center; font-weight:800; margin-top: 10px;'>Управление моделью</h2>", unsafe_allow_html=True)

if st.sidebar.button("🔓 Выйти из системы", use_container_width=True):
    st.session_state['authenticated'] = False
    st.rerun()

st.sidebar.markdown("---")

# Регион
region_select = st.sidebar.selectbox(
    "🌍 Выберите регион / тип города:",
    options=list(REGIONAL_PRESETS.keys()),
    help="Каждый пресет автоматически настраивает средние чеки сделок, стоимость аренды офиса, оклады бэк-офиса и fixed-роялти франшизы под экономический уровень выбранной территории."
)
preset = REGIONAL_PRESETS[region_select]

# Фиксированный сценарий "Бюджет Оптимальный"
st.sidebar.info("📌 **Выбран кейс: Бюджет Оптимальный (Интенсивный старт)**\\n\\nДанный кейс моделирует агрессивный запуск с первого месяца: в Месяце 0 производится закупка оборудования и ремонт офиса на сумму 1 335 000 ₽, а воронка масштабируется на базе 440 базовых лидов на старте с выходом на 11 000 лидов к 24 месяцу.")

st.sidebar.markdown("### 🎚️ Настройки калибровки")

# Разделы параметров
with st.sidebar.expander("📈 1. Воронка продаж", expanded=True):
    scaling_leads = st.slider(
        "Количество лидов (звонков), % к базе",
        min_value=50, max_value=200, value=100, step=10,
        help="Регулирует входящий поток клиентов. Масштабирует доходы от сделок по вторичному и первичному рынку."
    )
    conversion_meeting = st.slider(
        "Конверсия лид -> встреча, %",
        min_value=2.0, max_value=20.0, value=10.0, step=0.5,
        help="Какая часть позвонивших клиентов доходит до личной встречи в офисе. Базовый корпоративный стандарт Century 21 составляет 10%."
    )
    conversion_deal = st.slider(
        "Конверсия встреча -> договор, %",
        min_value=1.0, max_value=30.0, value=5.0, step=0.5,
        help="Эффективность работы агентов при переговорах. Показывает долю встреч, переходящих в подписанные эксклюзивные договоры. Базовый уровень для оптимального бюджета равен 5%."
    )

with st.sidebar.expander("💰 2. Средние комиссии", expanded=False):
    comm_secondary = st.number_input(
        "Вторичный рынок, ₽", min_value=50000, max_value=1000000, value=preset["comm_sec"], step=10000,
        help="Средний комиссионный доход агентства с одной сделки купли-продажи на вторичном рынке недвижимости."
    )
    comm_primary = st.number_input(
        "Первичный рынок, ₽", min_value=50000, max_value=1000000, value=preset["comm_prim"], step=10000,
        help="Средняя комиссия, получаемая от девелоперов за продажу новостройки."
    )
    comm_suburban = st.number_input(
        "Загородная недвижимость, ₽", min_value=50000, max_value=1000000, value=preset["comm_sub"], step=10000,
        help="Средний комиссионный чек со сделок по загородным домам, коттеджам и участкам."
    )
    comm_rent_val = st.number_input(
        "Аренда (жилая/коммерческая), ₽", min_value=10000, max_value=200000, value=preset["comm_rent"], step=5000,
        help="Комиссионный доход со сделок аренды жилой и коммерческой недвижимости."
    )

with st.sidebar.expander("👥 3. Персонал и Выплаты", expanded=False):
    agent_commission_pct = st.slider(
        "Средний % выплат агентам", min_value=25, max_value=60, value=38, step=1,
        help="Средний процент от комиссии по сделкам, выплачиваемый агентам. Отражает мотивационную сетку Century 21 (35%-50% в зависимости от категории)."
    )
    backoffice_salary_mult = st.slider(
        "Индекс окладов бэк-офиса, % к базе", min_value=50, max_value=150, value=100, step=5,
        help="Шкала фиксированных окладов бэк-офиса (РОП, HR/рекрутер, юрист, ипотечный брокер, листинг-менеджер, администратор)."
    )

with st.sidebar.expander("🏢 4. OPEX и Налоги", expanded=False):
    office_rent_custom = st.number_input(
        "Аренда офиса в месяц, ₽", min_value=10000, max_value=500000, value=preset["rent_base"], step=5000,
        help="Стоимость аренды офисного помещения. Применяются поправки на ранних этапах запуска офиса."
    )
    tax_rate = st.sidebar.slider(
        "Ставка налога УСН, %", min_value=3, max_value=15, value=7, step=1,
        help="Действующая налоговая ставка по УСН (доходы) с учетом региональных субсидий и льгот."
    )
    marketing_mult = st.slider(
        "Реклама и маркетинг, % к базе", min_value=50, max_value=200, value=100, step=10,
        help="Регулирует объем затрат на рекламу объектов и лидогенерацию."
    )

with st.sidebar.expander("🛠️ 5. Регулировка CAPEX", expanded=False):
    capex_renovation_val = st.number_input("Ремонт и брендирование (всего), ₽", min_value=100000, max_value=2000000, value=520000, step=50000, help="Затраты на ремонт офиса и рекламные вывески.")
    capex_equipment_val = st.number_input("Мебель и оборудование (всего), ₽", min_value=100000, max_value=2000000, value=815000, step=50000, help="Столы, стулья, компьютеры, принтеры и роутеры на старте.")
    pau_fee_val = st.number_input("Паушальный взнос (с НДС), ₽", min_value=150000, max_value=1500000, value=preset["pau_fee"], step=10000)

with st.sidebar.expander("🏢 6. Выплаты в ЦО", expanded=False):
    royalty_deal_custom = st.number_input(
        "Роялти за каждую сделку, ₽", min_value=1000, max_value=5000, value=preset["royalty_deal"], step=100,
        help="Фиксированный роялти за сделку. По умолчанию зафиксирован на уровне 2 100 ₽."
    )

# ----------------------------------------------------
# ВЫЧИСЛИТЕЛЬНОЕ ЯДРО МОДЕЛИ (Декомпозиция Оптимальный)
# ----------------------------------------------------
base = optimal_baseline.copy()
salary_mult = preset["salary_mult"]
royalty_fix_mult = preset["royalty_fix_factor"]

# Инициализация всех массивов
adj_leads = []
adj_meetings = []
adj_contracts = []
adj_prepayments = []

rev_secondary_list = []
rev_primary_list = []
rev_rent_list = []
rev_suburban_list = []
rev_overseas_list = []
rev_other_p_list = []
rev_mortgage_list = []
rev_insurance_list = []
rev_legal_list = []
rev_total_list = []

# Декомпозированный ФОТ
agent_D_list = []
agent_C_list = []
agent_B_list = []
agent_A_list = []
payouts_agents_list = []

rop_C_list = []
rop_B_list = []
admin_list = []
hr_list = []
ros_list = []
jurist_list = []
mort_broker_list = []
listing_list = []
photographer_list = []
smm_list = []
salaries_backoffice_list = []

taxes_payroll_list = []
total_payroll_list = []

# Расходы - OPEX
rent_list = []
internet_list = []
mobile_list = []
kanc_list = []
reklama_list = []
hh_list = []
buh_list = []
bank_list = []
cleaning_list = []
gsm_list = []
courier_list = []
events_list = []
total_opex_list = []

# Выплаты ЦО
hq_royalty_fix_list = []
hq_royalty_deal_list = []
hq_crm_list = []
hq_kc_list = []
total_hq_payments_list = []

taxes_usn_list = []
capex_franchise_list = [0]*24
capex_renovation_list = [0]*24
capex_equipment_list = [0]*24
total_capex_list = []
adj_net_profit = []

capex_renovation_list[14] = 520000
capex_equipment_list[14] = 815000
capex_equipment_list[19] = 790000

leads_mult = scaling_leads / 100.0
conv_meet_rate = conversion_meeting / 10.0
conv_deal_rate = conversion_deal / 5.0
overall_voeronka_factor = leads_mult * conv_meet_rate * conv_deal_rate

base_internet = [5000]*12 + [950]*12
base_mobile = [7800]*2 + [8400] + [9000] + [9600] + [10200] + [10800] + [11400] + [12000] + [12600] + [13200] + [13800] + [18000] + [18480] + [19800] + [21120] + [22440] + [23760] + [25080] + [26400] + [27720] + [29040] + [30360] + [31680]
base_rent_list = [150000]*12 + [150000]*2 + [300000]*10
base_kanc = [6500]*2 + [7000] + [7500] + [8000] + [8500] + [9000] + [9500] + [10000] + [10500] + [11000] + [11500] + [15000] + [15500] + [16500] + [17500] + [18500] + [20000] + [21500] + [22500] + [24000] + [24500] + [25500] + [26500]
base_reklama = [35000] + [70000] + [105000] + [140000] + [210000] + [280000] + [350000] + [420000] + [490000] + [560000] + [630000] + [700000] + [184800] + [215600] + [231000] + [246400] + [261800] + [277200] + [292600] + [308000] + [323400] + [338800] + [354200] + [369600]
base_hh = [45000]*12 + [90000]*12
base_buh = [20000]*12 + [40000]*12
base_bank = [2000]*12 + [6000]*12
base_cleaning = [15000]*12 + [45000]*12
base_gsm = [0]*10 + [5000]*2 + [10000]*4 + [15000]*2 + [20000]*2 + [25000]*4
base_courier = [0]*11 + [10000]*13

base_events = [0]*24
base_events[11] = 100000
base_events[15] = 350000
base_events[21] = 490000

for i in range(24):
    # 1. Лиды и воронка (с учетом ручной коррекции в Tab 3)
    default_leads_m = int(base['leads'][i] * leads_mult)
    L = int(get_driver_override('  ├─ Звонки/лиды', str(i+1), default_leads_m))
    adj_leads.append(L)
    
    default_meetings_m = int(base['meetings'][i] * leads_mult * conv_meet_rate)
    meetings = int(get_driver_override('  ├─ Встречи', str(i+1), default_meetings_m))
    adj_meetings.append(meetings)
    
    default_contracts_m = int(base['contracts'][i] * overall_voeronka_factor)
    contracts = int(get_driver_override('  ├─ Договоры', str(i+1), default_contracts_m))
    adj_contracts.append(contracts)
    
    default_prepayments_m = base['prepayments'][i] * overall_voeronka_factor
    prepayments = float(get_driver_override('  ├─ Задаток/аванс', str(i+1), default_prepayments_m))
    adj_prepayments.append(prepayments)
    
    # Сделки в штуках (динамические по умолчанию)
    deals_sec_vol = float(get_driver_override('  ├─ Сделки: вторичный рынок', str(i+1), prepayments * 0.90))
    deals_prim_vol = float(get_driver_override('  ├─ Сделки: первичный рынок', str(i+1), prepayments * 0.10))
    deals_rent_vol = float(get_driver_override('  ├─ Сделки: аренда', str(i+1), base_deals_rent[i]))
    deals_sub_vol = float(get_driver_override('  ├─ Сделки: загородная', str(i+1), base_deals_suburban[i]))
    deals_overseas_vol = float(get_driver_override('  ├─ Сделки: зарубежная', str(i+1), base_deals_overseas[i]))
    deals_other_vol = float(get_driver_override('  ├─ Сделки: прочее (МЛС, срочновыкуп, сайт)', str(i+1), base_deals_other[i]))
    
    serv_mort_vol = float(get_driver_override('  ├─ Сервисы: ипотека', str(i+1), base_services_mortgage[i]))
    serv_ins_vol = float(get_driver_override('  ├─ Сервисы: страхование', str(i+1), base_services_insurance[i]))
    serv_legal_vol = float(get_driver_override('  └─ Сервисы: юр. сопровождение', str(i+1), base_services_legal[i]))
    
    # 2. ДОХОДЫ (1-3 месяцы строго равны нулю, если не скорректированы вручную в P&L)
    if i < 3:
        rev_sec = get_override('  ├─ Вторичный рынок', str(i+1), 0.0)
        rev_prim = get_override('  ├─ Первичный рынок', str(i+1), 0.0)
        rev_rent = get_override('  ├─ Аренда (жилая/коммерческая)', str(i+1), 0.0)
        rev_sub = get_override('  ├─ Загородная недвижимость', str(i+1), 0.0)
        rev_overseas = get_override('  ├─ Зарубежная недвижимость', str(i+1), 0.0)
        rev_other_p = get_override('  ├─ Сделки: прочее (МЛС, срочновыкуп)', str(i+1), 0.0)
        rev_mort = get_override('  ├─ Сервисы: ипотека', str(i+1), 0.0)
        rev_ins = get_override('  ├─ Сервисы: страхование', str(i+1), 0.0)
        rev_legal = get_override('  └─ Сервисы: юр. сопровождение', str(i+1), 0.0)
    else:
        rev_sec = get_override('  ├─ Вторичный рынок', str(i+1), deals_sec_vol * comm_secondary)
        rev_prim = get_override('  ├─ Первичный рынок', str(i+1), deals_prim_vol * comm_primary)
        rev_rent = get_override('  ├─ Аренда (жилая/коммерческая)', str(i+1), deals_rent_vol * comm_rent_val)
        rev_sub = get_override('  ├─ Загородная недвижимость', str(i+1), deals_sub_vol * comm_suburban)
        rev_overseas = get_override('  ├─ Зарубежная недвижимость', str(i+1), deals_overseas_vol * 220000.0)
        rev_other_p = get_override('  ├─ Сделки: прочее (МЛС, срочновыкуп)', str(i+1), deals_other_vol * 252000.0)
        rev_mort = get_override('  ├─ Сервисы: ипотека', str(i+1), serv_mort_vol * 70000.0)
        rev_ins = get_override('  ├─ Сервисы: страхование', str(i+1), serv_ins_vol * 18000.0)
        rev_legal = get_override('  └─ Сервисы: юр. сопровождение', str(i+1), serv_legal_vol * 150000.0)
        
    rev_secondary_list.append(rev_sec)
    rev_primary_list.append(rev_prim)
    rev_rent_list.append(rev_rent)
    rev_suburban_list.append(rev_sub)
    rev_overseas_list.append(rev_overseas)
    rev_other_p_list.append(rev_other_p)
    rev_mortgage_list.append(rev_mort)
    rev_insurance_list.append(rev_ins)
    rev_legal_list.append(rev_legal)
    
    total_revenue_month = rev_sec + rev_prim + rev_rent + rev_sub + rev_overseas + rev_other_p + rev_mort + rev_ins + rev_legal
    rev_total_list.append(total_revenue_month)
    
    # 3. РАСХОДЫ - ФОТ АГЕНТЫ
    agent_payouts_base = base['agent_comm'][i] * (agent_commission_pct / 38.0) * overall_voeronka_factor
    compact_agent_override = get_override('  ├─ Выплаты агентам (% комиссионных)', str(i+1), None)
    if compact_agent_override is not None:
        agent_payouts_base = compact_agent_override
        
    agent_D_val = get_override('  ├─ Агент: категория D', str(i+1), agent_payouts_base * 0.30)
    agent_C_val = get_override('  ├─ Агент: категория C', str(i+1), agent_payouts_base * 0.35)
    agent_B_val = get_override('  ├─ Агент: категория B', str(i+1), agent_payouts_base * 0.25)
    agent_A_val = get_override('  ├─ Агент: категория A', str(i+1), agent_payouts_base * 0.10)
    
    agent_payouts_total_month = agent_D_val + agent_C_val + agent_B_val + agent_A_val
    payouts_agents_list.append(agent_payouts_total_month)
    agent_D_list.append(agent_D_val)
    agent_C_list.append(agent_C_val)
    agent_B_list.append(agent_B_val)
    agent_A_list.append(agent_A_val)
    
    # 4. РАСХОДЫ - ФОТ БЭК-ОФИС (Кадры и Штатное расписание)
    salaries_back_base = base['salaries'][i] * (backoffice_salary_mult / 100.0) * salary_mult
    compact_backoffice_override = get_override('  ├─ Оклады бэк-офиса (оклады)', str(i+1), None)
    if compact_backoffice_override is not None:
        salaries_back_base = compact_backoffice_override
        
    # Пропорциональные базовые оклады
    def calc_role_salary(role_name, default_count, prop_factor):
        role_count = get_headcount_override('  ├─ ' + role_name if 'Агент' not in role_name else '  ├─ ' + role_name, str(i+1), default_count)
        default_salary = salaries_back_base * prop_factor
        if default_count > 0:
            return (role_count / default_count) * default_salary
        else:
            # Если по умолчанию сотрудников 0, но юзер нанял в Tab 4
            rate = BASE_SALARY_RATES.get(role_name, 0.0)
            return role_count * rate * (backoffice_salary_mult / 100.0) * salary_mult
            
    v_rop_C = get_override('  ├─ РОП: категория C', str(i+1), calc_role_salary('РОП: категория C', DEFAULT_HEADCOUNTS['РОП: категория C'][i], 0.15))
    v_rop_B = get_override('  ├─ РОП: категория B', str(i+1), calc_role_salary('РОП: категория B', DEFAULT_HEADCOUNTS['РОП: категория B'][i], 0.15))
    v_admin = get_override('  ├─ Администратор офиса', str(i+1), calc_role_salary('Администратор офиса', DEFAULT_HEADCOUNTS['Администратор офиса'][i], 0.10))
    v_hr = get_override('  ├─ HR/рекрутер', str(i+1), calc_role_salary('HR/рекрутер', DEFAULT_HEADCOUNTS['HR/рекрутер'][i], 0.10))
    v_ros = get_override('  ├─ РОС/тренер', str(i+1), calc_role_salary('РОС/тренер', DEFAULT_HEADCOUNTS['РОС/тренер'][i], 0.10))
    v_jurist = get_override('  ├─ Юрист', str(i+1), calc_role_salary('Юрист', DEFAULT_HEADCOUNTS['Юрист'][i], 0.10))
    v_mort_broker = get_override('  ├─ Ипотечный Брокер', str(i+1), calc_role_salary('Ипотечный Брокер', DEFAULT_HEADCOUNTS['Ипотечный Брокер'][i], 0.10))
    v_listing = get_override('  ├─ Листинг-менеджер', str(i+1), calc_role_salary('Листинг-менеджер', DEFAULT_HEADCOUNTS['Листинг-менеджер'][i], 0.08))
    v_photographer = get_override('  ├─ Фотограф', str(i+1), calc_role_salary('Фотограф', DEFAULT_HEADCOUNTS['Фотограф'][i], 0.04))
    v_smm = get_override('  ├─ Маркетолог/SMM', str(i+1), calc_role_salary('Маркетолог/SMM', DEFAULT_HEADCOUNTS['Маркетолог/SMM'][i], 0.08))
    
    salaries_back_total_month = v_rop_C + v_rop_B + v_admin + v_hr + v_ros + v_jurist + v_mort_broker + v_listing + v_photographer + v_smm
    salaries_backoffice_list.append(salaries_back_total_month)
    
    rop_C_list.append(v_rop_C)
    rop_B_list.append(v_rop_B)
    admin_list.append(v_admin)
    hr_list.append(v_hr)
    ros_list.append(v_ros)
    jurist_list.append(v_jurist)
    mort_broker_list.append(v_mort_broker)
    listing_list.append(v_listing)
    photographer_list.append(v_photographer)
    smm_list.append(v_smm)
    
    # Налоги на ФОТ оклады (с учетом ручной корректировки в P&L)
    taxes_payroll_base = base['salary_taxes'][i] * (backoffice_salary_mult / 100.0) * salary_mult
    taxes_payroll = get_override('  └─ Налоги на ФОТ оклады', str(i+1), taxes_payroll_base)
    taxes_payroll_list.append(taxes_payroll)
    
    total_payroll_list.append(agent_payouts_total_month + salaries_back_total_month + taxes_payroll)
    
    # 5. OPEX + КОРРЕКТИРОВКИ
    rent_val = get_override('  ├─ Аренда офиса', str(i+1), office_rent_custom * (base_rent_list[i] / 150000.0))
    rent_list.append(rent_val)
    
    internet_val = get_override('  ├─ Интернет', str(i+1), base_internet[i])
    internet_list.append(internet_val)
    
    mobile_val = get_override('  ├─ Сотовая связь', str(i+1), base_mobile[i])
    mobile_list.append(mobile_val)
    
    kanc_val = get_override('  ├─ Канцелярия', str(i+1), base_kanc[i])
    kanc_list.append(kanc_val)
    
    reklama_val = get_override('  ├─ Реклама объектов', str(i+1), base_reklama[i] * leads_mult * (marketing_mult / 100.0))
    reklama_list.append(reklama_val)
    
    hh_val = get_override('  ├─ HeadHunter.ru', str(i+1), base_hh[i])
    hh_list.append(hh_val)
    
    buh_val = get_override('  ├─ Бухгалтерия: аутсорс', str(i+1), base_buh[i])
    buh_list.append(buh_val)
    
    bank_val = get_override('  ├─ Услуги банка', str(i+1), base_bank[i])
    bank_list.append(bank_val)
    
    cleaning_val = get_override('  ├─ Уборка офиса', str(i+1), base_cleaning[i])
    cleaning_list.append(cleaning_val)
    
    gsm_val = get_override('  ├─ ГСМ', str(i+1), base_gsm[i])
    gsm_list.append(gsm_val)
    
    courier_val = get_override('  ├─ Доставка/курьер', str(i+1), base_courier[i])
    courier_list.append(courier_val)
    
    events_val = get_override('  └─ OPEX: Корпоративы', str(i+1), base_events[i])
    events_list.append(events_val)
    
    total_opex = rent_val + internet_val + mobile_val + kanc_val + reklama_val + hh_val + buh_val + bank_val + cleaning_val + gsm_val + courier_val + events_val
    total_opex_list.append(total_opex)
    
    # 6. ВЫПЛАТЫ ЦО (Франшиза)
    hq_fix = get_override('  ├─ Роялти FIX (вкл. НРФ)', str(i+1), base['hq_royalty_fix'][i] * royalty_fix_mult)
    hq_royalty_fix_list.append(hq_fix)
    
    hq_deal = get_override('  ├─ Роялти со сделок', str(i+1), (deals_sec_vol + deals_prim_vol + deals_sub_vol + deals_rent_vol) * royalty_deal_custom)
    hq_royalty_deal_list.append(hq_deal)
    
    hq_crm = get_override('  ├─ CRM-система', str(i+1), 5500 if i < 12 else 11000)
    hq_crm_list.append(hq_crm)
    
    hq_kc_val = get_override('  └─ Колл-центр и прочие сервисы', str(i+1), max(0.0, base['hq_crm_and_others'][i] - hq_crm))
    hq_kc_list.append(hq_kc_val)
    
    total_hq_payments = hq_fix + hq_deal + hq_crm + hq_kc_val
    total_hq_payments_list.append(total_hq_payments)
    
    # 7. НАЛОГИ
    tax_month = get_override('★ НАЛОГ УСН', str(i+1), total_revenue_month * (tax_rate / 100.0))
    taxes_usn_list.append(tax_month)
    
    # 8. CAPEX
    capex_fran = get_override('  ├─ Франшиза (Паушальный взнос + Роспатент)', str(i+1), capex_franchise_list[i])
    capex_renov = get_override('  ├─ Ремонт и Брендирование офиса', str(i+1), capex_renovation_list[i])
    capex_equip = get_override('  └─ Мебель, компьютеры и оборудование', str(i+1), capex_equipment_list[i])
    
    capex_franchise_list[i] = capex_fran
    capex_renovation_list[i] = capex_renov
    capex_equipment_list[i] = capex_equip
    
    tot_capex_month = capex_fran + capex_renov + capex_equip
    total_capex_list.append(tot_capex_month)
    
    total_expenses_month = agent_payouts_total_month + salaries_back_total_month + taxes_payroll + total_opex + total_hq_payments + tax_month
    net_profit_month = total_revenue_month - total_expenses_month - tot_capex_month
    adj_net_profit.append(net_profit_month)

# Расчет Месяца 0 (Старт)
capex_m0_franchise = get_override('  ├─ Франшиза (Паушальный взнос + Роспатент)', 'Старт', pau_fee_val + 17000)
capex_m0_renovation = get_override('  ├─ Ремонт и Брендирование офиса', 'Старт', capex_renovation_val)
capex_m0_equipment = get_override('  └─ Мебель, компьютеры и оборудование', 'Старт', capex_equipment_val)
total_capex_m0 = capex_m0_franchise + capex_m0_renovation + capex_m0_equipment

start_rent = get_override('  ├─ Аренда офиса', 'Старт', office_rent_custom)
start_opex_total = start_rent
for metric in ['  ├─ Интернет', '  ├─ Сотовая связь', '  ├─ Канцелярия', '  ├─ Реклама объектов', '  ├─ HeadHunter.ru', '  ├─ Бухгалтерия: аутсорс', '  ├─ Услуги банка', '  ├─ Уборка офиса', '  ├─ ГСМ', '  ├─ Доставка/курьер', '  └─ OPEX: Корпоративы']:
    start_opex_total += get_override(metric, 'Старт', 0.0)

cum_flow = []
running_sum = -total_capex_m0 - start_opex_total
cum_flow.append(running_sum)

for i in range(24):
    running_sum += adj_net_profit[i]
    cum_flow.append(running_sum)

# KPI
total_revenue_2years = sum(rev_total_list)
total_profit_2years = sum(adj_net_profit) - total_capex_m0 - start_opex_total
max_investment = abs(min(cum_flow))

payback_month = "Н/Д"
for idx, val in enumerate(cum_flow):
    if val >= 0:
        payback_month = f"{idx} мес."
        break

breakeven_month = "Н/Д"
for idx, val in enumerate(adj_net_profit):
    if val > 0:
        breakeven_month = f"{idx + 1} мес."
        break

# Карточки KPI
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("💼 Общая выручка (2 года)", f"{total_revenue_2years:,.0f} ₽".replace(",", " "))
with col2:
    st.metric("📈 Итоговая чистая прибыль", f"{total_profit_2years:,.0f} ₽".replace(",", " "))
with col3:
    st.metric("📉 Пик инвестиций (CAPEX+OPEX)", f"{max_investment:,.0f} ₽".replace(",", " "))
with col4:
    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("🎯 Без-ость", breakeven_month)
    with col_b:
        st.metric("⏳ Окупаемость", payback_month)

# Вспомогательная функция для агрегации метрик по Кварталам / Годам
def aggregate_metric(m_list, m0_val, scale):
    if scale == "Месячный (24 месяца)":
        return [m0_val] + list(m_list)
    elif scale == "Квартальный (8 кварталов)":
        q_vals = [m0_val]
        for q in range(8):
            sub_list = m_list[q*3:(q+1)*3]
            q_vals.append(sum(sub_list))
        return q_vals
    else:
        y_vals = [m0_val]
        y1 = sum(m_list[:12])
        y2 = sum(m_list[12:])
        y_vals.extend([y1, y2])
        return y_vals

# ----------------------------------------------------
# ТАБЫ: ГРАФИКИ, ДЕТАЛИЗАЦИЯ И СРАВНЕНИЕ
# ----------------------------------------------------
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📊 Интерактивные Графики", 
    "📋 Финансовый отчет P&L (₽)", 
    "📈 Операционная воронка (шт.)", 
    "👥 Кадровое планирование (чел.)", 
    "🌍 Сравнение Регионов", 
    "📝 Справочник и Методология"
])

with tab1:
    st.markdown("### Визуализация финансового развития")
    fig_cum = go.Figure()
    fig_cum.add_trace(go.Scatter(
        x=list(range(25)), y=cum_flow, mode='lines+markers', name='Накопленный кэш-флоу',
        line=dict(color='#BEAF87', width=3),
        marker=dict(size=6, color='#FFFFFF', line=dict(color='#BEAF87', width=1.5))
    ))
    fig_cum.add_shape(type="line", x0=0, y0=0, x1=24, y1=0, line=dict(color="#777779", width=1.5, dash="dash"))
    fig_cum.update_layout(
        title="Кривая накопленного кэш-флоу (J-Curve)", autosize=True,
        xaxis_title="Месяцы проекта (0 - старт)", yaxis_title="Баланс кэша, ₽",
        template="plotly_white", plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF", font=dict(color="#252526")
    )
    st.plotly_chart(fig_cum, use_container_width=True)
    
    fig_rev_exp = go.Figure()
    fig_rev_exp.add_trace(go.Bar(x=months, y=rev_total_list, name='Доходы (Комиссия ВКД)', marker_color='#BEAF87'))
    expenses_no_capex = [rev_total_list[idx] - adj_net_profit[idx] - total_capex_list[idx] for idx in range(24)]
    fig_rev_exp.add_trace(go.Bar(x=months, y=expenses_no_capex, name='Операционные расходы + Налоги', marker_color='#252526'))
    fig_rev_exp.update_layout(
        title="Ежемесячные доходы и операционные расходы", autosize=True,
        xaxis_title="Месяцы", yaxis_title="Сумма, ₽", barmode='group',
        template="plotly_white", plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF", font=dict(color="#252526")
    )
    st.plotly_chart(fig_rev_exp, use_container_width=True)

    # Круговая диаграмма структуры расходов
    avg_expenses = {
        'Выплаты агентам': sum(payouts_agents_list),
        'Оклады бэк-офиса': sum(salaries_backoffice_list),
        'Налоги на ФОТ оклады': sum(taxes_payroll_list),
        'Аренда офиса': sum(rent_list),
        'Реклама и маркетинг': sum(reklama_list),
        'HH.ru и подбор': sum(hh_list),
        'Роялти и ЦО': sum(total_hq_payments_list),
        'УСН налоги': sum(taxes_usn_list),
        'Прочие OPEX': sum(internet_list) + sum(mobile_list) + sum(kanc_list) + sum(buh_list) + sum(bank_list) + sum(cleaning_list) + sum(gsm_list) + sum(courier_list) + sum(events_list)
    }
    fig_pie = px.pie(
        names=list(avg_expenses.keys()),
        values=list(avg_expenses.values()),
        color_discrete_sequence=['#BEAF87', '#A19276', '#252526', '#777779', '#CCCCCC', '#E6E7E8', '#999999', '#BEAF87', '#EAEAEA'],
        title="Детализированная структура расходов за 2 года"
    )
    fig_pie.update_layout(template="plotly_white", paper_bgcolor="#FFFFFF", font=dict(color="#252526"))
    st.plotly_chart(fig_pie, use_container_width=True)

# ----------------------------------------------------
# ПОМОЩНИК ДЛЯ СЕКЦИОННОЙ ОТРИСОВКИ ТАБЛИЦ P&L
# ----------------------------------------------------
def render_section(rows_def, scale, editor_key):
    table_data = {'Показатель': [r[0] for r in rows_def]}
    
    if scale == "Месячный (24 месяца)":
        col_names = ["Старт"] + [str(i) for i in range(1, 25)]
    elif scale == "Квартальный (8 кварталов)":
        col_names = ["Старт", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8"]
    else:
        col_names = ["Старт", "Год 1", "Год 2"]
        
    for r in rows_def:
        ui_name, m_list, m0_val, db_override_key = r
        agg_values = aggregate_metric(m_list, m0_val, scale)
        for c_idx, col in enumerate(col_names):
            if col not in table_data:
                table_data[col] = []
            table_data[col].append(agg_values[c_idx])
            
    df = pd.DataFrame(table_data)
    col_config = {"Показатель": st.column_config.TextColumn(disabled=True)}
    for col in df.columns[1:]:
        col_config[col] = st.column_config.NumberColumn(format="%d ₽", min_value=0)
        
    is_editable = (scale == "Месячный (24 месяца)")
    
    edited_df = st.data_editor(
        df,
        use_container_width=True,
        disabled=["Показатель"] if is_editable else True,
        column_config=col_config,
        key=editor_key,
        height=min(len(df) * 35 + 45, 450)
    )
    
    changes_detected = False
    if is_editable:
        for idx_row in range(len(df)):
            ui_name, m_list, m0_val, db_override_key = rows_def[idx_row]
            if "★" in ui_name or "►" in ui_name or db_override_key is None:
                continue
                
            val_base_m0 = df.at[idx_row, "Старт"]
            val_edited_m0 = edited_df.at[idx_row, "Старт"]
            if float(val_base_m0) != float(val_edited_m0):
                st.session_state['pl_overrides'][(db_override_key, "Старт")] = parse_value(val_edited_m0)
                changes_detected = True
                
            for col_m in col_names[1:]:
                val_base = df.at[idx_row, col_m]
                val_edited = edited_df.at[idx_row, col_m]
                if float(val_base) != float(val_edited):
                    st.session_state['pl_overrides'][(db_override_key, col_m)] = parse_value(val_edited)
                    changes_detected = True
                        
    return changes_detected

# Пре-расчет значений для колонки Старт (Месяц 0)
v_m0_rev_sec = get_override('  ├─ Вторичный рынок', 'Старт', 0.0)
v_m0_rev_prim = get_override('  ├─ Первичный рынок', 'Старт', 0.0)
v_m0_rev_rent = get_override('  ├─ Аренда (жилая/коммерческая)', 'Старт', 0.0)
v_m0_rev_sub = get_override('  ├─ Загородная недвижимость', 'Старт', 0.0)
v_m0_rev_overseas = get_override('  ├─ Зарубежная недвижимость', 'Старт', 0.0)
v_m0_rev_other_p = get_override('  ├─ Сделки: прочее (МЛС, срочновыкуп)', 'Старт', 0.0)
v_m0_rev_mort = get_override('  ├─ Сервисы: ипотека', 'Старт', 0.0)
v_m0_rev_ins = get_override('  ├─ Сервисы: страхование', 'Старт', 0.0)
v_m0_rev_legal = get_override('  └─ Сервисы: юр. сопровождение', 'Старт', 0.0)
v_m0_rev_total = v_m0_rev_sec + v_m0_rev_prim + v_m0_rev_rent + v_m0_rev_sub + v_m0_rev_overseas + v_m0_rev_other_p + v_m0_rev_mort + v_m0_rev_ins + v_m0_rev_legal

v_m0_agent_D = get_override('  ├─ Агент: категория D', 'Старт', 0.0)
v_m0_agent_C = get_override('  ├─ Агент: категория C', 'Старт', 0.0)
v_m0_agent_B = get_override('  ├─ Агент: категория B', 'Старт', 0.0)
v_m0_agent_A = get_override('  ├─ Агент: категория A', 'Старт', 0.0)

v_m0_agent_compact = get_override('  ├─ Выплаты агентам (% комиссионных)', 'Старт', 0.0)
if v_m0_agent_compact > 0.0 and (v_m0_agent_D + v_m0_agent_C + v_m0_agent_B + v_m0_agent_A == 0.0):
    v_m0_agent_D = v_m0_agent_compact * 0.30
    v_m0_agent_C = v_m0_agent_compact * 0.35
    v_m0_agent_B = v_m0_agent_compact * 0.25
    v_m0_agent_A = v_m0_agent_compact * 0.10
v_m0_agent_total = v_m0_agent_D + v_m0_agent_C + v_m0_agent_B + v_m0_agent_A
if v_m0_agent_compact == 0.0 and v_m0_agent_total > 0.0:
    v_m0_agent_compact = v_m0_agent_total

v_m0_rop_C = get_override('  ├─ РОП: категория C', 'Старт', 0.0)
v_m0_rop_B = get_override('  ├─ РОП: категория B', 'Старт', 0.0)
v_m0_admin = get_override('  ├─ Администратор офиса', 'Старт', 0.0)
v_m0_hr = get_override('  ├─ HR/рекрутер', 'Старт', 0.0)
v_m0_ros = get_override('  ├─ РОС/тренер', 'Старт', 0.0)
v_m0_jurist = get_override('  ├─ Юрист', 'Старт', 0.0)
v_m0_mort_broker = get_override('  ├─ Ипотечный Брокер', 'Старт', 0.0)
v_m0_listing = get_override('  ├─ Листинг-менеджер', 'Старт', 0.0)
v_m0_photographer = get_override('  ├─ Фотограф', 'Старт', 0.0)
v_m0_smm = get_override('  ├─ Маркетолог/SMM', 'Старт', 0.0)

v_m0_back_compact = get_override('  ├─ Оклады бэк-офиса (оклады)', 'Старт', 0.0)
if v_m0_back_compact > 0.0 and (v_m0_rop_C + v_m0_rop_B + v_m0_admin + v_m0_hr + v_m0_ros + v_m0_jurist + v_m0_mort_broker + v_m0_listing + v_m0_photographer + v_m0_smm == 0.0):
    v_m0_rop_C = v_m0_back_compact * 0.15
    v_m0_rop_B = v_m0_back_compact * 0.15
    v_m0_admin = v_m0_back_compact * 0.10
    v_m0_hr = v_m0_back_compact * 0.10
    v_m0_ros = v_m0_back_compact * 0.10
    v_m0_jurist = v_m0_back_compact * 0.10
    v_m0_mort_broker = v_m0_back_compact * 0.10
    v_m0_listing = v_m0_back_compact * 0.08
    v_m0_photographer = v_m0_back_compact * 0.04
    v_m0_smm = v_m0_back_compact * 0.08
v_m0_back_total = v_m0_rop_C + v_m0_rop_B + v_m0_admin + v_m0_hr + v_m0_ros + v_m0_jurist + v_m0_mort_broker + v_m0_listing + v_m0_photographer + v_m0_smm
if v_m0_back_compact == 0.0 and v_m0_back_total > 0.0:
    v_m0_back_compact = v_m0_back_total

v_m0_salary_taxes = get_override('  └─ Налоги на ФОТ оклады', 'Старт', 0.0)
v_m0_fot_total = v_m0_agent_total + v_m0_back_total + v_m0_salary_taxes

v_m0_rent = start_rent
v_m0_internet = get_override('  ├─ Интернет', 'Старт', 0.0)
v_m0_mobile = get_override('  ├─ Сотовая связь', 'Старт', 0.0)
v_m0_kanc = get_override('  ├─ Канцелярия', 'Старт', 0.0)
v_m0_reklama = get_override('  ├─ Реклама объектов', 'Старт', 0.0)
v_m0_hh = get_override('  ├─ HeadHunter.ru', 'Старт', 0.0)
v_m0_buh = get_override('  ├─ Бухгалтерия: аутсорс', 'Старт', 0.0)
v_m0_bank = get_override('  ├─ Услуги банка', 'Старт', 0.0)
v_m0_cleaning = get_override('  ├─ Уборка офиса', 'Старт', 0.0)
v_m0_gsm = get_override('  ├─ ГСМ', 'Старт', 0.0)
v_m0_courier = get_override('  ├─ Доставка/курьер', 'Старт', 0.0)
v_m0_events = get_override('  └─ OPEX: Корпоративы', 'Старт', 0.0)
v_m0_opex_total = v_m0_rent + v_m0_internet + v_m0_mobile + v_m0_kanc + v_m0_reklama + v_m0_hh + v_m0_buh + v_m0_bank + v_m0_cleaning + v_m0_gsm + v_m0_courier + v_m0_events

v_m0_hq_fix = get_override('  ├─ Роялти FIX (вкл. НРФ)', 'Старт', 0.0)
v_m0_hq_deal = get_override('  ├─ Роялти со сделок', 'Старт', 0.0)
v_m0_hq_crm = get_override('  ├─ CRM-система', 'Старт', 0.0)
v_m0_hq_kc = get_override('  └─ Колл-центр и прочие сервисы', 'Старт', 0.0)
v_m0_hq_total = v_m0_hq_fix + v_m0_hq_deal + v_m0_hq_crm + v_m0_hq_kc

v_m0_tax = get_override('★ НАЛОГ УСН', 'Старт', 0.0)
v_m0_opex_tax_total = v_m0_fot_total + v_m0_opex_total + v_m0_hq_total + v_m0_tax

with tab2:
    st.markdown("### 📋 Финансовый отчет P&L (₽)")
    st.markdown("ℹ️ *Вы можете отредактировать любое расчетное значение в ячейках месяцев 1-24 или Старт, и вся модель мгновенно пересчитается каскадом.*")
    
    time_scale = st.radio(
        "📅 Масштаб времени отчета:",
        options=["Месячный (24 месяца)", "Квартальный (8 кварталов)", "Годовой (2 года)"],
        horizontal=True,
        help="Выберите удобный масштаб времени. Редактирование ячеек доступно только в месячном режиме."
    )
    
    show_detailed_fot = st.checkbox(
        "🔍 **Показать полную декомпозицию ФОТ (по должностям и категориям)**", 
        value=False
    )
    
    has_changes = False
    
    # Секция 1: Доходы
    with st.expander("🟢 **РАЗДЕЛ P&L 1: ДОХОДЫ (ВКД И ДОП. СЕРВИСЫ)**", expanded=True):
        revenue_rows_def = [
            ('Вторичный рынок', rev_secondary_list, v_m0_rev_sec, '  ├─ Вторичный рынок'),
            ('Первичный рынок', rev_primary_list, v_m0_rev_prim, '  ├─ Первичный рынок'),
            ('Аренда (жилая/коммерческая)', rev_rent_list, v_m0_rev_rent, '  ├─ Аренда (жилая/коммерческая)'),
            ('Загородная недвижимость', rev_suburban_list, v_m0_rev_sub, '  ├─ Загородная недвижимость'),
            ('Зарубежная недвижимость', rev_overseas_list, v_m0_rev_overseas, '  ├─ Зарубежная недвижимость'),
            ('Сделки: прочее (МЛС, срочновыкуп)', rev_other_p_list, v_m0_rev_other_p, '  ├─ Сделки: прочее (МЛС, срочновыкуп)'),
            ('Сервисы: ипотека', rev_mortgage_list, v_m0_rev_mort, '  ├─ Сервисы: ипотека'),
            ('Сервисы: страхование', rev_insurance_list, v_m0_rev_ins, '  ├─ Сервисы: страхование'),
            ('Сервисы: юр. сопровождение', rev_legal_list, v_m0_rev_legal, '  └─ Сервисы: юр. сопровождение'),
            ('★ ИТОГО ДОХОДЫ', rev_total_list, v_m0_rev_total, None)
        ]
        if render_section(revenue_rows_def, time_scale, "editor_rev"): has_changes = True
        
        # --- NEW: VISUAL INTERACTIVE INCOME INSPECTOR ---
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div style='border-top: 1px solid #E6E7E8; padding-top: 15px;'></div>", unsafe_allow_html=True)
        st.markdown("#### 🎯 Интуитивный инспектор структуры доходов (ВКД)")
        st.markdown("ℹ️ *Выберите интересующий вас период, чтобы увидеть структуру и источники валового комиссионного дохода вашего офиса.*")
        
        selected_rev_ins_col = st.select_slider(
            "📍 Выберите месяц для инспектирования структуры доходов:",
            options=["Старт"] + [str(m) for m in range(1, 25)],
            value="13",
            key="rev_ins_slider"
        )
        
        rev_info = {
            'Вторичный рынок': {'emoji': '🏢', 'list': rev_secondary_list, 'm0': v_m0_rev_sec, 'desc': 'Комиссии со сделок купли-продажи на вторичном рынке жилья.'},
            'Первичный рынок': {'emoji': '🏗️', 'list': rev_primary_list, 'm0': v_m0_rev_prim, 'desc': 'Комиссии, выплачиваемые застройщиками за продажу новостроек.'},
            'Аренда (жилая/коммерческая)': {'emoji': '🔑', 'list': rev_rent_list, 'm0': v_m0_rev_rent, 'desc': 'Доходы со сделок аренды жилой и коммерческой недвижимости.'},
            'Загородная недвижимость': {'emoji': '🏡', 'list': rev_suburban_list, 'm0': v_m0_rev_sub, 'desc': 'Комиссии со сделок по домам, коттеджам и земельным участкам.'},
            'Зарубежная недвижимость': {'emoji': '✈️', 'list': rev_overseas_list, 'm0': v_m0_rev_overseas, 'desc': 'Доходы от партнерских сделок по продаже недвижимости за рубежом.'},
            'Сделки: прочее (МЛС, срочновыкуп)': {'emoji': '🔄', 'list': rev_other_p_list, 'm0': v_m0_rev_other_p, 'desc': 'Доходы от мультилистинга, срочного выкупа и прочих операций.'},
            'Сервисы: ипотека': {'emoji': '🏦', 'list': rev_mortgage_list, 'm0': v_m0_rev_mort, 'desc': 'Дополнительные вознаграждения от банков за ипотечные сделки.'},
            'Сервисы: страхование': {'emoji': '🛡️', 'list': rev_insurance_list, 'm0': v_m0_rev_ins, 'desc': 'Партнерские комиссии от страховых компаний при страховании заемщиков.'},
            'Сервисы: юр. сопровождение': {'emoji': '⚖️', 'list': rev_legal_list, 'm0': v_m0_rev_legal, 'desc': 'Доходы от платного юридического сопровождения сторонних сделок.'}
        }
        
        rev_vals = {}
        for r_name, r_meta in rev_info.items():
            if selected_rev_ins_col == 'Старт':
                val = r_meta['m0']
            else:
                m_idx = int(selected_rev_ins_col) - 1
                val = r_meta['list'][m_idx]
            if val > 0:
                rev_vals[r_name] = val
                
        total_rev_val = sum(rev_vals.values())
        
        if total_rev_val > 0:
            col_rev_chart, col_rev_cards = st.columns([1, 1.2])
            with col_rev_chart:
                fig_rev_pie = go.Figure(data=[go.Pie(
                    labels=list(rev_vals.keys()),
                    values=list(rev_vals.values()),
                    hole=.4,
                    marker=dict(colors=['#BEAF87', '#A19276', '#252526', '#777779', '#CCCCCC', '#E6E7E8', '#999999', '#D1C49D', '#8C8267']),
                    textinfo='percent+label',
                    showlegend=False
                )])
                fig_rev_pie.update_layout(
                    title=f"Валовый доход ({selected_rev_ins_col} период): {total_rev_val:,.0f} ₽".replace(",", " "),
                    template="plotly_white",
                    autosize=True,
                    margin=dict(t=40, b=10, l=10, r=10),
                    height=280
                )
                st.plotly_chart(fig_rev_pie, use_container_width=True)
                
            with col_rev_cards:
                st.markdown(f"**🟢 Источники поступлений ({selected_rev_ins_col} месяц):**")
                for r_name, r_val in rev_vals.items():
                    meta = rev_info[r_name]
                    st.markdown(f"""
                    <div style='background-color: #F8F9FA; padding: 10px 15px; border-radius: 6px; border-left: 4px solid #BEAF87; margin-bottom: 8px;'>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <span style='font-weight: 700; color: #252526; font-size: 13px;'>{meta['emoji']} {r_name}</span>
                            <span style='font-weight: 700; color: #A19276; font-size: 13px;'>{r_val:,.0f} ₽</span>
                        </div>
                        <div style='margin-top: 4px; font-size: 11px; color: #777779;'>
                            <span>{meta['desc']}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.info(f"ℹ️ В периоде '{selected_rev_ins_col}' выручка отсутствует. Первые 3 месяца запуска уходят на накопление базы объектов.")

        
    # Секция 2: ФОТ
    with st.expander("👥 **РАЗДЕЛ P&L 2: РАСХОДЫ НА ПЕРСОНАЛ (ФОТ И НАЛОГИ)**", expanded=False):
        if show_detailed_fot:
            fot_rows_def = [
                ('Агент: категория D', agent_D_list, v_m0_agent_D, '  ├─ Агент: категория D'),
                ('Агент: категория C', agent_C_list, v_m0_agent_C, '  ├─ Агент: категория C'),
                ('Агент: категория B', agent_B_list, v_m0_agent_B, '  ├─ Агент: категория B'),
                ('Агент: категория A', agent_A_list, v_m0_agent_A, '  ├─ Агент: категория A'),
                ('РОП: категория C', rop_C_list, v_m0_rop_C, '  ├─ РОП: категория C'),
                ('РОП: категория B', rop_B_list, v_m0_rop_B, '  ├─ РОП: категория B'),
                ('Администратор офиса', admin_list, v_m0_admin, '  ├─ Администратор офиса'),
                ('HR/рекрутер', hr_list, v_m0_hr, '  ├─ HR/рекрутер'),
                ('РОС/тренер', ros_list, v_m0_ros, '  ├─ РОС/тренер'),
                ('Юрист', jurist_list, v_m0_jurist, '  ├─ Юрист'),
                ('Ипотечный Брокер', mort_broker_list, v_m0_mort_broker, '  ├─ Ипотечный Брокер'),
                ('Листинг-менеджер', listing_list, v_m0_listing, '  ├─ Листинг-менеджер'),
                ('Фотограф', photographer_list, v_m0_photographer, '  ├─ Фотограф'),
                ('Маркетолог/SMM', smm_list, v_m0_smm, '  ├─ Маркетолог/SMM'),
                ('Налоги на ФОТ оклады', taxes_payroll_list, v_m0_salary_taxes, '  └─ Налоги на ФОТ оклады'),
                ('★ ИТОГО ФОТ', total_payroll_list, v_m0_fot_total, None)
            ]
        else:
            fot_rows_def = [
                ('Выплаты агентам (% комиссионных)', payouts_agents_list, v_m0_agent_total, '  ├─ Выплаты агентам (% комиссионных)'),
                ('Оклады бэк-офиса (оклады)', salaries_backoffice_list, v_m0_back_total, '  ├─ Оклады бэк-офиса (оклады)'),
                ('Налоги на ФОТ оклады', taxes_payroll_list, v_m0_salary_taxes, '  └─ Налоги на ФОТ оклады'),
                ('★ ИТОГО ФОТ', total_payroll_list, v_m0_fot_total, None)
            ]
        if render_section(fot_rows_def, time_scale, "editor_fot"): has_changes = True
        
        # --- NEW: VISUAL INTERACTIVE FOT INSPECTOR ---
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div style='border-top: 1px solid #E6E7E8; padding-top: 15px;'></div>", unsafe_allow_html=True)
        st.markdown("#### 🎯 Интуитивный инспектор структуры ФОТ бэк-офиса")
        st.markdown("ℹ️ *Выберите интересующий вас месяц, чтобы мгновенно увидеть визуальный состав команды бэк-офиса, оклады сотрудников и их долю в бюджете.*")
        
        selected_ins_col = st.select_slider(
            "📍 Выберите месяц для инспектирования структуры окладов бэк-офиса:",
            options=["Старт"] + [str(m) for m in range(1, 25)],
            value="13",
            key="fot_ins_slider"
        )
        
        roles_info = {
            'РОП: категория C': {'emoji': '👔', 'list': rop_C_list, 'm0': v_m0_rop_C, 'desc': 'Руководитель отдела продаж (Категория С) — управление группой до 10 агентов.'},
            'РОП: категория B': {'emoji': '👑', 'list': rop_B_list, 'm0': v_m0_rop_B, 'desc': 'Руководитель отдела продаж (Категория B) — опытный РОП для крупных дивизионов.'},
            'Администратор офиса': {'emoji': '📝', 'list': admin_list, 'm0': v_m0_admin, 'desc': 'Администратор — жизнеобеспечение офиса, координация звонков и листинга.'},
            'HR/рекрутер': {'emoji': '👥', 'list': hr_list, 'm0': v_m0_hr, 'desc': 'HR-менеджер — непрерывный наем и формирование новых групп стажеров.'},
            'РОС/тренер': {'emoji': '🎓', 'list': ros_list, 'm0': v_m0_ros, 'desc': 'Руководитель обучения (РОС) — подготовка агентов по стандартам Century 21.'},
            'Юрист': {'emoji': '⚖️', 'list': jurist_list, 'm0': v_m0_jurist, 'desc': 'Юрист — экспертиза документов, сопровождение сделок любой сложности.'},
            'Ипотечный Брокер': {'emoji': '💰', 'list': mort_broker_list, 'm0': v_m0_mort_broker, 'desc': 'Ипотечный брокер — согласование преференций по ставкам в банках-партнерах.'},
            'Листинг-менеджер': {'emoji': '📋', 'list': listing_list, 'm0': v_m0_listing, 'desc': 'Листинг-менеджер — координация выгрузки базы объектов на классифайды.'},
            'Фотограф': {'emoji': '📷', 'list': photographer_list, 'm0': v_m0_photographer, 'desc': 'Штатный фотограф — профессиональная фото/видео съемка объектов.'},
            'Маркетолог/SMM': {'emoji': '📢', 'list': smm_list, 'm0': v_m0_smm, 'desc': 'Маркетолог — лидогенерация, ведение соцсетей и локальный пиар офиса.'}
        }
        
        role_salaries = {}
        for r_name, r_meta in roles_info.items():
            if selected_ins_col == 'Старт':
                val = r_meta['m0']
            else:
                m_idx = int(selected_ins_col) - 1
                val = r_meta['list'][m_idx]
            if val > 0:
                role_salaries[r_name] = val
                
        total_bo_salary = sum(role_salaries.values())
        
        if total_bo_salary > 0:
            col_chart, col_cards = st.columns([1, 1.2])
            with col_chart:
                fig_bo_pie = go.Figure(data=[go.Pie(
                    labels=list(role_salaries.keys()),
                    values=list(role_salaries.values()),
                    hole=.4,
                    marker=dict(colors=['#BEAF87', '#A19276', '#252526', '#777779', '#CCCCCC', '#E6E7E8', '#999999', '#D1C49D', '#8C8267', '#3C3C3D']),
                    textinfo='percent+label',
                    showlegend=False
                )])
                fig_bo_pie.update_layout(
                    title=f"Оклады бэк-офиса ({selected_ins_col} период): {total_bo_salary:,.0f} ₽".replace(",", " "),
                    template="plotly_white",
                    margin=dict(t=40, b=10, l=10, r=10),
                    height=280
                )
                st.plotly_chart(fig_bo_pie, use_container_width=True)
                
            with col_cards:
                st.markdown(f"**🕵️ Активный состав бэк-офиса ({selected_ins_col} месяц):**")
                # Show beautiful badges
                for r_name, r_val in role_salaries.items():
                    meta = roles_info[r_name]
                    if selected_ins_col == 'Старт':
                        hc = 0
                    else:
                        hc = get_headcount_override('  ├─ ' + r_name if r_name != 'Маркетолог/SMM' else '  ├─ Маркетолог/SMM', selected_ins_col, DEFAULT_HEADCOUNTS[r_name][int(selected_ins_col)-1])
                        
                    st.markdown(f"""
                    <div style='background-color: #F8F9FA; padding: 10px 15px; border-radius: 6px; border-left: 4px solid #BEAF87; margin-bottom: 8px;'>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <span style='font-weight: 700; color: #252526; font-size: 13px;'>{meta['emoji']} {r_name}</span>
                            <span style='background-color: #E6E7E8; color: #252526; padding: 1px 6px; border-radius: 10px; font-size: 10px; font-weight: 700;'>{hc} чел.</span>
                        </div>
                        <div style='display: flex; justify-content: space-between; align-items: center; margin-top: 4px; font-size: 11px;'>
                            <span style='color: #777779;'>{meta['desc']}</span>
                            <span style='font-weight: 700; color: #A19276;'>{r_val:,.0f} ₽</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.info(f"ℹ️ В периоде '{selected_ins_col}' в бэк-офисе нет активных окладов. Все сотрудники будут наняты на последующих этапах проекта.")

        
    # Секция 3: OPEX
    with st.expander("🏢 **РАЗДЕЛ P&L 3: ОПЕРАЦИОННЫЕ РАСХОДЫ (OPEX ОФИСА)**", expanded=False):
        opex_rows_def = [
            ('Аренда офиса', rent_list, v_m0_rent, '  ├─ Аренда офиса'),
            ('Интернет', internet_list, v_m0_internet, '  ├─ Интернет'),
            ('Сотовая связь', mobile_list, v_m0_mobile, '  ├─ Сотовая связь'),
            ('Канцелярия', kanc_list, v_m0_kanc, '  ├─ Канцелярия'),
            ('Реклама объектов', reklama_list, v_m0_reklama, '  ├─ Реклама объектов'),
            ('HeadHunter.ru', hh_list, v_m0_hh, '  ├─ HeadHunter.ru'),
            ('Бухгалтерия: аутсорс', buh_list, v_m0_buh, '  ├─ Бухгалтерия: аутсорс'),
            ('Услуги банка', bank_list, v_m0_bank, '  ├─ Услуги банка'),
            ('Уборка офиса', cleaning_list, v_m0_cleaning, '  ├─ Уборка офиса'),
            ('ГСМ', gsm_list, v_m0_gsm, '  ├─ ГСМ'),
            ('Доставка/курьер', courier_list, v_m0_courier, '  ├─ Доставка/курьер'),
            ('Корпоративные мероприятия', events_list, v_m0_events, '  └─ OPEX: Корпоративы'),
            ('★ ИТОГО OPEX', total_opex_list, v_m0_opex_total, None)
        ]
        if render_section(opex_rows_def, time_scale, "editor_opex"): has_changes = True
        
        # --- NEW: VISUAL INTERACTIVE OPEX INSPECTOR ---
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div style='border-top: 1px solid #E6E7E8; padding-top: 15px;'></div>", unsafe_allow_html=True)
        st.markdown("#### 🎯 Интуитивный инспектор структуры OPEX (Операционных расходов)")
        st.markdown("ℹ️ *Выберите интересующий вас месяц, чтобы мгновенно увидеть визуальный состав операционных затрат офиса, доли расходов и бизнес-рекомендации.*")
        
        selected_opex_ins_col = st.select_slider(
            "📍 Выберите месяц для инспектирования структуры OPEX:",
            options=["Старт"] + [str(m) for m in range(1, 25)],
            value="13",
            key="opex_ins_slider"
        )
        
        opex_info = {
            'Аренда офиса': {'emoji': '🏢', 'list': rent_list, 'm0': v_m0_rent, 'desc': 'Аренда офиса — размещение агентов, представительские переговоры и встречи с клиентами.'},
            'Интернет': {'emoji': '🌐', 'list': internet_list, 'm0': v_m0_internet, 'desc': 'Интернет — высокоскоростной выделенный доступ для работы CRM и телефонии.'},
            'Сотовая связь': {'emoji': '📞', 'list': mobile_list, 'm0': v_m0_mobile, 'desc': 'Сотовая связь — корпоративная связь агентов, IP-телефония и связь с клиентами.'},
            'Канцелярия': {'emoji': '✏️', 'list': kanc_list, 'm0': v_m0_kanc, 'desc': 'Канцелярия — закупка бумаги, папок, ручек и расходников для оргтехники.'},
            'Реклама объектов': {'emoji': '📢', 'list': reklama_list, 'm0': v_m0_reklama, 'desc': 'Реклама объектов — размещение и продвижение объявлений на Циан, Авито, Яндекс.'},
            'HeadHunter.ru': {'emoji': '💼', 'list': hh_list, 'm0': v_m0_hh, 'desc': 'HeadHunter.ru — покупка доступов и вакансий для непрерывного найма новых агентов.'},
            'Бухгалтерия: аутсорс': {'emoji': '🧮', 'list': buh_list, 'm0': v_m0_buh, 'desc': 'Бухгалтерия — аутсорсинг налогового, кадрового и бухгалтерского учета.'},
            'Услуги банка': {'emoji': '🏦', 'list': bank_list, 'm0': v_m0_bank, 'desc': 'Услуги банка — эквайринг, комиссия за расчетный счет, банковские переводы.'},
            'Уборка офиса': {'emoji': '🧹', 'list': cleaning_list, 'm0': v_m0_cleaning, 'desc': 'Уборка офиса — ежедневный клининг для поддержания премиального вида офиса.'},
            'ГСМ': {'emoji': '🚗', 'list': gsm_list, 'm0': v_m0_gsm, 'desc': 'ГСМ — компенсация транспортных затрат агентов при выездах на показы.'},
            'Доставка/курьер': {'emoji': '📦', 'list': courier_list, 'm0': v_m0_courier, 'desc': 'Доставка/курьер — срочная доставка оригиналов договоров и корреспонденции.'},
            'Корпоративные мероприятия': {'emoji': '🎉', 'list': events_list, 'm0': v_m0_events, 'desc': 'Корпоративы — представительские расходы, мероприятия и награждения лучших агентов.'}
        }
        
        opex_salaries = {}
        for o_name, o_meta in opex_info.items():
            if selected_opex_ins_col == 'Старт':
                val = o_meta['m0']
            else:
                m_idx = int(selected_opex_ins_col) - 1
                val = o_meta['list'][m_idx]
            if val > 0:
                opex_salaries[o_name] = val
                
        total_opex_val = sum(opex_salaries.values())
        
        if total_opex_val > 0:
            col_opex_chart, col_opex_cards = st.columns([1, 1.2])
            with col_opex_chart:
                fig_opex_pie = go.Figure(data=[go.Pie(
                    labels=list(opex_salaries.keys()),
                    values=list(opex_salaries.values()),
                    hole=.4,
                    marker=dict(colors=['#BEAF87', '#A19276', '#252526', '#777779', '#CCCCCC', '#E6E7E8', '#999999', '#D1C49D', '#8C8267', '#3C3C3D', '#EFEFEF', '#E4D3C5']),
                    textinfo='percent+label',
                    showlegend=False
                )])
                fig_opex_pie.update_layout(
                    title=f"Операционные расходы ({selected_opex_ins_col} период): {total_opex_val:,.0f} ₽".replace(",", " "),
                    template="plotly_white",
                    margin=dict(t=40, b=10, l=10, r=10),
                    height=280
                )
                st.plotly_chart(fig_opex_pie, use_container_width=True)
                
            with col_opex_cards:
                st.markdown(f"**🏢 Расшифровка затрат OPEX ({selected_opex_ins_col} месяц):**")
                # Show beautiful badges
                for o_name, o_val in opex_salaries.items():
                    meta = opex_info[o_name]
                    st.markdown(f"""
                    <div style='background-color: #F8F9FA; padding: 10px 15px; border-radius: 6px; border-left: 4px solid #BEAF87; margin-bottom: 8px;'>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <span style='font-weight: 700; color: #252526; font-size: 13px;'>{meta['emoji']} {o_name}</span>
                            <span style='font-weight: 700; color: #A19276; font-size: 13px;'>{o_val:,.0f} ₽</span>
                        </div>
                        <div style='margin-top: 4px; font-size: 11px; color: #777779;'>
                            <span>{meta['desc']}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.info(f"ℹ️ В периоде '{selected_opex_ins_col}' в OPEX нет активных расходов. Все статьи равны нулю.")
        
    # Секция 4: Выплаты ЦО
    with st.expander("👑 **РАЗДЕЛ P&L 4: ПЛАТЕЖИ В ЦЕНТРАЛЬНЫЙ ОФИС (ФРАНШИЗА)**", expanded=False):
        hq_rows_def = [
            ('Роялти FIX (вкл. НРФ)', hq_royalty_fix_list, v_m0_hq_fix, '  ├─ Роялти FIX (вкл. НРФ)'),
            ('Роялти со сделок', hq_royalty_deal_list, v_m0_hq_deal, '  ├─ Роялти со сделок'),
            ('CRM-система', hq_crm_list, v_m0_hq_crm, '  ├─ CRM-система'),
            ('Колл-центр и прочие сервисы', hq_kc_list, v_m0_hq_kc, '  └─ Колл-центр и прочие сервисы'),
            ('★ ИТОГО ВЫПЛАТЫ ЦО', total_hq_payments_list, v_m0_hq_total, None)
        ]
        if render_section(hq_rows_def, time_scale, "editor_hq"): has_changes = True
        
        # --- NEW: VISUAL INTERACTIVE HQ INSPECTOR ---
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div style='border-top: 1px solid #E6E7E8; padding-top: 15px;'></div>", unsafe_allow_html=True)
        st.markdown("#### 🎯 Интуитивный инспектор выплат в Центральный Офис (ЦО)")
        st.markdown("ℹ️ *Выберите интересующий вас месяц, чтобы увидеть структуру роялти и расходов на ИТ-сервисы франшизы.*")
        
        selected_hq_ins_col = st.select_slider(
            "📍 Выберите месяц для инспектирования выплат в ЦО:",
            options=["Старт"] + [str(m) for m in range(1, 25)],
            value="13",
            key="hq_ins_slider"
        )
        
        hq_info = {
            'Роялти FIX (вкл. НРФ)': {'emoji': '👑', 'list': hq_royalty_fix_list, 'm0': v_m0_hq_fix, 'desc': 'Фиксированные роялти — каникулы первые 3 месяца, далее шкала под регион.'},
            'Роялти со сделок': {'emoji': '🤝', 'list': hq_royalty_deal_list, 'm0': v_m0_hq_deal, 'desc': 'Переменные роялти с каждой закрытой сделки вашего офиса.'},
            'CRM-система': {'emoji': '💻', 'list': hq_crm_list, 'm0': v_m0_hq_crm, 'desc': 'Плата за пользование CRM, лицензии и ИТ-инфраструктуру Century 21.'},
            'Колл-центр и прочие сервисы': {'emoji': '📞', 'list': hq_kc_list, 'm0': v_m0_hq_kc, 'desc': 'Услуги федерального колл-центра, поддержка листинга и обучение.'}
        }
        
        hq_vals = {}
        for h_name, h_meta in hq_info.items():
            if selected_hq_ins_col == 'Старт':
                val = h_meta['m0']
            else:
                m_idx = int(selected_hq_ins_col) - 1
                val = h_meta['list'][m_idx]
            if val > 0:
                hq_vals[h_name] = val
                
        total_hq_val = sum(hq_vals.values())
        
        if total_hq_val > 0:
            col_hq_chart, col_hq_cards = st.columns([1, 1.2])
            with col_hq_chart:
                fig_hq_pie = go.Figure(data=[go.Pie(
                    labels=list(hq_vals.keys()),
                    values=list(hq_vals.values()),
                    hole=.4,
                    marker=dict(colors=['#BEAF87', '#A19276', '#252526', '#777779']),
                    textinfo='percent+label',
                    showlegend=False
                )])
                fig_hq_pie.update_layout(
                    title=f"Выплаты в ЦО ({selected_hq_ins_col} период): {total_hq_val:,.0f} ₽".replace(",", " "),
                    template="plotly_white",
                    autosize=True,
                    margin=dict(t=40, b=10, l=10, r=10),
                    height=280
                )
                st.plotly_chart(fig_hq_pie, use_container_width=True)
                
            with col_hq_cards:
                st.markdown(f"**👑 Структура роялти ({selected_hq_ins_col} месяц):**")
                for h_name, h_val in hq_vals.items():
                    meta = hq_info[h_name]
                    st.markdown(f"""
                    <div style='background-color: #F8F9FA; padding: 10px 15px; border-radius: 6px; border-left: 4px solid #BEAF87; margin-bottom: 8px;'>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <span style='font-weight: 700; color: #252526; font-size: 13px;'>{meta['emoji']} {h_name}</span>
                            <span style='font-weight: 700; color: #A19276; font-size: 13px;'>{h_val:,.0f} ₽</span>
                        </div>
                        <div style='margin-top: 4px; font-size: 11px; color: #777779;'>
                            <span>{meta['desc']}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.info(f"ℹ️ В периоде '{selected_hq_ins_col}' выплаты в ЦО отсутствуют (действует льготный беспроцентный период запуска франшизы).")

        
    # Секция 5: CAPEX
    with st.expander("🛠️ **РАЗДЕЛ P&L 5: КАПИТАЛЬНЫЕ ЗАТРАТЫ И ИНВЕСТИЦИИ (CAPEX)**", expanded=False):
        capex_rows_def = [
            ('Франшиза (Паушальный взнос + Роспатент)', capex_franchise_list, capex_m0_franchise, '  ├─ Франшиза (Паушальный взнос + Роспатент)'),
            ('Ремонт и Брендирование офиса', capex_renovation_list, capex_m0_renovation, '  ├─ Ремонт и Брендирование офиса'),
            ('Мебель, компьютеры и оборудование', capex_equipment_list, capex_m0_equipment, '  └─ Мебель, компьютеры и оборудование'),
            ('★ ИТОГО CAPEX', total_capex_list, total_capex_m0, None)
        ]
        if render_section(capex_rows_def, time_scale, "editor_capex"): has_changes = True
        
        # --- NEW: VISUAL INTERACTIVE CAPEX INSPECTOR ---
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<div style='border-top: 1px solid #E6E7E8; padding-top: 15px;'></div>", unsafe_allow_html=True)
        st.markdown("#### 🎯 Интуитивный инспектор структуры инвестиций и CAPEX")
        st.markdown("ℹ️ *Выберите интересующий вас месяц, чтобы увидеть структуру капитальных затрат.*")
        
        selected_cap_ins_col = st.select_slider(
            "📍 Выберите месяц для инспектирования CAPEX:",
            options=["Старт"] + [str(m) for m in range(1, 25)],
            value="Старт",
            key="capex_ins_slider"
        )
        
        cap_info = {
            'Франшиза (Паушальный взнос + Роспатент)': {'emoji': '📜', 'list': capex_franchise_list, 'm0': capex_m0_franchise, 'desc': 'Паушальный взнос за покупку франшизы и регистрацию в Роспатенте.'},
            'Ремонт и Брендирование офиса': {'emoji': '🔨', 'list': capex_renovation_list, 'm0': capex_m0_renovation, 'desc': 'Дизайн-проект, материалы, строительные работы по брендбуку Century 21.'},
            'Мебель, компьютеры и оборудование': {'emoji': '🖥️', 'list': capex_equipment_list, 'm0': capex_m0_equipment, 'desc': 'Офисные столы, стулья, компьютеры агентам, оргтехника и телефония.'}
        }
        
        cap_vals = {}
        for c_name, c_meta in cap_info.items():
            if selected_cap_ins_col == 'Старт':
                val = c_meta['m0']
            else:
                m_idx = int(selected_cap_ins_col) - 1
                val = c_meta['list'][m_idx]
            if val > 0:
                cap_vals[c_name] = val
                
        total_cap_val = sum(cap_vals.values())
        
        if total_cap_val > 0:
            col_cap_chart, col_cap_cards = st.columns([1, 1.2])
            with col_cap_chart:
                fig_cap_pie = go.Figure(data=[go.Pie(
                    labels=list(cap_vals.keys()),
                    values=list(cap_vals.values()),
                    hole=.4,
                    marker=dict(colors=['#BEAF87', '#A19276', '#252526']),
                    textinfo='percent+label',
                    showlegend=False
                )])
                fig_cap_pie.update_layout(
                    title=f"Капитальные вложения ({selected_cap_ins_col} период): {total_cap_val:,.0f} ₽".replace(",", " "),
                    template="plotly_white",
                    autosize=True,
                    margin=dict(t=40, b=10, l=10, r=10),
                    height=280
                )
                st.plotly_chart(fig_cap_pie, use_container_width=True)
                
            with col_cap_cards:
                st.markdown(f"**🛠️ Капитальные затраты ({selected_cap_ins_col} месяц):**")
                for c_name, c_val in cap_vals.items():
                    meta = cap_info[c_name]
                    st.markdown(f"""
                    <div style='background-color: #F8F9FA; padding: 10px 15px; border-radius: 6px; border-left: 4px solid #BEAF87; margin-bottom: 8px;'>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <span style='font-weight: 700; color: #252526; font-size: 13px;'>{meta['emoji']} {c_name}</span>
                            <span style='font-weight: 700; color: #A19276; font-size: 13px;'>{c_val:,.0f} ₽</span>
                        </div>
                        <div style='margin-top: 4px; font-size: 11px; color: #777779;'>
                            <span>{meta['desc']}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.info(f"ℹ️ В периоде '{selected_cap_ins_col}' нет дополнительных капитальных инвестиций (все основные инвестиции осуществлены на Старте).")

        
    if has_changes:
        st.rerun()
        
    # Сводная таблица основных показателей (Красивое форматирование итогов)
    st.markdown("#### 🏆 Сводный финансовый баланс проекта")
    
    summary_labels = [
        '★ ИТОГО ДОХОДЫ',
        '★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ',
        '★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ',
        '★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК'
    ]
    
    sum_data = {'Показатель': summary_labels}
    col_names = ["Старт"] + [str(i) for i in range(1, 25)] if time_scale == "Месячный (24 месяца)" else (["Старт", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8"] if time_scale == "Квартальный (8 кварталов)" else ["Старт", "Год 1", "Год 2"])
    
    m0_values = [v_m0_rev_total, v_m0_opex_tax_total, -total_capex_m0 - v_m0_opex_tax_total, cum_flow[0]]
    
    for l_idx, label in enumerate(summary_labels):
        m_list = rev_total_list if l_idx == 0 else ( [total_payroll_list[i] + total_opex_list[i] + total_hq_payments_list[i] + taxes_usn_list[i] for i in range(24)] if l_idx == 1 else (adj_net_profit if l_idx == 2 else cum_flow[1:]) )
        m0_v = m0_values[l_idx]
        agg_v = aggregate_metric(m_list, m0_v, time_scale)
        for c_idx, col in enumerate(col_names):
            if col not in sum_data:
                sum_data[col] = []
            sum_data[col].append(agg_v[c_idx])
            
    df_sum = pd.DataFrame(sum_data)
    
    # Стильное форматирование с помощью Pandas Styler
    def style_dataframe(df):
        styled = df.style.format(formatter={col: "{:,.0f} ₽" for col in df.columns[1:]})
        
        profit_row_idx = df[df['Показатель'] == '★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ'].index[0]
        flow_row_idx = df[df['Показатель'] == '★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК'].index[0]
        
        def color_profit(val):
            try:
                v = float(str(val).replace(' ','').replace('₽','').replace(',',''))
                if v > 0: return 'background-color: #E2F0D9; color: #385723; font-weight: bold;'
                if v < 0: return 'background-color: #FCE4D6; color: #C00000;'
            except: pass
            return ''
            
        def color_flow(val):
            try:
                v = float(str(val).replace(' ','').replace('₽','').replace(',',''))
                if v >= 0: return 'background-color: #DDEBF7; color: #1F4E78; font-weight: bold;'
                if v < 0: return 'background-color: #FFF2CC; color: #7F6000;'
            except: pass
            return ''
            
        styled = styled.map(color_profit, subset=(df.index[profit_row_idx], df.columns[1:]))
        styled = styled.map(color_flow, subset=(df.index[flow_row_idx], df.columns[1:]))
        return styled
        
    st.dataframe(style_dataframe(df_sum), use_container_width=True, height=200)
    
    # Умный групповой заполнитель
    st.markdown("--- ")
    st.markdown("⚙️ **Умный групповой заполнитель показателей (Bulk Fill)**")
    col_fill_1, col_fill_2, col_fill_3, col_fill_4 = st.columns([1.5, 1, 1, 1])
    
    fillable_metrics = []
    for r in revenue_rows_def[:-1]: fillable_metrics.append(r[3])
    if show_detailed_fot:
        for r in fot_rows_def[:-1]: fillable_metrics.append(r[3])
    else:
        for r in fot_rows_def[:-1]: fillable_metrics.append(r[3])
    for r in opex_rows_def[:-1]: fillable_metrics.append(r[3])
    for r in hq_rows_def[:-1]: fillable_metrics.append(r[3])
    for r in capex_rows_def[:-1]: fillable_metrics.append(r[3])
    
    with col_fill_1:
        fill_metric = st.selectbox("Выберите статью:", options=fillable_metrics, key="fill_m")
    with col_fill_2:
        fill_start = st.number_input("С месяца:", min_value=1, max_value=24, value=1)
    with col_fill_3:
        fill_end = st.number_input("По месяц:", min_value=1, max_value=24, value=24)
    with col_fill_4:
        fill_value = st.number_input("Значение, ₽:", min_value=0.0, value=0.0, step=1000.0)
        
    if st.button("⚡ Заполнить ячейки интервала", use_container_width=True):
        for m_idx in range(fill_start, fill_end + 1):
            st.session_state['pl_overrides'][(fill_metric, str(m_idx))] = fill_value
        st.success(f"Успешно заполнена статья '{fill_metric}' во всех ячейках с {fill_start} по {fill_end} месяцы!")
        st.rerun()

    # Добавляем кнопку сброса ручных корректировок
    st.markdown(" ")
    col_reset_1, col_reset_2 = st.columns([3, 1])
    with col_reset_2:
        if st.button("🔄 Сбросить ручные корректировки P&L", use_container_width=True):
            st.session_state['pl_overrides'] = {}
            st.success("Все ручные изменения успешно сброшены!")
            st.rerun()

# ----------------------------------------------------
# TAB 3: ОПЕРАЦИОННАЯ ВОРОНКА (шт.)
# ----------------------------------------------------
with tab3:
    st.markdown("### 📈 Операционные показатели и Драйверы воронки (штук/конверсии)")
    st.markdown("ℹ️ *Вы можете скорректировать объем входящего потока лидов или количество закрываемых сделок любого типа за любой месяц. Модель мгновенно пересчитает доходы P&L.*")
    
    drivers_labels = [
        '  ├─ Звонки/лиды',
        '  ├─ Встречи',
        '  ├─ Договоры',
        '  ├─ Задаток/аванс',
        '  ├─ Сделки: вторичный рынок',
        '  ├─ Сделки: первичный рынок',
        '  ├─ Сделки: аренда',
        '  ├─ Сделки: загородная',
        '  ├─ Сделки: зарубежная',
        '  ├─ Сделки: прочее (МЛС, срочновыкуп, сайт)',
        '  ├─ Сервисы: ипотека',
        '  ├─ Сервисы: страхование',
        '  └─ Сервисы: юр. сопровождение'
    ]
    
    df_drivers_data = {'Категория / Статья': drivers_labels}
    df_drivers_data['Старт'] = [0.0]*len(drivers_labels)
    
    for idx, col in enumerate(months):
        df_drivers_data[col] = [
            get_driver_override('  ├─ Звонки/лиды', col, int(base['leads'][idx] * leads_mult)),
            get_driver_override('  ├─ Встречи', col, int(base['meetings'][idx] * leads_mult * conv_meet_rate)),
            get_driver_override('  ├─ Договоры', col, int(base['contracts'][idx] * overall_voeronka_factor)),
            get_driver_override('  ├─ Задаток/аванс', col, base['prepayments'][idx] * overall_voeronka_factor),
            get_driver_override('  ├─ Сделки: вторичный рынок', col, (base['prepayments'][idx] * overall_voeronka_factor) * 0.90),
            get_driver_override('  ├─ Сделки: первичный рынок', col, (base['prepayments'][idx] * overall_voeronka_factor) * 0.10),
            get_driver_override('  ├─ Сделки: аренда', col, base_deals_rent[idx]),
            get_driver_override('  ├─ Сделки: загородная', col, base_deals_suburban[idx]),
            get_driver_override('  ├─ Сделки: зарубежная', col, base_deals_overseas[idx]),
            get_driver_override('  ├─ Сделки: прочее (МЛС, срочновыкуп, сайт)', col, base_deals_other[idx]),
            get_driver_override('  ├─ Сервисы: ипотека', col, base_services_mortgage[idx]),
            get_driver_override('  ├─ Сервисы: страхование', col, base_services_insurance[idx]),
            get_driver_override('  └─ Сервисы: юр. сопровождение', col, base_services_legal[idx])
        ]
        
    df_drivers = pd.DataFrame(df_drivers_data)
    
    col_config_drv = {"Категория / Статья": st.column_config.TextColumn(disabled=True)}
    for col in df_drivers.columns[1:]:
        col_config_drv[col] = st.column_config.NumberColumn(format="%.1f", min_value=0.0)
        
    edited_drivers_df = st.data_editor(
        df_drivers,
        use_container_width=True,
        disabled=["Категория / Статья"],
        column_config=col_config_drv,
        key="drivers_editor",
        height=450
    )
    
    drivers_changed = False
    for idx_row in range(len(df_drivers)):
        row_name = df_drivers.at[idx_row, 'Категория / Статья']
        for col in df_drivers.columns[1:]:
            val_base = df_drivers.at[idx_row, col]
            val_edited = edited_drivers_df.at[idx_row, col]
            if float(val_base) != float(val_edited):
                st.session_state['driver_overrides'][(row_name, col)] = parse_value(val_edited)
                drivers_changed = True
                
    if drivers_changed:
        st.success("Драйверы воронки обновлены! Пересчет финансовой модели...")
        st.rerun()
        
    if st.button("🔄 Сбросить ручные корректировки воронки", use_container_width=True):
        st.session_state['driver_overrides'] = {}
        st.success("Все операционные корректировки сброшены!")
        st.rerun()

# ----------------------------------------------------
# TAB 4: КАДРОВОЕ ПЛАНИРОВАНИЕ (чел.)
# ----------------------------------------------------
with tab4:
    st.markdown("### 👥 Штатное расписание офиса (чел.)")
    st.markdown("ℹ️ *Здесь настраивается количество сотрудников в офисе. Увеличение или уменьшение численности персонала автоматически пересчитывает ФОТ в отчете P&L.*")
    
    staff_labels = [
        '  ├─ РОП: категория C',
        '  ├─ РОП: категория B',
        '  ├─ Администратор офиса',
        '  ├─ HR/рекрутер',
        '  ├─ РОС/тренер',
        '  ├─ Юрист',
        '  ├─ Ипотечный Брокер',
        '  ├─ Листинг-менеджер',
        '  ├─ Фотограф',
        '  ├─ Маркетолог/SMM',
        '  ├─ Агент: категория D',
        '  ├─ Агент: категория C',
        '  ├─ Агент: категория B',
        '  └─ Агент: категория A'
    ]
    
    df_staff_data = {'Штатная должность / Грейд': staff_labels}
    df_staff_data['Старт'] = [0]*len(staff_labels)
    
    for idx, col in enumerate(months):
        df_staff_data[col] = [
            get_headcount_override('  ├─ РОП: категория C', col, DEFAULT_HEADCOUNTS['РОП: категория C'][idx]),
            get_headcount_override('  ├─ РОП: категория B', col, DEFAULT_HEADCOUNTS['РОП: категория B'][idx]),
            get_headcount_override('  ├─ Администратор офиса', col, DEFAULT_HEADCOUNTS['Администратор офиса'][idx]),
            get_headcount_override('  ├─ HR/рекрутер', col, DEFAULT_HEADCOUNTS['HR/рекрутер'][idx]),
            get_headcount_override('  ├─ РОС/тренер', col, DEFAULT_HEADCOUNTS['РОС/тренер'][idx]),
            get_headcount_override('  ├─ Юрист', col, DEFAULT_HEADCOUNTS['Юрист'][idx]),
            get_headcount_override('  ├─ Ипотечный Брокер', col, DEFAULT_HEADCOUNTS['Ипотечный Брокер'][idx]),
            get_headcount_override('  ├─ Листинг-менеджер', col, DEFAULT_HEADCOUNTS['Листинг-менеджер'][idx]),
            get_headcount_override('  ├─ Фотограф', col, DEFAULT_HEADCOUNTS['Фотограф'][idx]),
            get_headcount_override('  ├─ Маркетолог/SMM', col, DEFAULT_HEADCOUNTS['Маркетолог/SMM'][idx]),
            get_headcount_override('  ├─ Агент: категория D', col, DEFAULT_HEADCOUNTS['Агент: категория D'][idx]),
            get_headcount_override('  ├─ Агент: категория C', col, DEFAULT_HEADCOUNTS['Агент: категория C'][idx]),
            get_headcount_override('  ├─ Агент: категория B', col, DEFAULT_HEADCOUNTS['Агент: категория B'][idx]),
            get_headcount_override('  └─ Агент: категория A', col, DEFAULT_HEADCOUNTS['Агент: категория A'][idx])
        ]
        
    df_staff = pd.DataFrame(df_staff_data)
    col_config_stf = {"Штатная должность / Грейд": st.column_config.TextColumn(disabled=True)}
    for col in df_staff.columns[1:]:
        col_config_stf[col] = st.column_config.NumberColumn(format="%d чел.", min_value=0)
        
    edited_staff_df = st.data_editor(
        df_staff,
        use_container_width=True,
        disabled=["Штатная должность / Грейд"],
        column_config=col_config_stf,
        key="staff_editor",
        height=480
    )
    
    staff_changed = False
    for idx_row in range(len(df_staff)):
        row_name = df_staff.at[idx_row, 'Штатная должность / Грейд']
        for col in df_staff.columns[1:]:
            val_base = df_staff.at[idx_row, col]
            val_edited = edited_staff_df.at[idx_row, col]
            if int(val_base) != int(val_edited):
                st.session_state['headcount_overrides'][(row_name, col)] = int(parse_value(val_edited))
                staff_changed = True
                
    if staff_changed:
        st.success("Кадровая структура обновлена! Пересчет ФОТ...")
        st.rerun()
        
    if st.button("🔄 Сбросить ручные корректировки штатного расписания", use_container_width=True):
        st.session_state['headcount_overrides'] = {}
        st.success("Все кадровые корректировки сброшены!")
        st.rerun()

# ----------------------------------------------------
# EXCEL EXPORT
# ----------------------------------------------------
if openpyxl_available:
    def generate_excel():
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "C21 Decomposed P&L"
        ws.views.sheetView[0].showGridLines = True
        
        gold_fill = PatternFill(start_color="BEAF87", end_color="BEAF87", fill_type="solid")
        dark_fill = PatternFill(start_color="252526", end_color="252526", fill_type="solid")
        light_gold_fill = PatternFill(start_color="F9F6EE", end_color="F9F6EE", fill_type="solid")
        category_fill = PatternFill(start_color="F1F1F3", end_color="F1F1F3", fill_type="solid")
        
        font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        
        thin_border_side = Side(border_style="thin", color="#D3D3D3")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        double_bottom_border = Border(bottom=Side(border_style="double", color="111111"), top=thin_border_side)
        
        ws.merge_cells("A1:Z1")
        ws["A1"] = f"ДЕТАЛИЗИРОВАННАЯ ФИНАНСОВАЯ МОДЕЛЬ CENTURY 21 - РЕГИОН: {region_select.upper()} (КЕЙС: БЮДЖЕТ ОПТИМАЛЬНЫЙ)"
        ws["A1"].font = font_title
        ws["A1"].fill = dark_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["Показатель / Период", "Старт"] + months
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = gold_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 28
        
        metrics_rows = [
            ('► ДОХОДЫ (ВКД)', [0], [""], ""),
            ('  ├─ Вторичный рынок', [0], rev_secondary_list, "#,##0"),
            ('  ├─ Первичный рынок', [0], rev_primary_list, "#,##0"),
            ('  ├─ Аренда (жилая/коммерческая)', [0], rev_rent_list, "#,##0"),
            ('  ├─ Загородная недвижимость', [0], rev_suburban_list, "#,##0"),
            ('  ├─ Зарубежная недвижимость', [0], rev_overseas_list, "#,##0"),
            ('  ├─ Сделки: прочее (МЛС, срочновыкуп)', [0], rev_other_p_list, "#,##0"),
            ('  ├─ Сервисы: ипотека', [0], rev_mortgage_list, "#,##0"),
            ('  ├─ Сервисы: страхование', [0], rev_insurance_list, "#,##0"),
            ('  └─ Сервисы: юр. сопровождение', [0], rev_legal_list, "#,##0"),
            ('★ ИТОГО ДОХОДЫ', [0], rev_total_list, "#,##0"),
            
            ('► РАСХОДЫ: ФОТ', [0], [""], ""),
            ('  ├─ Выплаты агентам (% комиссионных)', [0], payouts_agents_list, "#,##0"),
            ('  ├─ Оклады бэк-офиса (оклады)', [0], salaries_backoffice_list, "#,##0"),
            ('  └─ Налоги на ФОТ оклады', [0], taxes_payroll_list, "#,##0"),
            ('★ ИТОГО ФОТ', [0], total_payroll_list, "#,##0"),
            
            ('► РАСХОДЫ: OPEX', [0], [""], ""),
            ('  ├─ Аренда офиса', [office_rent_custom], rent_list, "#,##0"),
            ('  ├─ Интернет', [0], internet_list, "#,##0"),
            ('  ├─ Сотовая связь', [0], mobile_list, "#,##0"),
            ('  ├─ Канцелярия', [0], kanc_list, "#,##0"),
            ('  ├─ Реклама объектов', [0], reklama_list, "#,##0"),
            ('  ├─ HeadHunter.ru', [0], hh_list, "#,##0"),
            ('  ├─ Бухгалтерия: аутсорс', [0], buh_list, "#,##0"),
            ('  ├─ Услуги банка', [0], bank_list, "#,##0"),
            ('  ├─ Уборка офиса', [0], cleaning_list, "#,##0"),
            ('  ├─ ГСМ', [0], gsm_list, "#,##0"),
            ('  ├─ Доставка/курьер', [0], courier_list, "#,##0"),
            ('  └─ OPEX: Корпоративы', [0], events_list, "#,##0"),
            ('★ ИТОГО OPEX', [office_rent_custom], total_opex_list, "#,##0"),
            
            ('► ВЫПЛАТЫ ЦО (ФРАНШИЗА)', [0], [""], ""),
            ('  ├─ Роялти FIX (вкл. НРФ)', [0], hq_royalty_fix_list, "#,##0"),
            ('  ├─ Роялти со сделок', [0], hq_royalty_deal_list, "#,##0"),
            ('  ├─ CRM-система', [0], hq_crm_list, "#,##0"),
            ('  └─ Колл-центр и прочие сервисы', [0], hq_kc_list, "#,##0"),
            ('★ ИТОГО ВЫПЛАТЫ ЦО', [0], total_hq_payments_list, "#,##0"),
            
            ('★ НАЛОГ УСН', [0], taxes_usn_list, "#,##0"),
            ('★ ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ', [office_rent_custom], [total_payroll_list[idx]+total_opex_list[idx]+total_hq_payments_list[idx]+taxes_usn_list[idx] for idx in range(24)], "#,##0"),
            
            ('► КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)', [0], [""], ""),
            ('  ├─ Франшиза (Паушальный взнос + Роспатент)', [capex_m0_franchise], capex_franchise_list, "#,##0"),
            ('  ├─ Ремонт и Брендирование офиса', [capex_m0_renovation], capex_renovation_list, "#,##0"),
            ('  └─ Мебель, компьютеры и оборудование', [capex_m0_equipment], capex_equipment_list, "#,##0"),
            ('★ ИТОГО CAPEX', [total_capex_m0], total_capex_list, "#,##0"),
            
            ('★ ЧИСТАЯ ПРИБЫЛЬ ЗА МЕСЯЦ', [-total_capex_m0-office_rent_custom], adj_net_profit, "#,##0"),
            ('★ НАКОПЛЕННЫЙ ДЕНЕЖНЫЙ ПОТОК', [cum_flow[0]], cum_flow[1:], "#,##0")
        ]
        
        for r_idx, (m_name, m0_list, data_list, num_format) in enumerate(metrics_rows, 4):
            is_header = "►" in m_name
            is_total = "★" in m_name
            
            cell_a = ws.cell(row=r_idx, column=1, value=m_name)
            cell_a.font = font_bold if (is_header or is_total) else font_regular
            cell_a.border = thin_border
            
            if is_header:
                cell_a.fill = category_fill
            elif is_total:
                cell_a.fill = light_gold_fill
            
            cell_m0 = ws.cell(row=r_idx, column=2)
            if is_header:
                cell_m0.value = ""
                cell_m0.fill = category_fill
            else:
                cell_m0.value = m0_list[0]
                if num_format != "":
                    cell_m0.number_format = num_format
                cell_m0.alignment = Alignment(horizontal="right")
                if is_total:
                    cell_m0.font = font_bold
                    cell_m0.fill = light_gold_fill
                else:
                    cell_m0.font = font_regular
            cell_m0.border = thin_border
            
            for c_idx in range(1, 25):
                if is_header:
                    cell_val = ws.cell(row=r_idx, column=c_idx+2, value="")
                    cell_val.fill = category_fill
                    cell_val.border = thin_border
                    continue
                    
                val = data_list[c_idx-1] if len(data_list) >= c_idx else 0
                cell_val = ws.cell(row=r_idx, column=c_idx+2, value=val)
                if num_format != "":
                    cell_val.number_format = num_format
                cell_val.alignment = Alignment(horizontal="right")
                cell_val.border = thin_border
                
                if is_total:
                    cell_val.font = font_bold
                    cell_val.fill = light_gold_fill
                    if "ЧИСТАЯ ПРИБЫЛЬ" in m_name:
                        cell_val.border = double_bottom_border
                else:
                    cell_val.font = font_regular
            
            ws.row_dimensions[r_idx].height = 20
            
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 15
        for col in range(3, 27):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
            
        wb.save(output)
        return output.getvalue()
        
    excel_data = generate_excel()
    st.download_button(
        label="📥 Скачать детализированный P&L в Excel",
        data=excel_data,
        file_name=f"C21_P_and_L_Optimal_{region_select}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.warning("⚠️ **Экспорт в Excel временно недоступен.**")

with tab5:
    st.markdown("### 🌍 Сравнение ключевых показателей регионов")
    st.write("Сравнение базовых финансово-экономических параметров по региональным сегментам Century 21:")
    
    comp_data = {
        'Региональный сегмент': list(REGIONAL_PRESETS.keys()),
        'Паушальный взнос (база)': [f"{v['pau_fee']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Вторичка': [f"{v['comm_sec']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Первичка': [f"{v['comm_prim']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Чек: Загородная': [f"{v['comm_sub']:,} ₽".replace(",", " ") for v in REGIONAL_PRESETS.values()],
        'Оклады бэк-офиса': [f"{v['salary_mult']*100:.0f}% к Москве" for v in REGIONAL_PRESETS.values()],
        'Аренда офиса (база)': [f"{v['rent_base']:,} ₽/мес.".replace(",", " ") for v in REGIONAL_PRESETS.values()],
    }
    df_comp = pd.DataFrame(comp_data)
    st.table(df_comp)

with tab6:
    st.markdown("""
    ### 📖 Справочник и методология декомпозиции (Кейс: Бюджет Оптимальный)
    
    Эта интерактивная модель разработана специально для симуляции листа **«Бюджет Оптимальный»** финансового плана Century 21 за 24 месяца:
    
    1. **Доходы (Комиссия ВКД):** 
       * Рассчитывается динамически на основе воронки лидов. 
       * Месяцы 1–3 полностью освобождены от доходов в связи со стартовым лагом набора базы и закрытия первых сделок (0 ₽).
       * Включает доп. сервисы — **комиссионные вознаграждения от банков по ипотеке, страховых компаний и юр. услуг**.
       
    2. **ФОТ (Зарплатный фонд):**
       * **Переменная часть (агенты):** Привязана к объему сделок и настраивается слайдером (базово 38%).
       * **Постоянная часть (оклады):** Включает фиксированные оклады РОПа, HR/рекрутера, юриста, ипотечного брокера, листинг-менеджера и администратора, скорректированные под выбранный регион.
       
    3. **OPEX (Операционные расходы):**
       * Разделен на 12 детализированных статей, включая затраты на интернет, сотовую связь, рекламу объектов, подбор персонала, IT-лицензии и уборку.
       
    4. **Выплаты ЦО (Франшиза):**
       * Фиксированные роялти (шкала по месяцам), переменные роялти с каждой сделки (настраиваются слайдером) и IT-сервисы ЦО.
       
    5. **CAPEX (Капитальные вложения):**
       * Рассчитывается как агрессивный старт в Месяце 0: Паушальный взнос + Роспатент + Полная закупка оборудования и ремонт офиса (**1 335 000 ₽**).
    """)
